"""
Serwer FastAPI dla Systemu Zarządzania Produkcją v4.1
Deploy na Railway.app
"""
from fastapi import FastAPI, HTTPException, Header, Depends, Request, UploadFile, File, Body, Form
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, FileResponse, StreamingResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Any, Optional
import sqlite3, os, json, hashlib, time, io, datetime as _dt, traceback, urllib.request, urllib.error
import openpyxl as _openpyxl
from contextlib import contextmanager

def _now():
    """Bieżący czas UTC jako ISO string z 'Z' (JS-kompatybilny)."""
    return _dt.datetime.utcnow().isoformat() + "Z"

def _parse(s):
    """Parsuje ISO datetime string (obsługuje 'Z', Python <3.11)."""
    if not s: return _dt.datetime.utcnow()
    return _dt.datetime.fromisoformat(str(s).replace("Z", "").replace("+00:00", ""))

# ─── Konfiguracja ──────────────────────────────────────────────────────────────
DB_PATH       = os.environ.get("DB_PATH",      "/data/produkcja.db")
API_KEY       = os.environ.get("API_KEY",      "zmien-mnie-na-bezpieczny-klucz")
PORT          = int(os.environ.get("PORT",     8000))
BACKUP_PATH   = os.environ.get("BACKUP_PATH",  "/data/backup.json")   # persystentny backup JSON
BACKUP_INTERVAL = int(os.environ.get("BACKUP_INTERVAL", 120))        # co ile sekund auto-backup (domyślnie 2 min)

# ── Cloudinary – storage plików STEP ──────────────────────────────────────────
CLOUDINARY_CLOUD = os.environ.get("CLOUDINARY_CLOUD_NAME", "")
CLOUDINARY_KEY   = os.environ.get("CLOUDINARY_API_KEY", "")
CLOUDINARY_SECRET= os.environ.get("CLOUDINARY_API_SECRET", "")

# ── GitHub Gist backup (zalecany na Render bez dysku) ─────────────────────────
GIST_TOKEN   = os.environ.get("GIST_TOKEN", "")   # Personal Access Token (scope: gist)
GIST_ID      = os.environ.get("GIST_ID", "")      # ID Gista – opcjonalnie, wykrywane automatycznie
GIST_ID_FILE = os.path.join(os.path.dirname(__file__), ".gist_id")  # cache ID w kontenerze

# Jeżeli /data nie istnieje (Render bez dysku), używamy katalogu lokalnego
if not os.path.exists(os.path.dirname(DB_PATH)):
    _local = os.path.dirname(__file__)
    DB_PATH     = os.path.join(_local, "produkcja.db")
    BACKUP_PATH = os.path.join(_local, "backup.json")

app = FastAPI(title="Produkcja API", version="4.1")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    tb = traceback.format_exc()
    print(f"[500] {request.method} {request.url.path}\n{tb}", flush=True)
    return JSONResponse(
        status_code=500,
        content={"detail": f"Błąd serwera [{type(exc).__name__}]: {str(exc)}"}
    )

# ─── Auth ──────────────────────────────────────────────────────────────────────
def verify_key(x_api_key: str = Header(..., alias="x-api-key")):
    if x_api_key != API_KEY:
        raise HTTPException(status_code=401, detail="Nieprawidłowy klucz API")
    return True

# ─── Database helpers ──────────────────────────────────────────────────────────
@contextmanager
def get_db():
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

def _hash(p): return hashlib.sha256(p.encode()).hexdigest()

# ── Helpers: stawki per zlecenie ─────────────────────────────────────────────
def _init_stawki_zlecenia(conn, zlecenie_id: int) -> int:
    """Kopiuje globalne stawki dla stanowisk użytych w operacjach zlecenia."""
    ops = conn.execute(
        "SELECT DISTINCT stanowisko FROM operacje WHERE zlecenie_id=? AND stanowisko IS NOT NULL AND stanowisko!=''",
        (zlecenie_id,)
    ).fetchall()
    if not ops:
        return 0

    global_stawki = {
        r["stanowisko"]: dict(r)
        for r in conn.execute("SELECT stanowisko, stawka_godz, zbrojenie_stawka_godz FROM stawki").fetchall()
    }

    added = 0
    for op in ops:
        st = op["stanowisko"]
        if not st:
            continue
        existing = conn.execute(
            "SELECT id FROM stawki_zlecen WHERE zlecenie_id=? AND stanowisko=?",
            (zlecenie_id, st)
        ).fetchone()
        if existing:
            continue
        g = global_stawki.get(st, {})
        conn.execute(
            """INSERT INTO stawki_zlecen (zlecenie_id, stanowisko, stawka_godz, zbrojenie_stawka_godz)
               VALUES (?, ?, ?, ?)""",
            (zlecenie_id, st, g.get("stawka_godz", 0) or 0, g.get("zbrojenie_stawka_godz", 0) or 0)
        )
        added += 1
    return added

# ─── SQL Proxy ────────────────────────────────────────────────────────────────
class SQLRequest(BaseModel):
    sql: str
    params: List[Any] = []
    write: bool = False
    many: bool = False
    params_list: List[List[Any]] = []

class TransactionRequest(BaseModel):
    operations: List[dict]

@app.post("/sql", dependencies=[Depends(verify_key)])
def execute_sql(req: SQLRequest):
    try:
        conn = sqlite3.connect(DB_PATH, timeout=30)
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA foreign_keys=ON")
        cur = conn.cursor()
        if req.many and req.params_list:
            cur.executemany(req.sql, req.params_list)
            conn.commit()
            conn.close()
            return {"rowcount": cur.rowcount, "lastrowid": cur.lastrowid}
        cur.execute(req.sql, req.params)
        is_read = req.sql.strip().upper().startswith(("SELECT", "PRAGMA", "WITH"))
        if is_read:
            rows = cur.fetchall()
            conn.close()
            return {"rows": [list(r) for r in rows]}
        else:
            conn.commit()
            lastrowid = cur.lastrowid
            rowcount  = cur.rowcount
            conn.close()
            return {"lastrowid": lastrowid, "rowcount": rowcount}
    except sqlite3.Error as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/transaction", dependencies=[Depends(verify_key)])
def execute_transaction(req: TransactionRequest):
    try:
        conn = sqlite3.connect(DB_PATH, timeout=30)
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA foreign_keys=ON")
        results = []
        for op in req.operations:
            cur = conn.cursor()
            cur.execute(op["sql"], op.get("params", []))
            results.append({"lastrowid": cur.lastrowid, "rowcount": cur.rowcount})
        conn.commit()
        conn.close()
        return {"results": results}
    except sqlite3.Error as e:
        raise HTTPException(status_code=500, detail=str(e))

# ─── REST API ─────────────────────────────────────────────────────────────────
# Auth
class LoginRequest(BaseModel):
    username: str
    password: str

@app.post("/api/login")
def login(req: LoginRequest):
    with get_db() as conn:
        row = conn.execute(
            "SELECT id, username, full_name, role FROM users WHERE username=? AND password=?",
            (req.username, _hash(req.password))
        ).fetchone()
        if not row:
            raise HTTPException(401, "Nieprawidłowy login lub hasło")
        return {"id": row[0], "username": row[1], "full_name": row[2], "role": row[3]}

@app.get("/api/status/produkcja", dependencies=[Depends(verify_key)])
def status_produkcja():
    """Status produkcji na ekran logowania – zakończone/aktywne/następne operacje."""
    with get_db() as conn:
        zakonczone = conn.execute("""
            SELECT o.id, o.nazwa, o.kolejnosc, o.stanowisko, o.ilosc_wykonana,
                   z.ilosc_sztuk, z.numer as zl_numer, z.nazwa as zl_nazwa,
                   o.czas_norma, z.id as zlecenie_id
            FROM operacje o JOIN zlecenia z ON o.zlecenie_id=z.id
            WHERE o.status='zakonczona' AND z.status IN ('nowe','w_toku')
            ORDER BY z.numer, o.kolejnosc
        """).fetchall()
        aktywne = conn.execute("""
            SELECT DISTINCT o.id, o.nazwa, o.kolejnosc, o.stanowisko, o.ilosc_wykonana,
                            z.ilosc_sztuk, z.numer as zl_numer, z.nazwa as zl_nazwa, o.czas_norma
            FROM sesje_pracy s
            JOIN operacje o ON s.operacja_id=o.id
            JOIN zlecenia z ON o.zlecenie_id=z.id
            WHERE s.status='aktywna' AND s.typ IN ('operacja','inne_zlecenie')
            ORDER BY z.numer, o.kolejnosc
        """).fetchall()
        nastepne = conn.execute("""
            SELECT o.id, o.nazwa, o.kolejnosc, o.stanowisko, o.ilosc_wykonana,
                   z.ilosc_sztuk, z.numer as zl_numer, z.nazwa as zl_nazwa, o.czas_norma
            FROM operacje o JOIN zlecenia z ON o.zlecenie_id=z.id
            WHERE o.status='oczekuje' AND z.status IN ('nowe','w_toku')
            ORDER BY z.numer, o.kolejnosc LIMIT 15
        """).fetchall()

        next_map = {}
        for op in zakonczone:
            nxt = conn.execute("""
                SELECT nazwa, kolejnosc, stanowisko, czas_norma FROM operacje
                WHERE zlecenie_id=? AND kolejnosc>? AND status!='anulowane'
                ORDER BY kolejnosc LIMIT 1
            """, (op["zlecenie_id"], op["kolejnosc"])).fetchone()
            if nxt:
                next_map[str(op["id"])] = dict(nxt)

        return {
            "zakonczone": [dict(r) for r in zakonczone],
            "aktywne":    [dict(r) for r in aktywne],
            "nastepne":   [dict(r) for r in nastepne],
            "next_map":   next_map,
        }

# ─── QR Code scan ─────────────────────────────────────────────────────────────
@app.get("/api/scan/{qr}", dependencies=[Depends(verify_key)])
def scan_qr(qr: str):
    """Szuka operacji lub zlecenia po QR kodzie"""
    with get_db() as conn:
        op = conn.execute("""
            SELECT o.*, z.numer as zl_numer, z.nazwa as zl_nazwa, z.ilosc_sztuk
            FROM operacje o
            JOIN zlecenia z ON o.zlecenie_id = z.id
            WHERE o.qr_code=?
        """, (qr,)).fetchone()
        if op:
            od = dict(op)
            zbr_done = conn.execute(
                "SELECT COUNT(*) FROM sesje_pracy WHERE operacja_id=? AND typ='zbrojenie' AND status='zakonczona'",
                (op["id"],)
            ).fetchone()[0]
            od["zbrojenie_wykonane"] = bool(zbr_done)
            return {"type": "operacja", "data": od}

        zl = conn.execute("SELECT * FROM zlecenia WHERE qr_code=?", (qr,)).fetchone()
        if zl:
            ops = conn.execute(
                "SELECT * FROM operacje WHERE zlecenie_id=? ORDER BY kolejnosc", (zl["id"],)
            ).fetchall()
            ops_list = []
            for o in ops:
                od = dict(o)
                zbr_done = conn.execute(
                    "SELECT COUNT(*) FROM sesje_pracy WHERE operacja_id=? AND typ='zbrojenie' AND status='zakonczona'",
                    (o["id"],)
                ).fetchone()[0]
                od["zbrojenie_wykonane"] = bool(zbr_done)
                ops_list.append(od)
            return {"type": "zlecenie", "data": dict(zl), "operacje": ops_list}

        raise HTTPException(404, "Nie znaleziono kodu QR: " + qr)

# ─── Zlecenia CRUD ────────────────────────────────────────────────────────────
@app.get("/api/zlecenia", dependencies=[Depends(verify_key)])
def get_zlecenia(status: Optional[str] = None):
    with get_db() as conn:
        if status:
            rows = conn.execute(
                "SELECT * FROM zlecenia WHERE status=? ORDER BY created_at DESC", (status,)
            ).fetchall()
        else:
            rows = conn.execute("SELECT * FROM zlecenia ORDER BY created_at DESC").fetchall()
        result = []
        for z in rows:
            zd = dict(z)
            ops = conn.execute("""
                SELECT id, status, czas_norma, ilosc_wykonana, czas_zbrojenia_min FROM operacje WHERE zlecenie_id=?
            """, (z["id"],)).fetchall()
            ilosc = z["ilosc_sztuk"] or 1
            pozostale_min = 0
            for op in ops:
                if op["status"] == "zakonczona": continue
                norma = op["czas_norma"] or 0
                wykonano = op["ilosc_wykonana"] or 0
                zbrojenie = op["czas_zbrojenia_min"] or 0
                if norma:
                    pozostale_min += norma * max(0, ilosc - wykonano)
                if zbrojenie:
                    zbr_done = conn.execute(
                        "SELECT COUNT(*) FROM sesje_pracy WHERE operacja_id=? AND typ='zbrojenie' AND status='zakonczona'",
                        (op["id"],)
                    ).fetchone()[0]
                    if not zbr_done:
                        pozostale_min += zbrojenie
            zd["pozostale_min"] = round(pozostale_min, 1)
            result.append(zd)
        return result

class ZlecenieRequest(BaseModel):
    numer: str
    nazwa: str
    opis: Optional[str] = ""
    status: Optional[str] = "nowe"
    termin: Optional[str] = None
    ilosc_sztuk: Optional[int] = 1
    cena_brutto_szt: Optional[float] = 0
    material_od_klienta: Optional[int] = 0
    model_3d_url: Optional[str] = None

@app.post("/api/zlecenia", dependencies=[Depends(verify_key)])
def create_zlecenie(req: ZlecenieRequest):
    import uuid
    qr = "ZL-" + str(uuid.uuid4())[:8].upper()
    with get_db() as conn:
        try:
            cur = conn.execute(
                """INSERT INTO zlecenia (numer,nazwa,opis,status,termin,ilosc_sztuk,
                cena_brutto_szt,material_od_klienta,qr_code,model_3d_url)
                VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (req.numer, req.nazwa, req.opis, req.status, req.termin,
                 req.ilosc_sztuk, req.cena_brutto_szt, req.material_od_klienta, qr,
                 req.model_3d_url)
            )
            return {"id": cur.lastrowid, "qr_code": qr}
        except sqlite3.IntegrityError:
            raise HTTPException(400, "Zlecenie o tym numerze już istnieje")
        except Exception as e:
            raise HTTPException(500, f"Błąd bazy danych: {str(e)}")

@app.put("/api/zlecenia/{zid}", dependencies=[Depends(verify_key)])
def update_zlecenie(zid: int, req: ZlecenieRequest):
    with get_db() as conn:
        conn.execute(
            """UPDATE zlecenia SET numer=?,nazwa=?,opis=?,status=?,termin=?,
               ilosc_sztuk=?,cena_brutto_szt=?,material_od_klienta=?,model_3d_url=? WHERE id=?""",
            (req.numer, req.nazwa, req.opis, req.status, req.termin,
             req.ilosc_sztuk, req.cena_brutto_szt, req.material_od_klienta,
             req.model_3d_url, zid)
        )
        return {"ok": True}

@app.delete("/api/zlecenia/{zid}", dependencies=[Depends(verify_key)])
def delete_zlecenie(zid: int):
    with get_db() as conn:
        conn.execute("DELETE FROM sesje_pracy WHERE operacja_id IN (SELECT id FROM operacje WHERE zlecenie_id=?)", (zid,))
        conn.execute("DELETE FROM produkty_zlecenia WHERE zlecenie_id=?", (zid,))
        conn.execute("DELETE FROM stawki_zlecen WHERE zlecenie_id=?", (zid,))
        conn.execute("DELETE FROM operacje WHERE zlecenie_id=?", (zid,))
        conn.execute("DELETE FROM zlecenia WHERE id=?", (zid,))
        return {"ok": True}

@app.patch("/api/zlecenia/{zid}/status", dependencies=[Depends(verify_key)])
def change_zlecenie_status(zid: int, body: dict):
    status = body.get("status")
    if status not in ("nowe", "w_toku", "zakonczone", "anulowane", "oczekuje_potwierdzenia", "wstrzymane"):
        raise HTTPException(400, "Nieprawidłowy status")
    with get_db() as conn:
        conn.execute("UPDATE zlecenia SET status=? WHERE id=?", (status, zid))
        return {"ok": True}

# ─── Operacje CRUD ────────────────────────────────────────────────────────────
@app.get("/api/zlecenia/{zid}/operacje", dependencies=[Depends(verify_key)])
def get_operacje(zid: int):
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM operacje WHERE zlecenie_id=? ORDER BY kolejnosc", (zid,)
        ).fetchall()
        return [dict(r) for r in rows]

class OperacjaRequest(BaseModel):
    zlecenie_id: int
    nazwa: str
    kolejnosc: Optional[int] = 0
    czas_norma: Optional[float] = 0
    stanowisko: Optional[str] = ""
    opis_czynnosci: Optional[str] = ""
    czas_zbrojenia_min: Optional[float] = 0.0

@app.post("/api/operacje", dependencies=[Depends(verify_key)])
def create_operacja(req: OperacjaRequest):
    import uuid
    qr = "OP-" + str(uuid.uuid4())[:8].upper()
    with get_db() as conn:
        cur = conn.execute(
            """INSERT INTO operacje (zlecenie_id,nazwa,kolejnosc,czas_norma,
            stanowisko,opis_czynnosci,qr_code,czas_zbrojenia_min)
            VALUES (?,?,?,?,?,?,?,?)""",
            (req.zlecenie_id, req.nazwa, req.kolejnosc, req.czas_norma,
             req.stanowisko, req.opis_czynnosci, qr, req.czas_zbrojenia_min or 0.0)
        )
        # ➤ Inicjalizuj stawkę zlecenia dla tego stanowiska (jeśli istnieje)
        if req.stanowisko:
            _init_stawki_zlecenia(conn, req.zlecenie_id)
        return {"id": cur.lastrowid, "qr_code": qr}

@app.put("/api/operacje/{oid}", dependencies=[Depends(verify_key)])
def update_operacja(oid: int, req: OperacjaRequest):
    with get_db() as conn:
        op_before = conn.execute("SELECT zlecenie_id, stanowisko FROM operacje WHERE id=?", (oid,)).fetchone()
        conn.execute(
            """UPDATE operacje SET nazwa=?,kolejnosc=?,czas_norma=?,
               stanowisko=?,opis_czynnosci=?,czas_zbrojenia_min=? WHERE id=?""",
            (req.nazwa, req.kolejnosc, req.czas_norma,
             req.stanowisko, req.opis_czynnosci, req.czas_zbrojenia_min or 0.0, oid)
        )
        # Przy zmianie stanowiska – dodaj nowe do stawek_zlecen
        if req.stanowisko and op_before:
            if req.stanowisko != op_before["stanowisko"]:
                _init_stawki_zlecenia(conn, op_before["zlecenie_id"])
        return {"ok": True}

@app.delete("/api/operacje/{oid}", dependencies=[Depends(verify_key)])
def delete_operacja(oid: int):
    with get_db() as conn:
        conn.execute("DELETE FROM operacje WHERE id=?", (oid,))
        return {"ok": True}

# Endpoint KJ
class KJRequest(BaseModel):
    wynik: str  # 'zgodny' | 'niezgodny'
    uwagi: Optional[str] = ""

@app.patch("/api/operacje/{oid}/kj", dependencies=[Depends(verify_key)])
def zapisz_wynik_kj(oid: int, req: KJRequest):
    if req.wynik not in ("zgodny", "niezgodny"):
        raise HTTPException(400, "Wynik musi być: zgodny lub niezgodny")
    with get_db() as conn:
        op = conn.execute("SELECT * FROM operacje WHERE id=?", (oid,)).fetchone()
        if not op:
            raise HTTPException(404, "Operacja nie istnieje")
        nowy_status = "zakonczona" if req.wynik == "zgodny" else "niezgodna_kj"
        conn.execute(
            "UPDATE operacje SET kj_wynik=?, status=?, opis_czynnosci=CASE WHEN ?!='' THEN opis_czynnosci||'\n[KJ '||datetime('now')||']: '||? ELSE opis_czynnosci END WHERE id=?",
            (req.wynik, nowy_status, req.uwagi, req.uwagi, oid)
        )
        if req.wynik == "niezgodny":
            conn.execute(
                "UPDATE zlecenia SET status='wstrzymane' WHERE id=(SELECT zlecenie_id FROM operacje WHERE id=?)",
                (oid,)
            )
        _threading.Thread(target=_db_backup_to_json, daemon=True).start()
        return {"ok": True, "wynik": req.wynik}

def _detect_steel_grade(opis: str) -> str:
    """Rozpoznaje gatunek stali/materiału z opisu pozycji BOM."""
    import re
    opis_up = opis.upper()
    # Profile stalowe (stal konstrukcyjna S235/S355)
    if re.search(r'\bUPE\d+|\bIPE\d+|\bHEB\d+|\bHEA\d+|\bIPN\d+|\bUPN\d+|\bCEOWNIK\b|\bDWUTE\b|\bKĄTOWNIK\b|\bPROFIL\b', opis_up):
        if re.search(r'S355|S 355', opis_up):
            return 'S355'
        return 'S235'  # domyślny dla profili
    # Gatunek wprost w opisie
    for grade in ['S355MC','S355J2','S355JR','S355','S235JR','S235','S275','S420','S460',
                  'ST52','ST37','1.4301','1.4307','304','316L','P265GH','P355GH']:
        if grade.upper() in opis_up:
            return grade
    # Pręty kwadratowe / okrągłe / płaskowniki / blachy → domyślnie S235
    if re.search(r'\bPRĘT\b|\bPŁASKOWNIK\b|\bBLACHA\b|\bRURA\b', opis_up):
        return 'S235'
    return 'S235'  # fallback

def _calc_mass_kg(opis: str, ilosc: float) -> tuple:
    """
    Oblicza masę w kg na podstawie opisu i ilości sztuk.
    Zwraca (masa_kg: float, wymiary_str: str, gestosc: float).
    Obsługuje: Blacha AxBxC, Pręt kw. AxAxL, Ceownik UPExL, Płaskownik AxBxL.
    Uwaga: używa re.IGNORECASE zamiast .upper(), bo .upper() zamienia 'x'→'X'
    co psuje regex. Polskie znaki obsługiwane przez '.' w nazwie.
    """
    import re, math
    gestosc = 7.85  # g/cm³ → 7850 kg/m³; wzór: mm³/1e6 * 7850 = kg

    # Ceownik UPExxx L (np. "Ceownik UPE400x4390") – masa liniowa wg EN 10279
    UPE_MASY = {80:8.13,100:10.6,120:13.4,140:16.0,160:18.8,180:22.4,
                200:26.2,220:29.4,240:33.2,270:36.1,300:46.5,360:57.0,400:65.5}
    m = re.search(r'UPE\s*(\d+)[xX](\d+)', opis, re.IGNORECASE)
    if m:
        h, L_mm = int(m.group(1)), float(m.group(2))
        ml = UPE_MASY.get(h, h * 0.165)
        return round(ml * L_mm / 1000 * ilosc, 2), f"UPE{h} L={L_mm:.0f} mm", gestosc

    # IPE / HEB / HEA – masy liniowe [kg/m]
    PROFILE_MASY = {
        'IPE': {80:6.0,100:8.1,120:10.4,140:12.9,160:15.8,180:18.8,200:22.4,
                220:26.2,240:30.7,270:36.1,300:42.2,330:49.1,360:57.1,400:66.3},
        'HEB': {100:20.4,120:26.7,140:33.7,160:42.6,180:51.2,200:61.3,220:71.5,
                240:83.2,260:93.0,280:103.0,300:117.0,320:127.0,340:134.0,360:142.0},
        'HEA': {100:16.7,120:19.9,140:24.7,160:30.4,180:35.5,200:42.3,220:50.5,
                240:60.3,260:68.2,280:76.4,300:88.3,320:97.6,340:105.0,360:112.0},
    }
    for prof, masy in PROFILE_MASY.items():
        m = re.search(prof + r'\s*(\d+)[xX](\d+)', opis, re.IGNORECASE)
        if m:
            h, L_mm = int(m.group(1)), float(m.group(2))
            ml = masy.get(h, h * 0.18)
            return round(ml * L_mm / 1000 * ilosc, 2), f"{prof}{h} L={L_mm:.0f} mm", gestosc

    # Płaskownik AxBxL (np. "Płaskownik 60x30x3500") – '.' pasuje do 'ł'
    m = re.search(r'P.askownik\s+(\d+)[xX](\d+)[xX](\d+)', opis, re.IGNORECASE)
    if m:
        a, b, L = float(m.group(1)), float(m.group(2)), float(m.group(3))
        return round(a * b * L * 7.85e-6 * ilosc, 2), f"{a:.0f}×{b:.0f}×{L:.0f} mm", gestosc

    # Blacha t x B x L (np. "Blacha 40x405x4390")
    m = re.search(r'Blacha\s+(\d+)[xX](\d+)[xX](\d+)', opis, re.IGNORECASE)
    if m:
        t, b, L = float(m.group(1)), float(m.group(2)), float(m.group(3))
        return round(t * b * L * 7.85e-6 * ilosc, 2), f"{t:.0f}×{b:.0f}×{L:.0f} mm", gestosc

    # Pręt kwadratowy AxAxL (np. "Pręt kw. 15x15x200") – '.' pasuje do 'ę'
    m = re.search(r'Pr.t\s+kw[a-z.]*\s+(\d+)[xX](\d+)[xX](\d+)', opis, re.IGNORECASE)
    if m:
        a, b, L = float(m.group(1)), float(m.group(2)), float(m.group(3))
        return round(a * b * L * 7.85e-6 * ilosc, 2), f"□{a:.0f}×{L:.0f} mm", gestosc

    # Pręt okrągły fi D x L (np. "Pręt fi 20x1000" lub "Pręt Ø20x1000")
    m = re.search(r'Pr.t\s+(?:fi|f|Ø|O)?\s*(\d+(?:[.,]\d+)?)[xX](\d+)', opis, re.IGNORECASE)
    if m:
        d, L = float(m.group(1).replace(',','.')), float(m.group(2))
        return round(math.pi * (d/2)**2 * L * 7.85e-6 * ilosc, 2), f"⌀{d}×{L:.0f} mm", gestosc

    return 0.0, "", gestosc


def _parse_bom_from_pdf_text(text: str) -> list:
    """Wydobywa wykaz materialow z karty technologicznej PDF.

    pdfminer czyta kolumny PDF rozdzielnie - ilosci i JM trafiaja
    PRZED naglowkiem 'Oznaczenie', a opisy/kody PO nim.
    Format bloku ilosci: "Ilosc jedn.\n\n  1,00\n\nJM\n\nszt\n\n  9,00\n\nszt\n\n..."
    Format bloku opisow: "78111 P19518\n\nCeownik UPE400...\n\n78112 P19519\n\n..."
    Zwraca pozycje wzbogacone o: masa_kg, gatunek_stali, wymiary_str.
    """
    import re

    bom_start = re.search(r'Oznaczenie\s+Kod\s+Indeks', text, re.IGNORECASE | re.DOTALL)
    if not bom_start:
        return []

    # Ilosci sa w kolumnie prawej, wyciagane przez pdfminer PRZED 'Oznaczenie'
    pre_text = text[:bom_start.start()]
    ilosc_blok = re.search(
        r'Ilo.{1,4}\s+jedn\.\s*\n([\s\S]+?)(?=RYSUNKI:|Nr\s+oper\.)',
        pre_text, re.IGNORECASE
    )
    ilosci = []
    jm_list = []
    if ilosc_blok:
        blok = ilosc_blok.group(1)
        blok = re.sub(r'\bJM\b', '', blok)  # usun naglowek kolumny JM
        tokens = re.findall(r'(\d+[,\.]\d+)|([a-zA-Z]{1,6})', blok)
        i = 0
        while i < len(tokens):
            num, word = tokens[i]
            if num:
                ilosc = float(num.replace(',', '.'))
                jm = 'szt'
                if i + 1 < len(tokens) and tokens[i + 1][1]:
                    jm = tokens[i + 1][1].strip()
                    i += 1
                ilosci.append(ilosc)
                jm_list.append(jm)
            i += 1

    # Wiersze materialow po naglowku: "78111 P19518\n\nCeownik ... Poz. 1"
    bom_text = text[bom_start.end():]
    bom_end = re.search(r'\n\s*\d{3}\s*\n\s*OT-', bom_text)
    if bom_end:
        bom_text = bom_text[:bom_end.start()]

    row_pat = re.compile(
        r'^\s*(\d{4,7})\s+(P\d+)\s*\n+\s*(.+?)\s*$',
        re.MULTILINE
    )
    materialy = []
    for idx, m in enumerate(row_pat.finditer(bom_text)):
        oznaczenie = m.group(1).strip()
        kod        = m.group(2).strip()
        opis       = m.group(3).strip()
        opis = re.sub(r'\s+Poz\.\s*\d+\s*$', '', opis, flags=re.IGNORECASE).strip()
        ilosc = ilosci[idx] if idx < len(ilosci) else 1.0
        jm    = jm_list[idx] if idx < len(jm_list) else 'szt'
        masa_kg, wymiary_str, _ = _calc_mass_kg(opis, ilosc)
        gatunek = _detect_steel_grade(opis)
        materialy.append({
            "oznaczenie": oznaczenie, "kod": kod, "opis": opis,
            "ilosc": ilosc, "jm": jm,
            "masa_kg": masa_kg, "wymiary_str": wymiary_str, "gatunek_stali": gatunek,
        })

    # Fallback: format jednoliniowy (inne extractory)
    if not materialy:
        row_pat2 = re.compile(
            r'(\d{4,7})\s+(P\d+)\s+(.+?)\s+(\d+[,\.]\d+)\s+([a-zA-Z]{1,6})\s*$',
            re.MULTILINE
        )
        for m in row_pat2.finditer(bom_text):
            opis = m.group(3).strip()
            opis = re.sub(r'\s+Poz\.\s*\d+\s*$', '', opis, flags=re.IGNORECASE).strip()
            try:
                ilosc = float(m.group(4).replace(',', '.'))
            except Exception:
                ilosc = 1.0
            masa_kg, wymiary_str, _ = _calc_mass_kg(opis, ilosc)
            gatunek = _detect_steel_grade(opis)
            materialy.append({
                "oznaczenie": m.group(1).strip(),
                "kod":        m.group(2).strip(),
                "opis":       opis,
                "ilosc":      ilosc,
                "jm":         m.group(5).strip(),
                "masa_kg":    masa_kg, "wymiary_str": wymiary_str, "gatunek_stali": gatunek,
            })

    return materialy


# ─── Import technologii z PDF ─────────────────────────────────────────────────
def _parse_technologia_pdf(pdf_bytes: bytes) -> dict:
    """Parsuje kartę technologiczną PDF → dict z numerem, nazwą i operacjami."""
    import re, io
    try:
        from pdfminer.high_level import extract_text as _extract
        text = _extract(io.BytesIO(pdf_bytes))
    except Exception as e:
        raise ValueError(f"Nie można odczytać PDF: {e}")

    hdr = re.search(r'WYRÓB / DETAL:.*?\n(\S+)\s*\n(.+?)\n', text, re.DOTALL)
    numer = hdr.group(1).strip() if hdr else ""
    nazwa = hdr.group(2).strip() if hdr else ""
    if not numer:
        m2 = re.search(r'\n(P\d+)\n', text)
        numer = m2.group(1) if m2 else "IMPORT"
    if not nazwa:
        m3 = re.search(numer + r'\s*\n(.+?)\n', text)
        nazwa = m3.group(1).strip() if m3 else "Importowana technologia"

    op_pattern = re.compile(
        r'(\d{3})\s*\n\s*(OT-[A-Z0-9]+-\d+\s*-\s*[^\n]+?)\s*\n'
        r'\s*([\d\s,\.]+)\s*\n\s*([\d\s,\.]+)\s*\n\s*([^\n]+?)\s*\n'
        r'\s*Uwagi:\s*\n\s*(.*?)(?=\n\d{3}\s*\n|Suma Tj:|\Z)',
        re.DOTALL
    )

    operacje = []
    for m in op_pattern.finditer(text):
        nr_str   = m.group(1)
        kod_nazwa = m.group(2).strip()
        tj_str   = m.group(3).strip().replace(' ','').replace(',','.')
        tpz_str  = m.group(4).strip().replace(' ','').replace(',','.')
        stanowisko_raw = m.group(5).strip()
        opis_raw = m.group(6).strip()

        try: tj = float(re.search(r'[\d\.]+', tj_str).group())
        except: tj = 0.0
        try: tpz = float(re.search(r'[\d\.]+', tpz_str).group())
        except: tpz = 0.0

        st_match = re.match(r'^[A-Z0-9]+ - (.+)$', stanowisko_raw)
        stanowisko = st_match.group(1).strip() if st_match else stanowisko_raw

        kod_match = re.match(r'(OT-[A-Z]+)-\d+', kod_nazwa)
        kod_prefix = kod_match.group(1) if kod_match else "OT"

        if 'KJ' in kod_prefix or 'KJ' in stanowisko_raw:
            typ_op = 'kj'
        elif 'KOOP' in kod_prefix or 'KOOP' in stanowisko_raw:
            typ_op = 'kooperacja'
        elif 'ZP' in kod_prefix or 'Zbrojenie' in stanowisko or 'zbrojenie' in opis_raw.lower():
            typ_op = 'zbrojenie_zewn'
        else:
            typ_op = 'produkcja'

        nazwa_op_match = re.match(r'OT-[A-Z0-9]+-\d+\s*-\s*(.+)', kod_nazwa)
        nazwa_op = nazwa_op_match.group(1).strip() if nazwa_op_match else kod_nazwa

        kj_params = []
        for line in opis_raw.splitlines():
            if 'Niezgodny' in line and 'Zgodny' in line:
                kj_params.append(line.strip())

        opis_clean = re.sub(r'Parametry KJ:.*', '', opis_raw, flags=re.DOTALL).strip()

        operacje.append({
            "kolejnosc": int(nr_str),
            "nazwa": nazwa_op,
            "stanowisko_raw": stanowisko_raw,
            "stanowisko": stanowisko,
            "czas_norma": tj,
            "czas_tpz_min": tpz,
            "opis_czynnosci": opis_clean,
            "typ_operacji": typ_op,
            "parametry_kj": json.dumps(kj_params, ensure_ascii=False) if kj_params else None,
        })

    if not any(o["kolejnosc"] == 10 for o in operacje):
        m010 = re.search(r'010\s*\n\s*(OT-[^\n]+)\s*\n.*?([\d,]+)\s*\n\s*([\d,]+)\s*\n\s*([^\n]+?)\s*\n', text, re.DOTALL)
        if m010:
            nazwa_010 = re.sub(r'OT-[A-Z0-9]+-\d+\s*-\s*','', m010.group(1).strip())
            operacje.insert(0, {
                "kolejnosc": 10, "nazwa": nazwa_010,
                "stanowisko_raw": m010.group(4).strip(),
                "stanowisko": re.sub(r'^[A-Z0-9]+ - ','', m010.group(4).strip()),
                "czas_norma": 0.0, "czas_tpz_min": 0.0,
                "opis_czynnosci": "Kontrola jakości materiału",
                "typ_operacji": "kj",
                "parametry_kj": json.dumps(["Materiał wejściowy: Niezgodny - Zgodny"], ensure_ascii=False),
            })

    operacje.sort(key=lambda x: x["kolejnosc"])
    return {"numer": numer, "nazwa": nazwa, "operacje": operacje}

class ImportTechnologiaResponse(BaseModel):
    zlecenie_id: int
    numer: str
    nazwa: str
    operacje_created: int
    nowe_stanowiska: List[str]
    bom_added: int
    bom_new_materialy: int
    errors: List[str]


@app.post("/api/import-technologia/parse", dependencies=[Depends(verify_key)])
async def parse_technologia_pdf(file: UploadFile = File(...)):
    """Parsuje PDF bez zapisu do bazy. Zwraca podgląd operacji i materiałów BOM."""
    pdf_bytes = await file.read()
    if not pdf_bytes:
        raise HTTPException(400, "Pusty plik")
    try:
        from pdfminer.high_level import extract_text as _extract
        import io as _io
        text = _extract(_io.BytesIO(pdf_bytes))
        parsed = _parse_technologia_pdf(pdf_bytes)
    except ValueError as e:
        raise HTTPException(400, str(e))

    bom_raw = _parse_bom_from_pdf_text(text)

    # Sprawdź które materiały istnieją w bazie (po Kod = indeks)
    with get_db() as conn:
        for mat in bom_raw:
            # Szukaj po kodzie P-xxxxx jako indeks, albo po fragmencie opisu
            row = conn.execute(
                "SELECT id, indeks, opis, jm, do_dyspozycji, stan_rzeczywisty FROM materialy WHERE indeks=? LIMIT 1",
                (mat["kod"],)
            ).fetchone()
            if not row:
                # Próba dopasowania po fragmencie opisu (pierwsze 2 słowa)
                words = mat["opis"].split()[:2]
                like_q = "%" + " ".join(words) + "%"
                row = conn.execute(
                    "SELECT id, indeks, opis, jm, do_dyspozycji, stan_rzeczywisty FROM materialy WHERE opis LIKE ? LIMIT 1",
                    (like_q,)
                ).fetchone()
            if row:
                mat["w_bazie"] = True
                mat["material_id"] = row["id"]
                mat["indeks_bazy"] = row["indeks"]
                mat["opis_bazy"] = row["opis"]
                mat["do_dyspozycji"] = row["do_dyspozycji"] or 0
            else:
                mat["w_bazie"] = False
                mat["material_id"] = None
                mat["indeks_bazy"] = mat["kod"]
                mat["opis_bazy"] = mat["opis"]
                mat["do_dyspozycji"] = 0

    return {
        "numer": parsed["numer"],
        "nazwa": parsed["nazwa"],
        "operacje": parsed["operacje"],
        "bom": bom_raw,
    }


@app.post("/api/import-technologia", dependencies=[Depends(verify_key)])
async def import_technologia(file: UploadFile = File(...), force: bool = False, bom_json: str = Form(default="")):

    pdf_bytes = await file.read()
    if not pdf_bytes:
        raise HTTPException(400, "Pusty plik")
    try:
        parsed = _parse_technologia_pdf(pdf_bytes)
    except ValueError as e:
        raise HTTPException(400, str(e))

    errors = []
    nowe_stanowiska = []

    with get_db() as conn:
        base_numer = parsed["numer"]
        numer_do_uzycia = base_numer
        # Jeśli numer już istnieje, dodaj sufiks :2, :3 itd.
        suffix = 2
        while conn.execute("SELECT id FROM zlecenia WHERE numer=?", (numer_do_uzycia,)).fetchone():
            numer_do_uzycia = f"{base_numer}:{suffix}"
            suffix += 1
        parsed["numer"] = numer_do_uzycia

        import uuid as _uuid
        qr_zl = "ZL-" + str(_uuid.uuid4())[:8].upper()
        cur = conn.execute(
            """INSERT INTO zlecenia (numer, nazwa, opis, status, ilosc_sztuk, qr_code)
               VALUES (?, ?, ?, 'oczekuje_potwierdzenia', 1, ?)""",
            (parsed["numer"], parsed["nazwa"],
             f"Zaimportowano z karty technologicznej {_now()[:10]}", qr_zl)
        )
        zlecenie_id = cur.lastrowid

        import re as _re
        # Kazde zbrojenie_zewn staje sie osobna operacja z typ_operacji='zbrojenie'
        # zachowujac swoja kolejnosc z PDF – pojawia sie przed wlasciwa operacja produkcja
        operacje_finalne = []
        for op in parsed["operacje"]:
            if op["typ_operacji"] == "zbrojenie_zewn":
                op["typ_operacji"] = "zbrojenie"
                czas_zbr = op.get("czas_tpz_min") or op.get("czas_norma") or 0.0
                op["czas_zbrojenia_min"] = czas_zbr
                op["czas_norma"] = 0.0
                op["czas_tpz_min"] = 0.0
                operacje_finalne.append(op)
            else:
                operacje_finalne.append(op)

        operacje_finalne.sort(key=lambda x: x["kolejnosc"])

        istniejace_stanowiska = {
            r["stanowisko"] for r in conn.execute("SELECT stanowisko FROM stawki").fetchall()
        }
        for op in operacje_finalne:
            st = op["stanowisko"]
            if st and st not in istniejace_stanowiska:
                try:
                    conn.execute(
                        "INSERT INTO stawki (stanowisko, stawka_godz, opis) VALUES (?, 0.0, ?)",
                        (st, f"Dodano automatycznie podczas importu {_now()[:10]}")
                    )
                    istniejace_stanowiska.add(st)
                    nowe_stanowiska.append(st)
                except Exception as e:
                    errors.append(f"Nie można dodać stanowiska {st}: {e}")

        for op in operacje_finalne:
            if (op.get("czas_zbrojenia_min") or 0.0) > 0:
                try:
                    conn.execute(
                        "UPDATE stawki SET zbrojenie_aktywne=1 WHERE stanowisko=? AND zbrojenie_aktywne=0",
                        (op["stanowisko"],)
                    )
                except Exception:
                    pass

        op_count = 0
        for op in operacje_finalne:
            try:
                qr_op = "OP-" + str(_uuid.uuid4())[:8].upper()
                czas_norma = op["czas_norma"] if op["typ_operacji"] != "zbrojenie_zewn" else 0.0
                czas_zbrojenia = op.get("czas_zbrojenia_min") or 0.0
                conn.execute(
                    """INSERT INTO operacje
                       (zlecenie_id, nazwa, kolejnosc, czas_norma, stanowisko,
                        opis_czynnosci, qr_code, czas_zbrojenia_min,
                        typ_operacji, parametry_kj, czas_tpz_min)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                    (zlecenie_id, op["nazwa"], op["kolejnosc"],
                     czas_norma, op["stanowisko"], op["opis_czynnosci"],
                     qr_op, czas_zbrojenia,
                     op["typ_operacji"], op["parametry_kj"], op.get("czas_tpz_min", 0.0))
                )
                op_count += 1
            except Exception as e:
                errors.append(f"Operacja {op['kolejnosc']} {op['nazwa']}: {e}")

        # ➤ NOWE: Zainicjalizuj indywidualne stawki zlecenia na podstawie stanowisk z operacji
        _init_stawki_zlecenia(conn, zlecenie_id)

        # ➤ BOM: Przetwórz potwierdzone materiały
        bom_added = 0
        bom_new_materialy = 0
        if bom_json:
            try:
                bom_items = json.loads(bom_json)
                for item in bom_items:
                    if not item.get("included", True):
                        continue
                    indeks = item.get("indeks_bazy") or item.get("kod") or ""
                    opis   = item.get("opis_bazy") or item.get("opis") or ""
                    jm     = item.get("jm", "szt")
                    ilosc  = float(item.get("ilosc", 1))
                    mat_id = item.get("material_id")

                    if not mat_id:
                        # Utwórz nowy materiał w bazie z zerowymi stanami
                        cur_m = conn.execute(
                            """INSERT INTO materialy (kod, indeks, opis, jm, do_dyspozycji, stan_rzeczywisty, updated_at)
                               VALUES (?,?,?,?, 0, 0, ?)
                               ON CONFLICT(indeks) DO UPDATE SET opis=excluded.opis, jm=excluded.jm, updated_at=excluded.updated_at
                               RETURNING id""",
                            (item.get("kod",""), indeks, opis, jm, _now())
                        ).fetchone()
                        mat_id = cur_m["id"] if cur_m else conn.execute(
                            "SELECT id FROM materialy WHERE indeks=?", (indeks,)
                        ).fetchone()["id"]
                        bom_new_materialy += 1

                    # Dodaj do BOM (ignoruj duplikaty)
                    try:
                        _masa_kg = float(item.get("masa_kg") or 0)
                        _gatunek = item.get("gatunek_stali") or _detect_steel_grade(opis)
                        _wymiary = item.get("wymiary_str") or ""
                        if not _masa_kg:
                            _masa_kg, _wymiary, _ = _calc_mass_kg(opis, ilosc)
                        conn.execute(
                            "INSERT OR IGNORE INTO bom_pozycje (zlecenie_id, material_id, ilosc, uwagi, masa_kg, gatunek_stali, wymiary_str, created_at) VALUES (?,?,?,?,?,?,?,?)",
                            (zlecenie_id, mat_id, ilosc, item.get("uwagi",""), _masa_kg, _gatunek, _wymiary, _now())
                        )
                        bom_added += 1
                    except Exception as e:
                        errors.append(f"BOM {opis}: {e}")
            except Exception as e:
                errors.append(f"Błąd przetwarzania BOM: {e}")

        _threading.Thread(target=_db_backup_to_json, daemon=True).start()
        return {
            "zlecenie_id": zlecenie_id,
            "numer": parsed["numer"],
            "nazwa": parsed["nazwa"],
            "operacje_created": op_count,
            "nowe_stanowiska": nowe_stanowiska,
            "bom_added": bom_added,
            "bom_new_materialy": bom_new_materialy,
            "errors": errors,
        }

@app.get("/api/operacje/aktywne", dependencies=[Depends(verify_key)])
def get_aktywne_operacje():
    with get_db() as conn:
        rows = conn.execute("""
            SELECT o.*, z.numer as zl_numer, z.nazwa as zl_nazwa, z.ilosc_sztuk
            FROM operacje o
            JOIN zlecenia z ON o.zlecenie_id = z.id
            WHERE o.status IN ('oczekuje','w_toku')
              AND z.status IN ('nowe','w_toku')
            ORDER BY z.numer, o.kolejnosc
        """).fetchall()
        return [dict(r) for r in rows]

@app.get("/api/operacje/zakonczone-do-transportu", dependencies=[Depends(verify_key)])
def get_zakonczone_transport():
    with get_db() as conn:
        rows = conn.execute("""
            SELECT o.*, z.numer as zl_numer, z.nazwa as zl_nazwa, z.ilosc_sztuk
            FROM operacje o
            JOIN zlecenia z ON o.zlecenie_id = z.id
            WHERE o.status = 'zakonczona'
              AND z.status IN ('nowe','w_toku')
            ORDER BY z.numer, o.kolejnosc
        """).fetchall()
        return [dict(r) for r in rows]

# ─── Sesje pracy ──────────────────────────────────────────────────────────────
class StartSesjaRequest(BaseModel):
    operacja_id: Optional[int] = None
    user_id: int
    typ: str = "operacja"
    opis_nieprodukcyjnej: Optional[str] = ""
    zlecenie_id_inne: Optional[int] = None
    sesja_glowna: int = 1
    rzeczywiste_stanowisko: Optional[str] = None  # gdy operacja wykonana na innej maszynie

class StopSesjaRequest(BaseModel):
    sesja_id: int
    ilosc_sztuk: int
    uwagi: Optional[str] = ""

class PauzaRequest(BaseModel):
    sesja_id: int
    powod: Optional[str] = ""

@app.get("/api/sesje/aktywne_operacja/{operacja_id}", dependencies=[Depends(verify_key)])
def get_aktywne_sesje_operacja(operacja_id: int):
    with get_db() as conn:
        rows = conn.execute("""
            SELECT s.id, s.user_id, s.typ, s.start_time, s.sesja_glowna, s.pauzy,
                   u.full_name
            FROM sesje_pracy s
            JOIN users u ON s.user_id = u.id
            WHERE s.operacja_id=? AND s.status='aktywna'
            ORDER BY s.start_time
        """, (operacja_id,)).fetchall()
        return [dict(r) for r in rows]

@app.post("/api/sesje/start", dependencies=[Depends(verify_key)])
def start_sesja(req: StartSesjaRequest):
    now = _now()
    with get_db() as conn:
        if req.typ in ("operacja", "inne_zlecenie", "zbrojenie") and req.operacja_id:
            aktywne = conn.execute(
                """SELECT s.id, s.typ, s.user_id, s.sesja_glowna, u.full_name
                   FROM sesje_pracy s JOIN users u ON s.user_id=u.id
                   WHERE s.operacja_id=? AND s.status='aktywna'""",
                (req.operacja_id,)
            ).fetchall()

            zbrojenie_akt = [r for r in aktywne if r["typ"] == "zbrojenie"]
            operacja_akt  = [r for r in aktywne if r["typ"] in ("operacja", "inne_zlecenie")]

            if req.typ == "zbrojenie":
                if zbrojenie_akt:
                    raise HTTPException(400, "Zbrojenie tej operacji jest już aktywne.")
                if operacja_akt:
                    raise HTTPException(400, "Operacja jest już w toku – nie można uruchomić zbrojenia.")
            elif req.typ in ("operacja", "inne_zlecenie"):
                if zbrojenie_akt:
                    raise HTTPException(400, "Trwa zbrojenie tej operacji – najpierw je zakończ.")
                if req.sesja_glowna == 1:
                    glowna_akt = [r for r in operacja_akt if r["sesja_glowna"] == 1]
                    if glowna_akt:
                        raise HTTPException(400,
                            f"GLOWNA_ZAJETA:{glowna_akt[0]['full_name']}:{glowna_akt[0]['id']}")
                moja = [r for r in operacja_akt if r["user_id"] == req.user_id]
                if moja:
                    raise HTTPException(400, "Masz już aktywną sesję tej operacji.")

        cur = conn.execute(
            """INSERT INTO sesje_pracy
               (operacja_id, user_id, typ, start_time, status, uwagi, zlecenie_id_inne, sesja_glowna, rzeczywiste_stanowisko)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (req.operacja_id, req.user_id, req.typ, now, "aktywna",
             req.opis_nieprodukcyjnej or "",
             req.zlecenie_id_inne if req.typ == "inne_zlecenie" else None,
             req.sesja_glowna,
             req.rzeczywiste_stanowisko or None)
        )
        sesja_id = cur.lastrowid
        if req.operacja_id and req.typ in ("operacja", "inne_zlecenie"):
            conn.execute(
                "UPDATE operacje SET status='w_toku' WHERE id=? AND status='oczekuje'",
                (req.operacja_id,)
            )
        _threading.Thread(target=_db_backup_to_json, daemon=True).start()
        return {"sesja_id": sesja_id, "start_time": now}

@app.post("/api/sesje/pauza/start", dependencies=[Depends(verify_key)])
def pauza_start(req: PauzaRequest):
    now = _now()
    with get_db() as conn:
        sesja = conn.execute("SELECT * FROM sesje_pracy WHERE id=?", (req.sesja_id,)).fetchone()
        if not sesja:
            raise HTTPException(404, "Sesja nie znaleziona")
        pauzy = json.loads(sesja["pauzy"] or "[]")
        if pauzy and pauzy[-1].get("koniec") is None:
            raise HTTPException(400, "Pauza już aktywna")
        pauzy.append({"start": now, "koniec": None, "powod": req.powod or ""})
        conn.execute("UPDATE sesje_pracy SET pauzy=? WHERE id=?",
                     (json.dumps(pauzy), req.sesja_id))
        _threading.Thread(target=_db_backup_to_json, daemon=True).start()
        return {"ok": True, "pauza_start": now}

@app.post("/api/sesje/pauza/stop", dependencies=[Depends(verify_key)])
def pauza_stop(req: PauzaRequest):
    now = _now()
    with get_db() as conn:
        sesja = conn.execute("SELECT * FROM sesje_pracy WHERE id=?", (req.sesja_id,)).fetchone()
        if not sesja:
            raise HTTPException(404, "Sesja nie znaleziona")
        pauzy = json.loads(sesja["pauzy"] or "[]")
        if not pauzy or pauzy[-1].get("koniec") is not None:
            raise HTTPException(400, "Brak aktywnej pauzy")
        pauzy[-1]["koniec"] = now
        conn.execute("UPDATE sesje_pracy SET pauzy=? WHERE id=?",
                     (json.dumps(pauzy), req.sesja_id))
        _threading.Thread(target=_db_backup_to_json, daemon=True).start()
        return {"ok": True, "pauza_koniec": now}

@app.post("/api/sesje/stop", dependencies=[Depends(verify_key)])
def stop_sesja(req: StopSesjaRequest):
    now = _now()
    with get_db() as conn:
        sesja = conn.execute("SELECT * FROM sesje_pracy WHERE id=?", (req.sesja_id,)).fetchone()
        if not sesja:
            raise HTTPException(404, "Sesja nie znaleziona")

        pauzy = json.loads(sesja["pauzy"] or "[]")
        if pauzy and pauzy[-1].get("koniec") is None:
            pauzy[-1]["koniec"] = now

        conn.execute(
            "UPDATE sesje_pracy SET end_time=?, ilosc_sztuk=?, uwagi=?, status='zakonczona', pauzy=? WHERE id=?",
            (now, req.ilosc_sztuk, req.uwagi, json.dumps(pauzy), req.sesja_id)
        )
        if sesja["operacja_id"] and sesja["typ"] in ("operacja", "inne_zlecenie"):
            conn.execute(
                "UPDATE operacje SET ilosc_wykonana = ilosc_wykonana + ? WHERE id=?",
                (req.ilosc_sztuk, sesja["operacja_id"])
            )
            op = conn.execute(
                "SELECT o.ilosc_wykonana, z.ilosc_sztuk FROM operacje o JOIN zlecenia z ON o.zlecenie_id=z.id WHERE o.id=?",
                (sesja["operacja_id"],)
            ).fetchone()
            if op and op[0] >= op[1]:
                conn.execute("UPDATE operacje SET status='zakonczona' WHERE id=?", (sesja["operacja_id"],))
                zl_id = conn.execute("SELECT zlecenie_id FROM operacje WHERE id=?", (sesja["operacja_id"],)).fetchone()[0]
                pozostale = conn.execute(
                    "SELECT COUNT(*) FROM operacje WHERE zlecenie_id=? AND status NOT IN ('zakonczona','anulowane')",
                    (zl_id,)
                ).fetchone()[0]
                if pozostale == 0:
                    conn.execute(
                        "UPDATE zlecenia SET status='zakonczone' WHERE id=? AND status NOT IN ('zakonczone','anulowane')",
                        (zl_id,)
                    )

        _threading.Thread(target=_db_backup_to_json, daemon=True).start()
        return {"status": "ok", "end_time": now}

class EditSesjaTimeRequest(BaseModel):
    start_time: Optional[str] = None
    end_time: Optional[str] = None

@app.patch("/api/sesje/{sesja_id}/czas", dependencies=[Depends(verify_key)])
def edit_sesja_czas(sesja_id: int, req: EditSesjaTimeRequest):
    with get_db() as conn:
        sesja = conn.execute("SELECT * FROM sesje_pracy WHERE id=?", (sesja_id,)).fetchone()
        if not sesja:
            raise HTTPException(404, "Sesja nie istnieje")
        new_start = req.start_time or sesja["start_time"]
        new_end   = req.end_time   or sesja["end_time"]
        if new_end:
            try:
                dt_start = _parse(new_start)
                dt_end   = _parse(new_end)
                if dt_end <= dt_start:
                    raise HTTPException(400, "Czas zakończenia musi być późniejszy niż rozpoczęcia")
            except HTTPException:
                raise
            except Exception:
                raise HTTPException(400, "Nieprawidłowy format czasu")
        conn.execute(
            "UPDATE sesje_pracy SET start_time=?, end_time=? WHERE id=?",
            (new_start, new_end, sesja_id)
        )
        return {"status": "ok", "sesja_id": sesja_id, "start_time": new_start, "end_time": new_end}

@app.delete("/api/sesje/{sesja_id}", dependencies=[Depends(verify_key)])
def delete_sesja(sesja_id: int):
    with get_db() as conn:
        sesja = conn.execute("SELECT * FROM sesje_pracy WHERE id=?", (sesja_id,)).fetchone()
        if not sesja:
            raise HTTPException(404, "Sesja nie istnieje")
        if sesja["status"] == "aktywna":
            raise HTTPException(400, "Nie można usunąć aktywnej sesji – najpierw ją zakończ")
        conn.execute("DELETE FROM sesje_pracy WHERE id=?", (sesja_id,))
        return {"status": "ok", "deleted": sesja_id}

@app.get("/api/sesje/aktywne/{user_id}", dependencies=[Depends(verify_key)])
def get_aktywne_sesje(user_id: int):
    with get_db() as conn:
        rows = conn.execute("""
            SELECT s.*, o.nazwa as op_nazwa, o.stanowisko,
                   COALESCE(z.numer, zi.numer) as zl_numer,
                   COALESCE(z.nazwa, zi.nazwa) as zl_nazwa,
                   COALESCE(z.id, zi.id) as zlecenie_id
            FROM sesje_pracy s
            LEFT JOIN operacje o ON s.operacja_id = o.id
            LEFT JOIN zlecenia z ON o.zlecenie_id = z.id
            LEFT JOIN zlecenia zi ON s.zlecenie_id_inne = zi.id
            WHERE s.user_id=? AND s.status='aktywna'
            ORDER BY s.start_time
        """, (user_id,)).fetchall()
        return [dict(r) for r in rows]

# backward compat
@app.get("/api/sesje/aktywna/{user_id}", dependencies=[Depends(verify_key)])
def get_aktywna_sesja(user_id: int):
    with get_db() as conn:
        row = conn.execute("""
            SELECT s.*, o.nazwa as op_nazwa, o.stanowisko,
                   z.numer as zl_numer, z.nazwa as zl_nazwa
            FROM sesje_pracy s
            LEFT JOIN operacje o ON s.operacja_id = o.id
            LEFT JOIN zlecenia z ON o.zlecenie_id = z.id
            WHERE s.user_id=? AND s.status='aktywna'
            ORDER BY s.start_time LIMIT 1
        """, (user_id,)).fetchone()
        return dict(row) if row else None

@app.get("/api/sesje/historia/{user_id}", dependencies=[Depends(verify_key)])
def get_historia(user_id: int, limit: int = 100):
    with get_db() as conn:
        rows = conn.execute("""
            SELECT s.*, o.nazwa as op_nazwa, o.stanowisko,
                   z.numer as zl_numer, z.nazwa as zl_nazwa
            FROM sesje_pracy s
            LEFT JOIN operacje o ON s.operacja_id = o.id
            LEFT JOIN zlecenia z ON o.zlecenie_id = z.id
            WHERE s.user_id=? AND s.status='zakonczona'
            ORDER BY s.end_time DESC LIMIT ?
        """, (user_id, limit)).fetchall()
        return [dict(r) for r in rows]

# ─── Stawki CRUD (globalne) ──────────────────────────────────────────────────
@app.get("/api/stawki", dependencies=[Depends(verify_key)])
def get_stawki():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM stawki ORDER BY stanowisko").fetchall()
        return [dict(r) for r in rows]

class StawkaRequest(BaseModel):
    stanowisko: str
    stawka_godz: float
    opis: Optional[str] = ""
    zbrojenie_aktywne: Optional[int] = 0
    zbrojenie_stawka_godz: Optional[float] = 0.0
    typ_maszyny: Optional[str] = ""

@app.post("/api/stawki", dependencies=[Depends(verify_key)])
def create_stawka(req: StawkaRequest):
    with get_db() as conn:
        try:
            cur = conn.execute(
                "INSERT INTO stawki (stanowisko,stawka_godz,opis,zbrojenie_aktywne,zbrojenie_stawka_godz,typ_maszyny) VALUES (?,?,?,?,?,?)",
                (req.stanowisko, req.stawka_godz, req.opis, req.zbrojenie_aktywne or 0, req.zbrojenie_stawka_godz or 0.0, req.typ_maszyny or '')
            )
            return {"id": cur.lastrowid}
        except sqlite3.IntegrityError:
            raise HTTPException(400, "Stanowisko już istnieje")

@app.put("/api/stawki/{sid}", dependencies=[Depends(verify_key)])
def update_stawka(sid: int, req: StawkaRequest):
    with get_db() as conn:
        conn.execute(
            "UPDATE stawki SET stanowisko=?,stawka_godz=?,opis=?,zbrojenie_aktywne=?,zbrojenie_stawka_godz=?,typ_maszyny=? WHERE id=?",
            (req.stanowisko, req.stawka_godz, req.opis, req.zbrojenie_aktywne or 0, req.zbrojenie_stawka_godz or 0.0, req.typ_maszyny or '', sid)
        )
        # Celowo NIE aktualizujemy stawki_zlecen – istniejące zlecenia mają zachować indywidualne stawki
        return {"ok": True}

@app.delete("/api/stawki/{sid}", dependencies=[Depends(verify_key)])
def delete_stawka(sid: int):
    with get_db() as conn:
        conn.execute("DELETE FROM stawki WHERE id=?", (sid,))
        return {"ok": True}

# ─── Stawki per zlecenie ─────────────────────────────────────────────────────
@app.get("/api/zlecenia/{zid}/stawki", dependencies=[Depends(verify_key)])
def get_stawki_zlecenia(zid: int):
    """Lista stanowisk ze stawkami dla konkretnego zlecenia."""
    with get_db() as conn:
        zl = conn.execute("SELECT id FROM zlecenia WHERE id=?", (zid,)).fetchone()
        if not zl:
            raise HTTPException(404, "Zlecenie nie istnieje")

        _init_stawki_zlecenia(conn, zid)

        rows = conn.execute("""
            SELECT sz.id, sz.zlecenie_id, sz.stanowisko, sz.stawka_godz, sz.zbrojenie_stawka_godz,
                   g.stawka_godz as global_stawka,
                   g.zbrojenie_stawka_godz as global_zbrojenie,
                   (CASE
                       WHEN g.stawka_godz IS NULL THEN 1
                       WHEN sz.stawka_godz != g.stawka_godz THEN 1
                       ELSE 0 END) as modified_stawka,
                   (CASE
                       WHEN COALESCE(g.zbrojenie_stawka_godz,0) = 0 AND COALESCE(sz.zbrojenie_stawka_godz,0) > 0 THEN 1
                       WHEN sz.zbrojenie_stawka_godz != COALESCE(g.zbrojenie_stawka_godz,0) THEN 1
                       ELSE 0 END) as modified_zbrojenie
            FROM stawki_zlecen sz
            LEFT JOIN stawki g ON sz.stanowisko = g.stanowisko
            WHERE sz.zlecenie_id=?
            ORDER BY sz.stanowisko
        """, (zid,)).fetchall()
        return [dict(r) for r in rows]


class StawkaZleceniaRequest(BaseModel):
    stawka_godz: float
    zbrojenie_stawka_godz: Optional[float] = 0.0


@app.put("/api/zlecenia/{zid}/stawki/{sid}", dependencies=[Depends(verify_key)])
def update_stawka_zlecenia(zid: int, sid: int, req: StawkaZleceniaRequest):
    """Edycja stawki dla konkretnego zlecenia."""
    with get_db() as conn:
        conn.execute(
            """UPDATE stawki_zlecen
               SET stawka_godz=?, zbrojenie_stawka_godz=?
               WHERE id=? AND zlecenie_id=?""",
            (req.stawka_godz, req.zbrojenie_stawka_godz or 0.0, sid, zid)
        )
        _threading.Thread(target=_db_backup_to_json, daemon=True).start()
        return {"ok": True}


@app.post("/api/zlecenia/{zid}/stawki/sync", dependencies=[Depends(verify_key)])
def sync_stawki_zlecenia(zid: int, force: bool = False):
    """Synchronizacja stawek zlecenia z globalnymi.
    force=false: dodaje tylko brakujące stanowiska
    force=true:  nadpisuje wszystkie stawki globalnymi"""
    with get_db() as conn:
        zl = conn.execute("SELECT id FROM zlecenia WHERE id=?", (zid,)).fetchone()
        if not zl:
            raise HTTPException(404, "Zlecenie nie istnieje")

        ops = conn.execute(
            "SELECT DISTINCT stanowisko FROM operacje WHERE zlecenie_id=? AND stanowisko IS NOT NULL AND stanowisko!=''",
            (zid,)
        ).fetchall()
        global_stawki = {
            r["stanowisko"]: dict(r)
            for r in conn.execute("SELECT stanowisko, stawka_godz, zbrojenie_stawka_godz FROM stawki").fetchall()
        }

        updated = 0
        for op in ops:
            st = op["stanowisko"]
            g = global_stawki.get(st, {})
            existing = conn.execute(
                "SELECT id FROM stawki_zlecen WHERE zlecenie_id=? AND stanowisko=?",
                (zid, st)
            ).fetchone()
            if existing:
                if force:
                    conn.execute(
                        """UPDATE stawki_zlecen
                           SET stawka_godz=?, zbrojenie_stawka_godz=?
                           WHERE zlecenie_id=? AND stanowisko=?""",
                        (g.get("stawka_godz", 0) or 0, g.get("zbrojenie_stawka_godz", 0) or 0, zid, st)
                    )
                    updated += 1
            else:
                conn.execute(
                    """INSERT INTO stawki_zlecen (zlecenie_id, stanowisko, stawka_godz, zbrojenie_stawka_godz)
                       VALUES (?, ?, ?, ?)""",
                    (zid, st, g.get("stawka_godz", 0) or 0, g.get("zbrojenie_stawka_godz", 0) or 0)
                )
                updated += 1

        # Usuń stanowiska które już nie są używane w operacjach tego zlecenia
        used = [op["stanowisko"] for op in ops]
        if used:
            conn.execute(
                f"DELETE FROM stawki_zlecen WHERE zlecenie_id=? AND stanowisko NOT IN ({','.join('?'*len(used))})",
                [zid] + used
            )
        else:
            conn.execute("DELETE FROM stawki_zlecen WHERE zlecenie_id=?", (zid,))

        _threading.Thread(target=_db_backup_to_json, daemon=True).start()
        return {"ok": True, "updated": updated, "force": force}


@app.delete("/api/zlecenia/{zid}/stawki/{sid}", dependencies=[Depends(verify_key)])
def delete_stawka_zlecenia(zid: int, sid: int):
    with get_db() as conn:
        conn.execute("DELETE FROM stawki_zlecen WHERE id=? AND zlecenie_id=?", (sid, zid))
        return {"ok": True}


# ─── Użytkownicy CRUD ─────────────────────────────────────────────────────────
@app.get("/api/users", dependencies=[Depends(verify_key)])
def get_users():
    with get_db() as conn:
        rows = conn.execute(
            "SELECT id, username, full_name, role FROM users ORDER BY full_name"
        ).fetchall()
        return [dict(r) for r in rows]

class NewUserRequest(BaseModel):
    username: str
    password: str
    full_name: str
    role: str

@app.post("/api/users", dependencies=[Depends(verify_key)])
def create_user(req: NewUserRequest):
    with get_db() as conn:
        try:
            cur = conn.execute(
                "INSERT INTO users (username, password, full_name, role) VALUES (?,?,?,?)",
                (req.username, _hash(req.password), req.full_name, req.role)
            )
            return {"id": cur.lastrowid}
        except sqlite3.IntegrityError:
            raise HTTPException(400, "Użytkownik już istnieje")

class EditUserRequest(BaseModel):
    full_name: str
    role: str
    password: Optional[str] = None

@app.put("/api/users/{uid}", dependencies=[Depends(verify_key)])
def update_user(uid: int, req: EditUserRequest):
    with get_db() as conn:
        if req.password:
            conn.execute(
                "UPDATE users SET full_name=?,role=?,password=? WHERE id=?",
                (req.full_name, req.role, _hash(req.password), uid)
            )
        else:
            conn.execute(
                "UPDATE users SET full_name=?,role=? WHERE id=?",
                (req.full_name, req.role, uid)
            )
        return {"ok": True}

@app.post("/api/users/{uid}/change-password", dependencies=[Depends(verify_key)])
def change_password(uid: int, req: dict = Body(...)):
    old_pass = req.get("old_password", "")
    new_pass = req.get("new_password", "")
    if not new_pass or len(new_pass) < 4:
        raise HTTPException(400, "Nowe hasło musi mieć co najmniej 4 znaki")
    with get_db() as conn:
        user = conn.execute(
            "SELECT id FROM users WHERE id=? AND password=?",
            (uid, _hash(old_pass))
        ).fetchone()
        if not user:
            raise HTTPException(400, "Aktualne hasło jest nieprawidłowe")
        conn.execute("UPDATE users SET password=? WHERE id=?", (_hash(new_pass), uid))
        return {"ok": True}

@app.post("/api/users/{uid}/reset-password", dependencies=[Depends(verify_key)])
def reset_password(uid: int):
    import random, string
    new_pass = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
    with get_db() as conn:
        user = conn.execute("SELECT full_name FROM users WHERE id=?", (uid,)).fetchone()
        if not user:
            raise HTTPException(404, "Użytkownik nie znaleziony")
        conn.execute("UPDATE users SET password=? WHERE id=?", (_hash(new_pass), uid))
        return {"ok": True, "new_password": new_pass, "full_name": user["full_name"]}

@app.delete("/api/users/{uid}", dependencies=[Depends(verify_key)])
def delete_user(uid: int):
    with get_db() as conn:
        conn.execute("PRAGMA foreign_keys = ON")
        conn.execute("DELETE FROM sesje_pracy WHERE user_id=?", (uid,))
        conn.execute("DELETE FROM user_permissions WHERE user_id=?", (uid,))
        conn.execute("DELETE FROM users WHERE id=?", (uid,))
        return {"ok": True}

# ─── Uprawnienia użytkowników ────────────────────────────────────────────────
@app.get("/api/users/{uid}/permissions", dependencies=[Depends(verify_key)])
def get_user_permissions(uid: int):
    with get_db() as conn:
        row = conn.execute(
            "SELECT tabs FROM user_permissions WHERE user_id=?", (uid,)
        ).fetchone()
        if row:
            return {"user_id": uid, "tabs": json.loads(row["tabs"])}
        return {"user_id": uid, "tabs": []}

@app.get("/api/permissions/all", dependencies=[Depends(verify_key)])
def get_all_permissions():
    with get_db() as conn:
        rows = conn.execute("SELECT user_id, tabs FROM user_permissions").fetchall()
        return {r["user_id"]: json.loads(r["tabs"]) for r in rows}

class PermissionsRequest(BaseModel):
    tabs: List[str]

@app.put("/api/users/{uid}/permissions", dependencies=[Depends(verify_key)])
def set_user_permissions(uid: int, req: PermissionsRequest):
    with get_db() as conn:
        conn.execute("""
            INSERT INTO user_permissions (user_id, tabs, updated_at)
            VALUES (?, ?, ?)
            ON CONFLICT(user_id) DO UPDATE SET tabs=excluded.tabs, updated_at=excluded.updated_at
        """, (uid, json.dumps(req.tabs), _now()))
        return {"ok": True, "user_id": uid, "tabs": req.tabs}

@app.delete("/api/users/{uid}/permissions", dependencies=[Depends(verify_key)])
def reset_user_permissions(uid: int):
    with get_db() as conn:
        conn.execute("DELETE FROM user_permissions WHERE user_id=?", (uid,))
        return {"ok": True}

# ─── Katalog produktów CRUD ──────────────────────────────────────────────────
@app.get("/api/katalog", dependencies=[Depends(verify_key)])
def get_katalog():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM katalog_produktow ORDER BY nazwa").fetchall()
        return [dict(r) for r in rows]

class KatalogRequest(BaseModel):
    nazwa: str
    opis: Optional[str] = ""
    ilosc_domyslna: Optional[int] = 1
    cena_szt: Optional[float] = 0.0

@app.post("/api/katalog", dependencies=[Depends(verify_key)])
def create_produkt(req: KatalogRequest):
    with get_db() as conn:
        try:
            cur = conn.execute(
                "INSERT INTO katalog_produktow (nazwa,opis,ilosc_domyslna,cena_szt) VALUES (?,?,?,?)",
                (req.nazwa, req.opis, req.ilosc_domyslna, req.cena_szt)
            )
            return {"id": cur.lastrowid}
        except sqlite3.IntegrityError:
            raise HTTPException(400, "Produkt już istnieje")

@app.put("/api/katalog/{kid}", dependencies=[Depends(verify_key)])
def update_produkt(kid: int, req: KatalogRequest):
    with get_db() as conn:
        conn.execute(
            "UPDATE katalog_produktow SET nazwa=?,opis=?,ilosc_domyslna=?,cena_szt=? WHERE id=?",
            (req.nazwa, req.opis, req.ilosc_domyslna, req.cena_szt, kid)
        )
        return {"ok": True}

@app.delete("/api/katalog/{kid}", dependencies=[Depends(verify_key)])
def delete_produkt(kid: int):
    with get_db() as conn:
        conn.execute("DELETE FROM katalog_produktow WHERE id=?", (kid,))
        return {"ok": True}

# ─── QR Code generation ──────────────────────────────────────────────────────
@app.get("/api/qr/{kod}", dependencies=[Depends(verify_key)])
def generate_qr(kod: str):
    try:
        import qrcode
        qr = qrcode.QRCode(version=1, box_size=10, border=4)
        qr.add_data(kod)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        return StreamingResponse(buf, media_type="image/png",
            headers={"Content-Disposition": f'inline; filename="{kod}.png"'})
    except ImportError:
        raise HTTPException(500, "Brak biblioteki qrcode. Zainstaluj: pip install qrcode[pil]")


# ═══════════════════════════════════════════════════════════════════════════════
# ─── Drzewo G/P – Wyroby i struktura BOM ─────────────────────────────────────
# ═══════════════════════════════════════════════════════════════════════════════

# ── Model Pydantic ─────────────────────────────────────────────────────────────
class WyrobRequest(BaseModel):
    symbol: str
    typ: str = "P"                   # 'G' lub 'P'
    nazwa: str
    numer_rysunku: Optional[str] = ""
    jednostka: Optional[str] = "szt"
    opis: Optional[str] = ""

class WyrobyBomRequest(BaseModel):
    skladnik_id: Optional[int] = None     # id z tabeli wyroby (dla typ=P)
    typ_skladnika: str = "P"              # 'P' lub 'M'
    material_indeks: Optional[str] = ""   # dla typ=M
    ilosc: float = 1
    jednostka: Optional[str] = "szt"
    pozycja: Optional[int] = 0
    uwagi: Optional[str] = ""

class ZapotrzebowanieRequest(BaseModel):
    zlecenie_g_id: int
    wyrob_p_symbol: str
    ilosc_wymagana: float = 1
    priorytet: Optional[int] = 0
    uwagi: Optional[str] = ""
    zlecenie_p_id: Optional[int] = None   # od razu linkuj do zlecenia P

# ── CRUD wyrobów ───────────────────────────────────────────────────────────────
@app.get("/api/wyroby", dependencies=[Depends(verify_key)])
def get_wyroby(typ: Optional[str] = None, q: Optional[str] = None):
    """Lista wyrobów G lub P, opcjonalnie filtrowana."""
    with get_db() as conn:
        clauses, params = [], []
        if typ:
            clauses.append("typ=?")
            params.append(typ)
        if q:
            clauses.append("(symbol LIKE ? OR nazwa LIKE ?)")
            params += [f"%{q}%", f"%{q}%"]
        where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
        rows = conn.execute(f"SELECT * FROM wyroby {where} ORDER BY symbol", params).fetchall()
        return [dict(r) for r in rows]

@app.get("/api/wyroby/{wid}", dependencies=[Depends(verify_key)])
def get_wyrob(wid: int):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM wyroby WHERE id=?", (wid,)).fetchone()
        if not row:
            raise HTTPException(404, "Wyrób nie znaleziony")
        return dict(row)

@app.post("/api/wyroby", dependencies=[Depends(verify_key)])
def create_wyrob(req: WyrobRequest):
    if req.typ not in ("G", "P"):
        raise HTTPException(400, "typ musi być 'G' lub 'P'")
    with get_db() as conn:
        try:
            cur = conn.execute(
                """INSERT INTO wyroby (symbol,typ,nazwa,numer_rysunku,jednostka,opis)
                   VALUES (?,?,?,?,?,?)""",
                (req.symbol.strip(), req.typ, req.nazwa.strip(),
                 req.numer_rysunku or "", req.jednostka or "szt", req.opis or "")
            )
            return {"id": cur.lastrowid}
        except Exception as e:
            if "UNIQUE" in str(e):
                raise HTTPException(409, f"Symbol '{req.symbol}' już istnieje")
            raise HTTPException(500, str(e))

@app.put("/api/wyroby/{wid}", dependencies=[Depends(verify_key)])
def update_wyrob(wid: int, req: WyrobRequest):
    if req.typ not in ("G", "P"):
        raise HTTPException(400, "typ musi być 'G' lub 'P'")
    with get_db() as conn:
        conn.execute(
            """UPDATE wyroby SET symbol=?,typ=?,nazwa=?,numer_rysunku=?,jednostka=?,opis=?
               WHERE id=?""",
            (req.symbol.strip(), req.typ, req.nazwa.strip(),
             req.numer_rysunku or "", req.jednostka or "szt", req.opis or "", wid)
        )
        return {"ok": True}

@app.delete("/api/wyroby/{wid}", dependencies=[Depends(verify_key)])
def delete_wyrob(wid: int):
    with get_db() as conn:
        conn.execute("DELETE FROM wyroby_bom WHERE wyrob_id=? OR (typ_skladnika='P' AND skladnik_id=?)", (wid, wid))
        conn.execute("DELETE FROM wyroby WHERE id=?", (wid,))
        return {"ok": True}

# ── BOM wyrobu ─────────────────────────────────────────────────────────────────
@app.get("/api/wyroby/{wid}/bom", dependencies=[Depends(verify_key)])
def get_wyrob_bom(wid: int):
    """Bezpośrednie dzieci BOM danego wyrobu (1 poziom)."""
    with get_db() as conn:
        rows = conn.execute("""
            SELECT wb.*,
                   w.symbol as skladnik_symbol, w.nazwa as skladnik_nazwa,
                   w.typ as skladnik_typ_wyrobu,
                   m.opis as material_opis, m.jm as material_jm,
                   m.do_dyspozycji as material_stan
            FROM wyroby_bom wb
            LEFT JOIN wyroby w ON wb.skladnik_id = w.id AND wb.typ_skladnika = 'P'
            LEFT JOIN materialy m ON wb.material_indeks = m.indeks AND wb.typ_skladnika = 'M'
            WHERE wb.wyrob_id=?
            ORDER BY wb.pozycja, wb.id
        """, (wid,)).fetchall()
        return [dict(r) for r in rows]

@app.post("/api/wyroby/{wid}/bom", dependencies=[Depends(verify_key)])
def add_wyrob_bom(wid: int, req: WyrobyBomRequest):
    if req.typ_skladnika not in ("P", "M"):
        raise HTTPException(400, "typ_skladnika musi być 'P' lub 'M'")
    if req.typ_skladnika == "P" and not req.skladnik_id:
        raise HTTPException(400, "Dla typ_skladnika='P' wymagane skladnik_id")
    if req.typ_skladnika == "M" and not req.material_indeks:
        raise HTTPException(400, "Dla typ_skladnika='M' wymagane material_indeks")
    if req.typ_skladnika == "P" and req.skladnik_id == wid:
        raise HTTPException(400, "Wyrób nie może być swoim własnym składnikiem")
    with get_db() as conn:
        try:
            cur = conn.execute(
                """INSERT OR REPLACE INTO wyroby_bom
                   (wyrob_id, skladnik_id, typ_skladnika, material_indeks, ilosc, jednostka, pozycja, uwagi)
                   VALUES (?,?,?,?,?,?,?,?)""",
                (wid, req.skladnik_id or 0, req.typ_skladnika,
                 req.material_indeks or "", req.ilosc,
                 req.jednostka or "szt", req.pozycja or 0, req.uwagi or "")
            )
            return {"id": cur.lastrowid}
        except Exception as e:
            raise HTTPException(500, str(e))

@app.delete("/api/wyroby/{wid}/bom/{bid}", dependencies=[Depends(verify_key)])
def delete_wyrob_bom(wid: int, bid: int):
    with get_db() as conn:
        conn.execute("DELETE FROM wyroby_bom WHERE id=? AND wyrob_id=?", (bid, wid))
        return {"ok": True}

# ── Drzewo pełne (rekurencyjne) ────────────────────────────────────────────────
@app.get("/api/wyroby/{wid}/drzewo", dependencies=[Depends(verify_key)])
def get_wyrob_drzewo(wid: int, max_depth: int = 10):
    """
    Pełne drzewo struktury wyrobu – 3 zapytania SQL zamiast rekurencji N+1.
    1. WITH RECURSIVE → wszystkie węzły drzewa (id, parent_id, depth)
    2. Batch SELECT wszystkich wyrobów P z drzewa
    3. Batch SELECT wszystkich zleceń dla tych wyrobów
    Budowanie drzewa w pamięci: O(n).
    """
    with get_db() as conn:
        root = conn.execute("SELECT * FROM wyroby WHERE id=?", (wid,)).fetchone()
        if not root:
            raise HTTPException(404, "Wyrób nie znaleziony")

        # ── Zapytanie 1: wszystkie krawędzie BOM osiągalne z korzenia ─────────
        # WITH RECURSIVE przechodzi drzewo wyroby_bom w SQLite natywnie.
        # Zwraca: (bom_id, wyrob_id=rodzic, skladnik_id, typ, material_indeks,
        #          ilosc, jednostka, pozycja, uwagi, depth)
        bom_rows = conn.execute("""
            WITH RECURSIVE tree(
                bom_id, wyrob_id, skladnik_id, typ_skladnika,
                material_indeks, ilosc, jednostka, pozycja, uwagi, depth
            ) AS (
                -- korzeń: bezpośrednie dzieci wid
                SELECT wb.id, wb.wyrob_id, wb.skladnik_id, wb.typ_skladnika,
                       wb.material_indeks, wb.ilosc, wb.jednostka,
                       wb.pozycja, wb.uwagi, 1
                FROM wyroby_bom wb
                WHERE wb.wyrob_id = ?

                UNION ALL

                -- rekurencja: dzieci dzieci (tylko P mają dalszych potomków)
                SELECT wb.id, wb.wyrob_id, wb.skladnik_id, wb.typ_skladnika,
                       wb.material_indeks, wb.ilosc, wb.jednostka,
                       wb.pozycja, wb.uwagi, tree.depth + 1
                FROM wyroby_bom wb
                JOIN tree ON wb.wyrob_id = tree.skladnik_id
                         AND tree.typ_skladnika = 'P'
                WHERE tree.depth < ?
            )
            SELECT * FROM tree
            ORDER BY wyrob_id, pozycja, bom_id
        """, (wid, max_depth)).fetchall()

        # ── Zapytanie 2: batch wyrobów P (wszystkie IDs z drzewa) ─────────────
        p_ids = list({r["skladnik_id"] for r in bom_rows if r["typ_skladnika"] == "P"})
        p_ids.append(wid)  # dodaj korzeń
        wyroby_map = {}
        if p_ids:
            placeholders = ",".join("?" * len(p_ids))
            for w in conn.execute(
                f"SELECT * FROM wyroby WHERE id IN ({placeholders})", p_ids
            ).fetchall():
                wyroby_map[w["id"]] = dict(w)

        # ── Zapytanie 3: batch materiałów M (wszystkie indeksy z drzewa) ──────
        m_indeksy = list({r["material_indeks"] for r in bom_rows
                          if r["typ_skladnika"] == "M" and r["material_indeks"]})
        materialy_map = {}
        if m_indeksy:
            placeholders = ",".join("?" * len(m_indeksy))
            for m in conn.execute(
                f"SELECT indeks, opis, jm, do_dyspozycji FROM materialy WHERE indeks IN ({placeholders})",
                m_indeksy
            ).fetchall():
                materialy_map[m["indeks"]] = dict(m)

        # ── Zapytanie 4: batch zleceń dla wyrobów P z drzewa ─────────────────
        p_symbols = [wyroby_map[pid]["symbol"] for pid in p_ids if pid in wyroby_map]
        zlecenia_map = {}  # symbol → [lista zleceń]
        if p_symbols:
            placeholders = ",".join("?" * len(p_symbols))
            for z in conn.execute(f"""
                SELECT z.id, z.numer, z.status, z.ilosc_sztuk,
                       COUNT(o.id)  as op_total,
                       SUM(CASE WHEN o.status='zakonczona' THEN 1 ELSE 0 END) as op_done
                FROM zlecenia z
                LEFT JOIN operacje o ON o.zlecenie_id = z.id
                WHERE z.numer IN ({placeholders})
                GROUP BY z.id
            """, p_symbols).fetchall():
                sym = z["numer"]
                zlecenia_map.setdefault(sym, []).append(dict(z))

        # ── Budowanie drzewa w pamięci ─────────────────────────────────────────
        # Kluczowa zmiana: węzły P są indeksowane po bom_id (unikalnym), nie po
        # wyrob_id – dzięki temu ten sam wyrób P występujący wielokrotnie w drzewie
        # (pod różnymi rodzicami lub z różnymi ilościami) tworzy osobne węzły
        # i nie dochodzi do scalania/sumowania ilości.
        #
        # children_by_parent[wyrob_id] = [lista węzłów-dzieci (obiekty słownikowe)]
        # nodes_by_bom_id[bom_id]      = węzeł P (referencja do obiektu w liście children)

        children_by_parent: dict = {}   # wyrob_id  → [child_node, ...]
        nodes_by_bom_id: dict    = {}   # bom_id    → node dict (tylko P)

        for r in bom_rows:
            parent_id = r["wyrob_id"]
            children_by_parent.setdefault(parent_id, [])

            if r["typ_skladnika"] == "P":
                child_wyrob = wyroby_map.get(r["skladnik_id"], {})
                # Filtruj zlecenia wg ilosc_sztuk – pobieramy te, których ilosc_sztuk
                # odpowiada ilości z BOM (lub wszystkie, jeśli nie ma dopasowania).
                all_zl = zlecenia_map.get(child_wyrob.get("symbol", ""), [])
                bom_ilosc = r["ilosc"]
                matching = [z for z in all_zl if z.get("ilosc_sztuk") == bom_ilosc]
                zlecenia_for_node = matching if matching else all_zl

                node = dict(child_wyrob)
                node["zlecenia"]       = zlecenia_for_node
                node["_bom_id"]        = r["bom_id"]
                node["_bom_ilosc"]     = bom_ilosc
                node["_bom_jednostka"] = r["jednostka"]
                node["_bom_pozycja"]   = r["pozycja"]
                node["_bom_uwagi"]     = r["uwagi"]
                node["children"]       = []  # wypełni się poniżej
                children_by_parent[parent_id].append(node)
                nodes_by_bom_id[r["bom_id"]] = node
            else:  # M
                mat = materialy_map.get(r["material_indeks"], {})
                m_node = {
                    "typ":            "M",
                    "material_indeks": r["material_indeks"],
                    # Opis: priorytet słownik materiałów → fallback uwagi BOM → indeks
                    "material_opis":   mat.get("opis") or r.get("uwagi") or r["material_indeks"],
                    "material_jm":     mat.get("jm", r["jednostka"]),
                    "material_stan":   mat.get("do_dyspozycji", 0),
                    "ilosc":           r["ilosc"],
                    "jednostka":       r["jednostka"],
                    "_bom_id":         r["bom_id"],
                    "_bom_pozycja":    r["pozycja"],
                    "_bom_uwagi":      r.get("uwagi", ""),
                }
                children_by_parent[parent_id].append(m_node)

        # Przypisz children do węzłów P (jeden przebieg po bom_rows).
        # Używamy nodes_by_bom_id żeby trafić w konkretny węzeł-instancję,
        # a nie w dowolne wystąpienie tego samego wyrob_id.
        def _attach_children(node: dict) -> dict:
            nid = node.get("id")
            if nid and nid in children_by_parent:
                node["children"] = [
                    _attach_children(dict(ch)) if ch.get("typ") != "M" else ch
                    for ch in children_by_parent[nid]
                ]
            elif "children" not in node:
                node["children"] = []
            return node

        # Zbuduj korzeń
        root_node = dict(wyroby_map.get(wid, dict(root)))
        root_node["zlecenia"] = zlecenia_map.get(root_node.get("symbol", ""), [])
        root_node["children"] = [
            _attach_children(dict(ch)) if ch.get("typ") != "M" else ch
            for ch in children_by_parent.get(wid, [])
        ]
        return root_node

# ── Zapotrzebowania (G→P linkowanie do zleceń) ─────────────────────────────────
@app.get("/api/zlecenia/{gid}/zapotrzebowania", dependencies=[Depends(verify_key)])
def get_zapotrzebowania(gid: int):
    """Lista półproduktów P wymaganych przez zlecenie G z ich statusem realizacji."""
    with get_db() as conn:
        rows = conn.execute("""
            SELECT z.*,
                   zp.numer as zlecenie_p_numer, zp.status as zlecenie_p_status,
                   zp.ilosc_sztuk as zlecenie_p_ilosc_sztuk,
                   (SELECT COUNT(*) FROM operacje o WHERE o.zlecenie_id=zp.id) as op_total,
                   (SELECT COUNT(*) FROM operacje o WHERE o.zlecenie_id=zp.id AND o.status='zakonczona') as op_done,
                   w.nazwa as wyrob_nazwa, w.numer_rysunku, w.id as wyrob_id
            FROM zapotrzebowania z
            LEFT JOIN zlecenia zp ON z.zlecenie_p_id = zp.id
            LEFT JOIN wyroby w ON w.symbol = z.wyrob_p_symbol
            WHERE z.zlecenie_g_id=?
            ORDER BY z.priorytet DESC, z.id
        """, (gid,)).fetchall()
        return [dict(r) for r in rows]

@app.post("/api/zlecenia/{gid}/zapotrzebowania", dependencies=[Depends(verify_key)])
def add_zapotrzebowanie(gid: int, req: ZapotrzebowanieRequest):
    with get_db() as conn:
        g = conn.execute("SELECT id FROM zlecenia WHERE id=?", (gid,)).fetchone()
        if not g:
            raise HTTPException(404, "Zlecenie G nie znalezione")
        cur = conn.execute(
            """INSERT INTO zapotrzebowania
               (zlecenie_g_id, wyrob_p_symbol, ilosc_wymagana, priorytet, uwagi, zlecenie_p_id)
               VALUES (?,?,?,?,?,?)""",
            (gid, req.wyrob_p_symbol.strip(), req.ilosc_wymagana,
             req.priorytet or 0, req.uwagi or "", req.zlecenie_p_id)
        )
        return {"id": cur.lastrowid}

@app.patch("/api/zapotrzebowania/{zap_id}/link", dependencies=[Depends(verify_key)])
def link_zapotrzebowanie_to_zlecenie(zap_id: int, body: dict = Body(...)):
    """Powiąż zapotrzebowanie P z istniejącym lub nowo utworzonym zleceniem P."""
    zlecenie_p_id = body.get("zlecenie_p_id")
    with get_db() as conn:
        if zlecenie_p_id:
            zp = conn.execute("SELECT id FROM zlecenia WHERE id=?", (zlecenie_p_id,)).fetchone()
            if not zp:
                raise HTTPException(404, "Zlecenie P nie znalezione")
        conn.execute(
            "UPDATE zapotrzebowania SET zlecenie_p_id=? WHERE id=?",
            (zlecenie_p_id, zap_id)
        )
        return {"ok": True}

@app.patch("/api/zapotrzebowania/{zap_id}/status", dependencies=[Depends(verify_key)])
def update_zapotrzebowanie_status(zap_id: int, body: dict = Body(...)):
    status = body.get("status")
    if status not in ("oczekuje", "w_toku", "zakonczone", "anulowane"):
        raise HTTPException(400, "Nieprawidłowy status")
    ilosc_wyk = body.get("ilosc_wykonana")
    with get_db() as conn:
        if ilosc_wyk is not None:
            conn.execute(
                "UPDATE zapotrzebowania SET status=?, ilosc_wykonana=? WHERE id=?",
                (status, ilosc_wyk, zap_id)
            )
        else:
            conn.execute("UPDATE zapotrzebowania SET status=? WHERE id=?", (status, zap_id))
        return {"ok": True}

@app.delete("/api/zapotrzebowania/{zap_id}", dependencies=[Depends(verify_key)])
def delete_zapotrzebowanie(zap_id: int):
    with get_db() as conn:
        conn.execute("DELETE FROM zapotrzebowania WHERE id=?", (zap_id,))
        return {"ok": True}

# ── Import drzewa G/P z PDF (Graffiti ERP) ─────────────────────────────────────

def _parse_graffiti_bom_pdf(pdf_bytes: bytes) -> dict:
    """
    Parsuje PDF 'Drzewo technologiczne' z Graffiti ERP używając analizy układu (x,y).
    Każdy wiersz BOM to zestaw textboxów na tym samym Y:
      - x < 160  : symbol G/P/M  (wcięcie x → głębokość hierarchii)
      - 160-380  : [nr] Opis
      - 380-440  : ilość
      - 440-490  : JMT
    Zwraca listę wierszy z polem 'depth' wyznaczonym z x0 symbolu.
    """
    import io as _io
    import re as _re
    from pdfminer.high_level import extract_pages as _extract_pages
    from pdfminer.layout import LTTextBox as _LTTextBox

    SKIP_Y_MAX   = 100    # nagłówek strony (tytuł, data, nr G, kolumny)
    SKIP_Y_MIN   = 730    # stopka (Wydrukowano, Strona X/70)
    X_SYM_MAX    = 160    # kolumna symboli G/P/M
    X_OPIS_MIN   = 160    # kolumna [nr] Opis
    X_OPIS_MAX   = 385
    X_ILOSC_MIN  = 385    # kolumna ilości
    X_ILOSC_MAX  = 445
    X_JM_MIN     = 445    # kolumna JMT
    X_JM_MAX     = 495
    ROW_TOL      = 4      # tolerancja grupowania wierszy (pt)
    X_BASE       = 28.8   # X lewej krawędzi dla głębokości 0
    X_STEP       = 14.3   # przyrost X na jeden poziom wcięcia

    SYM_RE = _re.compile(r'^([GPM]\d+)$')

    rows = []

    with _io.BytesIO(pdf_bytes) as buf:
        for page_num, page_layout in enumerate(_extract_pages(buf)):
            page_h = page_layout.height
            boxes = []
            for el in page_layout:
                if not isinstance(el, _LTTextBox):
                    continue
                raw = el.get_text().strip()
                if not raw:
                    continue
                y0 = page_h - el.y1   # 0 = góra strony
                boxes.append({'x0': el.x0, 'y0': y0, 'text': raw})

            # Sortuj: Y rosnąco (góra→dół), potem X rosnąco
            boxes.sort(key=lambda b: (round(b['y0']), b['x0']))

            # Grupuj boxy na tym samym Y w "wiersze"
            page_rows = []
            cur_row, cur_y = [], None
            for b in boxes:
                if cur_y is None or abs(b['y0'] - cur_y) <= ROW_TOL:
                    cur_row.append(b)
                    cur_y = b['y0']
                else:
                    if cur_row:
                        page_rows.append(cur_row)
                    cur_row, cur_y = [b], b['y0']
            if cur_row:
                page_rows.append(cur_row)

            for row in page_rows:
                y0 = row[0]['y0']
                if y0 < SKIP_Y_MAX or y0 > SKIP_Y_MIN:
                    continue

                sym_box = opis_box = ilosc_box = jm_box = None
                for b in row:
                    x, txt = b['x0'], b['text']
                    if x < X_SYM_MAX and SYM_RE.match(txt) and sym_box is None:
                        sym_box = b
                    elif X_OPIS_MIN <= x < X_OPIS_MAX and opis_box is None:
                        opis_box = b
                    elif X_ILOSC_MIN <= x < X_ILOSC_MAX and ilosc_box is None:
                        ilosc_box = b
                    elif X_JM_MIN <= x < X_JM_MAX and jm_box is None:
                        jm_box = b

                if sym_box is None:
                    continue

                sym  = sym_box['text']
                opis = opis_box['text'] if opis_box else ''
                # Usuń [nr] na początku opisu i złącz wieloliniowe
                opis = _re.sub(r'^\[\d+\]\s*', '', opis)
                opis = _re.sub(r'\s+', ' ', opis).strip()
                # Usuń "Cechy dostaw: ..." z opisu materiału
                opis = _re.sub(r'\s*Cechy dostaw:.*$', '', opis).strip()

                ilosc = 1.0
                if ilosc_box:
                    raw_i = ilosc_box['text'].replace('\xa0', '').replace(' ', '').replace(',', '.')
                    try:
                        ilosc = float(raw_i)
                    except Exception:
                        pass

                jm    = jm_box['text'].strip() if jm_box else 'szt'
                depth = max(0, round((sym_box['x0'] - X_BASE) / X_STEP))
                typ   = 'G' if sym.startswith('G') else ('P' if sym.startswith('P') else 'M')

                rows.append({
                    'symbol': sym,
                    'opis':   opis,
                    'ilosc':  ilosc,
                    'jm':     jm,
                    'depth':  depth,
                    'typ':    typ,
                })

    if not rows:
        raise ValueError("Nie znaleziono żadnych wierszy BOM w PDF")

    # Pierwszy wiersz to wyrób główny (G lub P), depth=0
    first = rows[0]
    return {
        'symbol_glowny': first['symbol'],
        'nazwa_glowna':  first['opis'],
        'typ_glowny':    first['typ'],
        'rows':          rows,
    }


@app.post("/api/import-drzewo-gp", dependencies=[Depends(verify_key)])
async def import_drzewo_gp(file: UploadFile = File(...)):
    """
    Parsuje plik PDF 'Drzewo technologiczne' z Graffiti ERP.
    Używa analizy układu PDF (współrzędne X) do odczytu hierarchii wcięć.
    Tworzy/aktualizuje wyroby G i P oraz ich pełną strukturę BOM.
    """
    import json as _json

    pdf_bytes = await file.read()
    if not pdf_bytes:
        raise HTTPException(400, "Pusty plik")

    try:
        parsed = _parse_graffiti_bom_pdf(pdf_bytes)
    except Exception as e:
        raise HTTPException(400, f"Błąd parsowania PDF: {e}")

    rows          = parsed['rows']
    glowny_symbol = parsed['symbol_glowny']
    glowna_nazwa  = parsed['nazwa_glowna']
    glowny_typ    = parsed['typ_glowny']

    errors        = []
    wyroby_created = 0
    bom_created    = 0

    with get_db() as conn:
        # ── 1. Utwórz/zaktualizuj wszystkie wyroby P i G z drzewa ────────────
        for row in rows:
            if row['typ'] not in ('G', 'P'):
                continue
            ex = conn.execute(
                "SELECT id FROM wyroby WHERE symbol=?", (row['symbol'],)
            ).fetchone()
            if ex:
                conn.execute(
                    "UPDATE wyroby SET nazwa=CASE WHEN nazwa='' OR nazwa IS NULL THEN ? ELSE nazwa END, "
                    "jednostka=COALESCE(NULLIF(jednostka,''),?) WHERE symbol=?",
                    (row['opis'], row['jm'], row['symbol'])
                )
            else:
                conn.execute(
                    "INSERT INTO wyroby (symbol, typ, nazwa, jednostka) VALUES (?,?,?,?)",
                    (row['symbol'], row['typ'], row['opis'], row['jm'])
                )
                wyroby_created += 1

        # Pobierz ID wyrobu głównego (po upsert)
        g_row = conn.execute(
            "SELECT id FROM wyroby WHERE symbol=?", (glowny_symbol,)
        ).fetchone()
        if not g_row:
            # Wstaw wyrób główny jeśli jeszcze nie istnieje
            cur = conn.execute(
                "INSERT INTO wyroby (symbol, typ, nazwa) VALUES (?,?,?)",
                (glowny_symbol, glowny_typ, glowna_nazwa)
            )
            gid = cur.lastrowid
            wyroby_created += 1
        else:
            gid = g_row['id']
            conn.execute(
                "UPDATE wyroby SET nazwa=?, typ=? WHERE id=?",
                (glowna_nazwa, glowny_typ, gid)
            )

        # ── 2. Usuń stary BOM tego wyrobu G ───────────────────────────────────
        conn.execute(
            "DELETE FROM wyroby_bom WHERE wyrob_id=?", (gid,)
        )

        # ── 3. Zbuduj hierarchię przez stos ojców ──────────────────────────────
        # parent_stack[depth] = id wyrobu na tym poziomie
        # depth 0 = G główne, depth 1 = dzieci G, depth 2 = dzieci P, ...
        # Wiersz rows[0] to sam G (depth=0) - pomijamy go jako rodzic zaczyna od depth=1
        parent_stack = {0: gid}

        # Usuń stary BOM dla wszystkich P z tego drzewa (żeby uniknąć duplikatów
        # gdy ten sam P występuje w wielu miejscach drzewa)
        all_p_symbols = [r['symbol'] for r in rows if r['typ'] == 'P']
        for psym in all_p_symbols:
            pr = conn.execute("SELECT id FROM wyroby WHERE symbol=?", (psym,)).fetchone()
            if pr:
                conn.execute("DELETE FROM wyroby_bom WHERE wyrob_id=?", (pr['id'],))

        # Przetwarzaj wiersze od indeksu 1 (pomijamy root G)
        # Zbierz BOM per wyrob_id żeby wstawiać w kolejności z pozycją
        bom_inserts = {}  # wyrob_id -> [(skladnik_info, pozycja)]

        for row in rows[1:]:
            depth  = row['depth']
            sym    = row['symbol']
            typ    = row['typ']
            opis   = row['opis']
            ilosc  = row['ilosc']
            jm     = row['jm']

            # Rodzic = element ze stosu na poziomie depth-1
            parent_depth = depth - 1
            if parent_depth < 0:
                parent_depth = 0
            parent_id = parent_stack.get(parent_depth, gid)

            # Aktualizuj stos: ten element jest teraz rodzicem na swoim poziomie
            if typ in ('G', 'P'):
                child_row = conn.execute(
                    "SELECT id FROM wyroby WHERE symbol=?", (sym,)
                ).fetchone()
                child_id = child_row['id'] if child_row else None
                if child_id:
                    parent_stack[depth] = child_id
                    # Wyczyść głębsze poziomy żeby nie "przeciekały" z poprzedniej gałęzi
                    for d in list(parent_stack.keys()):
                        if d > depth:
                            del parent_stack[d]

                if parent_id not in bom_inserts:
                    bom_inserts[parent_id] = []
                bom_inserts[parent_id].append({
                    'typ': 'P',
                    'symbol': sym,
                    'child_id': child_id,
                    'ilosc': ilosc,
                    'jm': jm,
                })
            else:  # M
                if parent_id not in bom_inserts:
                    bom_inserts[parent_id] = []
                bom_inserts[parent_id].append({
                    'typ': 'M',
                    'symbol': sym,
                    'ilosc': ilosc,
                    'jm': jm,
                })

        # ── 4. Wstaw BOM do bazy ───────────────────────────────────────────────
        for wyrob_id, items in bom_inserts.items():
            pozycja = 0
            for item in items:
                try:
                    if item['typ'] == 'M':
                        conn.execute(
                            """INSERT OR IGNORE INTO wyroby_bom
                               (wyrob_id, skladnik_id, typ_skladnika, material_indeks, ilosc, jednostka, pozycja)
                               VALUES (?, 0, 'M', ?, ?, ?, ?)""",
                            (wyrob_id, item['symbol'], item['ilosc'], item['jm'], pozycja)
                        )
                    else:  # P
                        if item['child_id']:
                            conn.execute(
                                """INSERT OR IGNORE INTO wyroby_bom
                                   (wyrob_id, skladnik_id, typ_skladnika, material_indeks, ilosc, jednostka, pozycja)
                                   VALUES (?, ?, 'P', '', ?, ?, ?)""",
                                (wyrob_id, item['child_id'], item['ilosc'], item['jm'], pozycja)
                            )
                    bom_created += 1
                except Exception as e:
                    errors.append(f"{item['typ']} {item['symbol']} (parent={wyrob_id}): {e}")
                pozycja += 1

        # ── 5. Log importu ─────────────────────────────────────────────────────
        conn.execute(
            """INSERT INTO import_log (typ, symbol_glowny, ilosc_wyrobow, ilosc_pozycji_bom, bledy)
               VALUES ('drzewo_gp', ?, ?, ?, ?)""",
            (glowny_symbol, wyroby_created, bom_created,
             _json.dumps(errors[:50], ensure_ascii=False))
        )

    _threading.Thread(target=_db_backup_to_json, daemon=True).start()
    return {
        "ok":             True,
        "symbol_glowny":  glowny_symbol,
        "nazwa_glowna":   glowna_nazwa,
        "wyroby_created": wyroby_created,
        "bom_created":    bom_created,
        "items_parsed":   len(rows),
        "errors":         errors[:20],
    }

# ── MRP: zapotrzebowanie materiałowe dla zlecenia G ───────────────────────────
@app.get("/api/zlecenia/{gid}/mrp", dependencies=[Depends(verify_key)])
def get_mrp_zlecenia(gid: int):
    """
    Oblicza sumaryczne zapotrzebowanie materiałowe dla zlecenia G.
    Uwzględnia cały BOM rekurencyjnie.
    Zwraca widok zbiorczy (per materiał) i szczegółowy (per P).
    """
    with get_db() as conn:
        zl_g = conn.execute("SELECT * FROM zlecenia WHERE id=?", (gid,)).fetchone()
        if not zl_g:
            raise HTTPException(404, "Zlecenie nie znalezione")

        ilosc_g = zl_g["ilosc_sztuk"] or 1

        # Znajdź wyrób G powiązany numerem zlecenia
        wyrob_g = conn.execute("SELECT * FROM wyroby WHERE symbol=?", (zl_g["numer"],)).fetchone()

        def _collect_materials(wyrob_id: int, mnoznik: float, visited: set) -> list:
            """Rekurencyjnie zbiera materiały z BOM."""
            if wyrob_id in visited:
                return []
            visited = visited | {wyrob_id}
            results = []
            bom_rows = conn.execute("""
                SELECT wb.*, w.id as w_id, w.symbol as w_symbol, w.nazwa as w_nazwa,
                       m.opis as m_opis, m.jm as m_jm,
                       m.do_dyspozycji as m_stan, m.stan_rzeczywisty as m_stan_rzecz,
                       m.indeks as m_indeks, m.id as m_id
                FROM wyroby_bom wb
                LEFT JOIN wyroby w ON wb.skladnik_id=w.id AND wb.typ_skladnika='P'
                LEFT JOIN materialy m ON wb.material_indeks=m.indeks AND wb.typ_skladnika='M'
                WHERE wb.wyrob_id=?
            """, (wyrob_id,)).fetchall()

            for row in bom_rows:
                rd = dict(row)
                ilosc_wymagana = (rd["ilosc"] or 1) * mnoznik
                if rd["typ_skladnika"] == "M":
                    results.append({
                        "material_indeks": rd["material_indeks"],
                        "material_opis":   rd.get("m_opis") or rd["material_indeks"],
                        "material_jm":     rd.get("m_jm", "szt"),
                        "material_id":     rd.get("m_id"),
                        "material_stan":   rd.get("m_stan") or 0,
                        "material_stan_rzecz": rd.get("m_stan_rzecz") or 0,
                        "ilosc_wymagana":  ilosc_wymagana,
                        "jednostka":       rd["jednostka"],
                    })
                elif rd["typ_skladnika"] == "P" and rd.get("w_id"):
                    # Rekurencja w głąb
                    sub = _collect_materials(rd["w_id"], ilosc_wymagana, visited)
                    results.extend(sub)
            return results

        # Zbierz wszystkie materiały
        if wyrob_g:
            raw_materials = _collect_materials(wyrob_g["id"], ilosc_g, set())
        else:
            # Fallback: użyj BOM z tabeli bom_pozycje
            bom_rows = conn.execute("""
                SELECT bp.*, m.opis, m.jm, m.do_dyspozycji, m.stan_rzeczywisty, m.indeks, m.id as m_id
                FROM bom_pozycje bp
                JOIN materialy m ON bp.material_id=m.id
                WHERE bp.zlecenie_id=?
            """, (gid,)).fetchall()
            raw_materials = [{
                "material_indeks": r["indeks"],
                "material_opis":   r["opis"],
                "material_jm":     r["jm"],
                "material_id":     r["m_id"],
                "material_stan":   r["do_dyspozycji"] or 0,
                "material_stan_rzecz": r["stan_rzeczywisty"] or 0,
                "ilosc_wymagana":  r["ilosc"] * ilosc_g,
                "jednostka":       r["jm"],
            } for r in bom_rows]

        # Agreguj per materiał (widok zbiorczy)
        zbiorczo: dict = {}
        for mat in raw_materials:
            key = mat["material_indeks"]
            if key not in zbiorczo:
                zbiorczo[key] = {**mat, "ilosc_wymagana": 0}
            zbiorczo[key]["ilosc_wymagana"] += mat["ilosc_wymagana"]

        zbiorczy = []
        for key, mat in zbiorczo.items():
            mat["ilosc_wymagana"] = round(mat["ilosc_wymagana"], 4)
            mat["braki"] = max(0, round(mat["ilosc_wymagana"] - mat["material_stan"], 4))
            mat["status_dostepnosci"] = (
                "ok" if mat["material_stan"] >= mat["ilosc_wymagana"]
                else ("czesciowo" if mat["material_stan"] > 0 else "brak")
            )
            zbiorczy.append(mat)

        zbiorczy.sort(key=lambda x: x.get("material_opis", ""))

        # Status zapotrzebowań P
        zapotrz = conn.execute("""
            SELECT z.*, zp.numer as p_numer, zp.status as p_status,
                   zp.ilosc_sztuk as p_ilosc,
                   w.nazwa as wyrob_nazwa
            FROM zapotrzebowania z
            LEFT JOIN zlecenia zp ON z.zlecenie_p_id = zp.id
            LEFT JOIN wyroby w ON w.symbol = z.wyrob_p_symbol
            WHERE z.zlecenie_g_id=?
            ORDER BY z.priorytet DESC
        """, (gid,)).fetchall()

        # ── Agregacja czasów operacji per stanowisko ──────────────────────────
        czasy_stanowisk: dict = {}
        if wyrob_g:
            # Zbierz wszystkie symbole P rekurencyjnie z BOM
            def _collect_p_symbols(wyrob_id: int, visited: set) -> set:
                if wyrob_id in visited:
                    return set()
                visited = visited | {wyrob_id}
                p_syms = set()
                sub_rows = conn.execute(
                    "SELECT typ_skladnika, skladnik_id FROM wyroby_bom WHERE wyrob_id=?",
                    (wyrob_id,)
                ).fetchall()
                for sr in sub_rows:
                    if sr["typ_skladnika"] == "P" and sr["skladnik_id"]:
                        sub_w = conn.execute(
                            "SELECT symbol FROM wyroby WHERE id=?", (sr["skladnik_id"],)
                        ).fetchone()
                        if sub_w:
                            p_syms.add(sub_w["symbol"])
                            p_syms |= _collect_p_symbols(sr["skladnik_id"], visited)
                return p_syms

            all_p_syms = _collect_p_symbols(wyrob_g["id"], set())
            # Dodaj też symbol G (zlecenie G może mieć własne operacje)
            all_syms = list(all_p_syms | {zl_g["numer"]})

            if all_syms:
                placeholders = ",".join("?" * len(all_syms))
                op_rows = conn.execute(f"""
                    SELECT o.stanowisko,
                           o.czas_norma,
                           COALESCE(o.czas_zbrojenia_min, 0) as czas_zbrojenia_min,
                           COALESCE(z.ilosc_sztuk, 1) as ilosc_sztuk
                    FROM operacje o
                    JOIN zlecenia z ON o.zlecenie_id = z.id
                    WHERE z.numer IN ({placeholders})
                      AND o.stanowisko IS NOT NULL
                      AND o.stanowisko != ''
                """, all_syms).fetchall()

                for op in op_rows:
                    st = op["stanowisko"]
                    if st not in czasy_stanowisk:
                        czasy_stanowisk[st] = {
                            "stanowisko": st,
                            "czas_norma_min": 0.0,
                            "zbrojenie_min": 0.0,
                            "czas_razem_min": 0.0,
                        }
                    ilosc = op["ilosc_sztuk"] or 1
                    norma = (op["czas_norma"] or 0.0) * ilosc
                    zbroj = op["czas_zbrojenia_min"] or 0.0
                    czasy_stanowisk[st]["czas_norma_min"] += norma
                    czasy_stanowisk[st]["zbrojenie_min"]  += zbroj
                    czasy_stanowisk[st]["czas_razem_min"] += norma + zbroj

        czasy_lista = sorted(
            [{"stanowisko": k, **v} for k, v in czasy_stanowisk.items()],
            key=lambda x: x["czas_razem_min"], reverse=True
        )

        return {
            "zlecenie": dict(zl_g),
            "wyrob_g": dict(wyrob_g) if wyrob_g else None,
            "zapotrzebowania_p": [dict(r) for r in zapotrz],
            "materialy_zbiorczy": zbiorczy,
            "materialy_szczegolowy": [dict(m) for m in raw_materials],
            "czasy_stanowisk": czasy_lista,
            "summary": {
                "material_count": len(zbiorczy),
                "brak_count": sum(1 for m in zbiorczy if m["status_dostepnosci"] == "brak"),
                "czesciowo_count": sum(1 for m in zbiorczy if m["status_dostepnosci"] == "czesciowo"),
                "ok_count": sum(1 for m in zbiorczy if m["status_dostepnosci"] == "ok"),
            }
        }

# ── MRP: Rezerwacja materiałów dla zlecenia G ─────────────────────────────────
@app.post("/api/zlecenia/{gid}/mrp/rezerwuj", dependencies=[Depends(verify_key)])
def mrp_rezerwuj(gid: int, body: dict = Body(...)):
    """
    Rezerwuje materiały MRP dla zlecenia G.
    Tworzy wpisy w mrp_rezerwacje i zmniejsza materialy.do_dyspozycji.
    body: { "pozycje": [{material_indeks, ilosc_do_rezerwacji}] }
    body.tryb: "dostepne" → rezerwuj tyle ile jest na stanie (min(wymagane, dostepne))
               "wszystko"  → rezerwuj pełne wymaganie (nawet jeśli brak na stanie)
    """
    pozycje = body.get("pozycje", [])
    tryb = body.get("tryb", "dostepne")  # dostepne | wszystko
    if not pozycje:
        raise HTTPException(400, "Brak pozycji do rezerwacji")

    with get_db() as conn:
        zl = conn.execute("SELECT numer FROM zlecenia WHERE id=?", (gid,)).fetchone()
        if not zl:
            raise HTTPException(404, "Zlecenie nie znalezione")
        zlecenie_nr = zl["numer"]

        zarezerwowane = []
        pominięte = []

        for poz in pozycje:
            indeks = poz.get("material_indeks", "").strip()
            ilosc_req = float(poz.get("ilosc_do_rezerwacji", 0) or 0)
            if not indeks or ilosc_req <= 0:
                continue

            mat = conn.execute(
                "SELECT id, do_dyspozycji FROM materialy WHERE indeks=?", (indeks,)
            ).fetchone()
            if not mat:
                pominięte.append({"indeks": indeks, "powod": "nie znaleziono"})
                continue

            dostepne = float(mat["do_dyspozycji"] or 0)
            if tryb == "dostepne":
                ilosc_rez = min(ilosc_req, max(dostepne, 0))
            else:
                ilosc_rez = ilosc_req

            if ilosc_rez <= 0:
                pominięte.append({"indeks": indeks, "powod": "brak stanu"})
                continue

            # Zmniejsz do_dyspozycji
            new_dysp = max(dostepne - ilosc_rez, 0)
            conn.execute(
                "UPDATE materialy SET do_dyspozycji=? WHERE indeks=?",
                (round(new_dysp, 4), indeks)
            )

            # Zapisz rezerwację
            conn.execute("""
                INSERT OR IGNORE INTO mrp_rezerwacje
                    (zlecenie_id, zlecenie_nr, material_indeks, ilosc)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(zlecenie_id, material_indeks) DO UPDATE SET
                    ilosc = ilosc + excluded.ilosc,
                    updated_at = CURRENT_TIMESTAMP
            """, (gid, zlecenie_nr, indeks, round(ilosc_rez, 4)))
            zarezerwowane.append({"indeks": indeks, "ilosc": round(ilosc_rez, 4)})

        return {
            "ok": True,
            "zarezerwowane": len(zarezerwowane),
            "pominięte": len(pominięte),
            "szczegoly": zarezerwowane,
        }


@app.delete("/api/zlecenia/{gid}/mrp/rezerwuj", dependencies=[Depends(verify_key)])
def mrp_anuluj_rezerwacje(gid: int):
    """Anuluje wszystkie rezerwacje MRP dla zlecenia i przywraca stany magazynowe."""
    with get_db() as conn:
        rezerwy = conn.execute(
            "SELECT material_indeks, ilosc FROM mrp_rezerwacje WHERE zlecenie_id=?", (gid,)
        ).fetchall()
        for r in rezerwy:
            conn.execute(
                "UPDATE materialy SET do_dyspozycji = do_dyspozycji + ? WHERE indeks=?",
                (r["ilosc"], r["material_indeks"])
            )
        deleted = conn.execute(
            "DELETE FROM mrp_rezerwacje WHERE zlecenie_id=?", (gid,)
        ).rowcount
        return {"ok": True, "zwolniono": deleted}


@app.get("/api/zlecenia/{gid}/mrp/rezerwacje", dependencies=[Depends(verify_key)])
def mrp_get_rezerwacje(gid: int):
    """Zwraca aktualne rezerwacje MRP dla zlecenia."""
    with get_db() as conn:
        rows = conn.execute("""
            SELECT r.*, m.opis as material_opis, m.jm as material_jm
            FROM mrp_rezerwacje r
            LEFT JOIN materialy m ON m.indeks = r.material_indeks
            WHERE r.zlecenie_id=?
            ORDER BY r.material_indeks
        """, (gid,)).fetchall()
        return [dict(r) for r in rows]


@app.get("/api/wyroby-g/postep", dependencies=[Depends(verify_key)])
def get_wyroby_g_postep():
    """Widok zarządczy: wyroby G ze statusem zlecenia i postępem P."""
    with get_db() as conn:
        wyroby_g = conn.execute(
            "SELECT * FROM wyroby WHERE typ='G' ORDER BY symbol"
        ).fetchall()

        result = []
        for wg in wyroby_g:
            wg_dict = dict(wg)

            # Zlecenie G
            zl_g = conn.execute(
                "SELECT * FROM zlecenia WHERE numer=?", (wg["symbol"],)
            ).fetchone()
            wg_dict["zlecenie"] = dict(zl_g) if zl_g else None

            # Zapotrzebowania P i ich realizacja
            if zl_g:
                zapotrz = conn.execute("""
                    SELECT z.*, zp.status as p_status, w.nazwa as wyrob_nazwa
                    FROM zapotrzebowania z
                    LEFT JOIN zlecenia zp ON z.zlecenie_p_id = zp.id
                    LEFT JOIN wyroby w ON w.symbol = z.wyrob_p_symbol
                    WHERE z.zlecenie_g_id=?
                """, (zl_g["id"],)).fetchall()
                zapotrz_list = [dict(r) for r in zapotrz]

                total_p = len(zapotrz_list)
                done_p  = sum(1 for z in zapotrz_list if z.get("p_status") == "zakonczone")
                wg_dict["zapotrzebowania_p"] = zapotrz_list
                wg_dict["p_total"] = total_p
                wg_dict["p_done"]  = done_p
                wg_dict["p_postep_pct"] = round(done_p / total_p * 100) if total_p > 0 else 0
            else:
                wg_dict["zapotrzebowania_p"] = []
                wg_dict["p_total"] = 0
                wg_dict["p_done"] = 0
                wg_dict["p_postep_pct"] = 0

            result.append(wg_dict)
        return result


# ─── Statystyki dla majstra ──────────────────────────────────────────────────
@app.get("/api/stats/majster", dependencies=[Depends(verify_key)])
def majster_stats():
    with get_db() as conn:
        aktywne = conn.execute("""
            SELECT s.start_time, s.pauzy, s.id as sesja_id, s.typ, s.uwagi,
                   u.full_name, u.id as user_id,
                   o.nazwa as op_nazwa, o.stanowisko, o.czas_norma,
                   o.ilosc_wykonana, o.id as op_id,
                   COALESCE(z.numer, zi.numer) as zl_numer,
                   COALESCE(z.nazwa, zi.nazwa) as zl_nazwa,
                   COALESCE(z.ilosc_sztuk, zi.ilosc_sztuk) as ilosc_sztuk
            FROM sesje_pracy s
            JOIN users u ON s.user_id = u.id
            LEFT JOIN operacje o ON s.operacja_id = o.id
            LEFT JOIN zlecenia z ON o.zlecenie_id = z.id
            LEFT JOIN zlecenia zi ON s.zlecenie_id_inne = zi.id
            WHERE s.status='aktywna'
            ORDER BY s.start_time
        """).fetchall()

        dzis = conn.execute("""
            SELECT COUNT(*) as cnt, COALESCE(SUM(s.ilosc_sztuk),0) as sztuki,
                   COALESCE(SUM(
                     (strftime('%s', COALESCE(s.end_time, datetime('now'))) -
                       strftime('%s', s.start_time)) / 3600.0
                   ), 0) as godz
            FROM sesje_pracy s
            WHERE s.status='zakonczona' AND date(s.end_time) = date('now')
              AND s.typ = 'operacja'
        """).fetchone()

        zlecenia = conn.execute("""
            SELECT z.id, z.numer, z.nazwa, z.status, z.ilosc_sztuk, z.termin,
                   z.cena_brutto_szt,
                   COUNT(o.id) as op_total,
                   SUM(CASE WHEN o.status='zakonczona' THEN 1 ELSE 0 END) as op_done,
                   MIN(o.ilosc_wykonana) as sztuki_wykonane
            FROM zlecenia z
            LEFT JOIN operacje o ON o.zlecenie_id = z.id
            WHERE z.status IN ('nowe','w_toku')
            GROUP BY z.id
            ORDER BY z.numer
        """).fetchall()

        wszystkie_zlecenia = conn.execute("""
            SELECT z.id, z.numer, z.nazwa, z.status, z.ilosc_sztuk, z.termin,
                   z.cena_brutto_szt,
                   COUNT(o.id) as op_total,
                   SUM(CASE WHEN o.status='zakonczona' THEN 1 ELSE 0 END) as op_done,
                   MIN(o.ilosc_wykonana) as sztuki_wykonane
            FROM zlecenia z
            LEFT JOIN operacje o ON o.zlecenie_id = z.id
            GROUP BY z.id
            ORDER BY z.id DESC
        """).fetchall()

        alerty = conn.execute("""
            SELECT s.id as sesja_id, u.full_name, o.nazwa as op_nazwa,
                   o.czas_norma, s.start_time, s.pauzy,
                   z.numer as zl_numer, z.ilosc_sztuk
            FROM sesje_pracy s
            JOIN users u ON s.user_id = u.id
            JOIN operacje o ON s.operacja_id = o.id
            JOIN zlecenia z ON o.zlecenie_id = z.id
            WHERE s.status='aktywna' AND o.czas_norma > 0
        """).fetchall()
        alert_list = []
        for a in alerty:
            elapsed_min = (_dt.datetime.utcnow() - _parse(a["start_time"])).total_seconds() / 60
            pauzy = json.loads(a["pauzy"] or "[]")
            for p in pauzy:
                if p.get("koniec"):
                    pause_sec = (_parse(p["koniec"]) - _parse(p["start"])).total_seconds()
                    elapsed_min -= pause_sec / 60
            ilosc_sztuk = max(1, a["ilosc_sztuk"] or 1)
            norma_calkowita = a["czas_norma"] * ilosc_sztuk
            if elapsed_min > norma_calkowita:
                alert_list.append({
                    "sesja_id": a["sesja_id"],
                    "pracownik": a["full_name"],
                    "operacja": a["op_nazwa"],
                    "zlecenie": a["zl_numer"],
                    "norma_min": a["czas_norma"],
                    "norma_calkowita_min": round(norma_calkowita, 1),
                    "ilosc_sztuk": ilosc_sztuk,
                    "elapsed_min": round(elapsed_min, 1),
                    "przekroczenie_pct": round((elapsed_min / norma_calkowita - 1) * 100)
                })

        # ➤ Używamy COALESCE(stawki_zlecen, stawki globalne)
        koszty = conn.execute("""
            SELECT s.user_id, u.full_name, o.stanowisko, o.zlecenie_id,
                   SUM(CASE WHEN s.typ='operacja' THEN
                     (strftime('%s', COALESCE(s.end_time, datetime('now'))) -
                       strftime('%s', s.start_time)) / 3600.0
                       * COALESCE(st_zl.stawka_godz, st.stawka_godz, 0)
                   ELSE
                     (strftime('%s', COALESCE(s.end_time, datetime('now'))) -
                       strftime('%s', s.start_time)) / 3600.0
                       * COALESCE(st_zl.zbrojenie_stawka_godz, st.zbrojenie_stawka_godz, 0)
                   END) as koszt,
                   SUM((strftime('%s', COALESCE(s.end_time, datetime('now'))) -
                        strftime('%s', s.start_time)) / 3600.0) as godz
            FROM sesje_pracy s
            JOIN users u ON s.user_id = u.id
            LEFT JOIN operacje o ON s.operacja_id = o.id
            LEFT JOIN stawki_zlecen st_zl ON st_zl.zlecenie_id=o.zlecenie_id AND st_zl.stanowisko=o.stanowisko
            LEFT JOIN stawki st ON o.stanowisko = st.stanowisko
            WHERE date(s.start_time) = date('now') AND s.typ IN ('operacja','zbrojenie')
            GROUP BY s.user_id, o.stanowisko
        """).fetchall()

        return {
            "aktywne_sesje": [dict(r) for r in aktywne],
            "dzis_sesji": dzis[0] if dzis else 0,
            "dzis_sztuk": dzis[1] or 0,
            "dzis_godz": round(dzis[2] or 0, 2),
            "zlecenia": [dict(r) for r in zlecenia],
            "wszystkie_zlecenia": [dict(r) for r in wszystkie_zlecenia],
            "alerty_norm": alert_list,
            "koszty_dzis": [dict(r) for r in koszty],
        }

# ─── Wydajność pracowników ────────────────────────────────────────────────────
@app.get("/api/stats/wydajnosc", dependencies=[Depends(verify_key)])
def stats_wydajnosc(okres: str = "dzis"):
    if okres == "tydzien":
        filter_sql = "s.end_time >= datetime('now', '-7 days')"
    elif okres == "miesiac":
        filter_sql = "s.end_time >= datetime('now', '-30 days')"
    else:
        filter_sql = "date(s.end_time) = date('now')"

    with get_db() as conn:
        users_rows = conn.execute(f"""
            SELECT u.id, u.full_name,
                   COUNT(s.id) as sesji,
                   COALESCE(SUM(s.ilosc_sztuk), 0) as sztuki,
                   COALESCE(SUM(
                     (strftime('%s', s.end_time) - strftime('%s', s.start_time)) / 60.0
                   ), 0) as min_total
            FROM sesje_pracy s
            JOIN users u ON s.user_id = u.id
            WHERE s.status='zakonczona' AND s.typ IN ('operacja','inne_zlecenie') AND {filter_sql}
            GROUP BY u.id ORDER BY sztuki DESC
        """).fetchall()

        wyniki = []
        for r in users_rows:
            # ➤ Używamy COALESCE(stawki_zlecen, stawki globalne)
            sesje = conn.execute(f"""
                SELECT s.id, s.ilosc_sztuk, s.start_time, s.end_time, s.pauzy,
                       o.nazwa as op_nazwa, o.czas_norma, o.czas_zbrojenia_min, o.stanowisko,
                       z.numer as zl_numer, z.nazwa as zl_nazwa, s.typ,
                       COALESCE(st_zl.stawka_godz, st.stawka_godz, 0) as stawka_godz,
                       COALESCE(st_zl.zbrojenie_stawka_godz, st.zbrojenie_stawka_godz, 0) as zbrojenie_stawka_godz
                FROM sesje_pracy s
                LEFT JOIN operacje o ON s.operacja_id = o.id
                LEFT JOIN zlecenia z ON o.zlecenie_id = z.id
                LEFT JOIN stawki_zlecen st_zl ON st_zl.zlecenie_id=o.zlecenie_id AND st_zl.stanowisko=o.stanowisko
                LEFT JOIN stawki st ON o.stanowisko = st.stanowisko
                WHERE s.user_id=? AND s.status='zakonczona' AND s.typ IN ('operacja','zbrojenie')
                  AND s.start_time IS NOT NULL AND s.end_time IS NOT NULL
                  AND {filter_sql}
                ORDER BY s.end_time DESC LIMIT 30
            """, (r["id"],)).fetchall()

            sesje_list = []
            normy_ok = 0
            normy_total = 0
            koszt_pracy = 0.0; koszt_zbrojenia = 0.0
            for s in sesje:
                try:
                    elapsed = (_parse(s["end_time"]) - _parse(s["start_time"])).total_seconds() / 60
                except Exception:
                    continue
                pauzy = json.loads(s["pauzy"] or "[]")
                for p in pauzy:
                    if p.get("koniec"):
                        elapsed -= (_parse(p["koniec"]) - _parse(p["start"])).total_seconds() / 60
                elapsed = max(0.1, elapsed)
                czas_norma = s["czas_norma"]
                ilosc = s["ilosc_sztuk"] or 1
                czas_zbrojenia_min = s["czas_zbrojenia_min"] or 0
                if s["typ"] == "operacja" and czas_norma:
                    wyd_pct = round(czas_norma * ilosc / elapsed * 100)
                elif s["typ"] == "zbrojenie" and czas_zbrojenia_min:
                    wyd_pct = round(czas_zbrojenia_min / elapsed * 100)
                else:
                    wyd_pct = None
                if wyd_pct is not None:
                    normy_total += 1
                    if wyd_pct >= 90:
                        normy_ok += 1
                if s["typ"] == "operacja":
                    koszt_pracy += (elapsed / 60.0) * float(s["stawka_godz"] or 0)
                elif s["typ"] == "zbrojenie":
                    koszt_zbrojenia += (elapsed / 60.0) * float(s["zbrojenie_stawka_godz"] or 0)
                sesje_list.append({
                    "sesja_id": s["id"],
                    "op_nazwa": s["op_nazwa"],
                    "stanowisko": s["stanowisko"],
                    "zl_numer": s["zl_numer"],
                    "zl_nazwa": (dict(s).get("zl_nazwa") or s["zl_numer"]),
                    "ilosc_sztuk": s["ilosc_sztuk"],
                    "czas_min": round(elapsed, 1),
                    "norma_min": czas_norma,
                    "wyd_pct": wyd_pct,
                    "data": (s["end_time"] or "")[:10],
                    "start_time": s["start_time"],
                    "end_time": s["end_time"],
                    "typ": s["typ"],
                })

            wyniki.append({
                "user_id": r["id"],
                "full_name": r["full_name"],
                "sesji": r["sesji"],
                "sztuki": r["sztuki"],
                "godz": round(r["min_total"] / 60, 2),
                "normy_ok": normy_ok,
                "normy_total": normy_total,
                "koszt_pracy": round(koszt_pracy, 2),
                "koszt_zbrojenia": round(koszt_zbrojenia, 2),
                "koszt_total": round(koszt_pracy + koszt_zbrojenia, 2),
                "sesje": sesje_list,
            })

        return {"okres": okres, "pracownicy": wyniki}

# ─── Wydajność jednego pracownika ─────────────────────────────────────────────
@app.get("/api/stats/wydajnosc/{user_id}", dependencies=[Depends(verify_key)])
def stats_wydajnosc_user(user_id: int, okres: str = "tydzien"):
    if okres == "dzis":
        filter_sql = "date(s.end_time) = date('now')"
    elif okres == "miesiac":
        filter_sql = "s.end_time >= datetime('now', '-30 days')"
    else:
        filter_sql = "s.end_time >= datetime('now', '-7 days')"

    with get_db() as conn:
        summary = conn.execute(f"""
            SELECT COUNT(s.id) as sesji,
                   COALESCE(SUM(s.ilosc_sztuk), 0) as sztuki,
                   COALESCE(SUM(
                     (strftime('%s', s.end_time) - strftime('%s', s.start_time)) / 60.0
                   ), 0) as min_total
            FROM sesje_pracy s
            WHERE s.user_id=? AND s.status='zakonczona' AND s.typ IN ('operacja','inne_zlecenie') AND {filter_sql}
        """, (user_id,)).fetchone()

        # ➤ Używamy COALESCE(stawki_zlecen, stawki globalne)
        sesje = conn.execute(f"""
            SELECT s.id, s.ilosc_sztuk, s.start_time, s.end_time, s.pauzy, s.typ,
                   o.nazwa as op_nazwa, o.czas_norma, o.czas_zbrojenia_min, o.stanowisko,
                   z.numer as zl_numer, z.nazwa as zl_nazwa,
                   COALESCE(st_zl.stawka_godz, st.stawka_godz, 0) as stawka_godz,
                   COALESCE(st_zl.zbrojenie_stawka_godz, st.zbrojenie_stawka_godz, 0) as zbrojenie_stawka_godz
            FROM sesje_pracy s
            LEFT JOIN operacje o ON s.operacja_id = o.id
            LEFT JOIN zlecenia z ON o.zlecenie_id = z.id
            LEFT JOIN stawki_zlecen st_zl ON st_zl.zlecenie_id=o.zlecenie_id AND st_zl.stanowisko=o.stanowisko
            LEFT JOIN stawki st ON o.stanowisko = st.stanowisko
            WHERE s.user_id=? AND s.status='zakonczona' AND s.typ IN ('operacja','zbrojenie')
              AND s.start_time IS NOT NULL AND s.end_time IS NOT NULL
              AND {filter_sql}
            ORDER BY s.end_time DESC
        """, (user_id,)).fetchall()

        sesje_list = []
        normy_ok = 0
        normy_total = 0
        koszt_pracy = 0.0; koszt_zbrojenia = 0.0
        for s in sesje:
            try:
                elapsed = (_parse(s["end_time"]) - _parse(s["start_time"])).total_seconds() / 60
            except Exception:
                continue
            pauzy = json.loads(s["pauzy"] or "[]")
            for p in pauzy:
                if p.get("koniec"):
                    elapsed -= (_parse(p["koniec"]) - _parse(p["start"])).total_seconds() / 60
            elapsed = max(0.1, elapsed)
            czas_norma = s["czas_norma"]
            ilosc = s["ilosc_sztuk"] or 1
            czas_zbrojenia_min = s["czas_zbrojenia_min"] or 0
            if s["typ"] == "operacja" and czas_norma:
                wyd_pct = round(czas_norma * ilosc / elapsed * 100)
            elif s["typ"] == "zbrojenie" and czas_zbrojenia_min:
                wyd_pct = round(czas_zbrojenia_min / elapsed * 100)
            else:
                wyd_pct = None
            if wyd_pct is not None:
                normy_total += 1
                if wyd_pct >= 90:
                    normy_ok += 1
            if s["typ"] == "operacja":
                koszt_pracy += (elapsed / 60.0) * float(s["stawka_godz"] or 0)
            elif s["typ"] == "zbrojenie":
                koszt_zbrojenia += (elapsed / 60.0) * float(s["zbrojenie_stawka_godz"] or 0)
            sesje_list.append({
                "sesja_id": s["id"],
                "op_nazwa": s["op_nazwa"],
                "stanowisko": s["stanowisko"],
                "zl_numer": s["zl_numer"],
                "zl_nazwa": s["zl_nazwa"],
                "ilosc_sztuk": s["ilosc_sztuk"],
                "czas_min": round(elapsed, 1),
                "norma_min": czas_norma if s["typ"] == "operacja" else czas_zbrojenia_min or None,
                "wyd_pct": wyd_pct,
                "typ": s["typ"],
                "data": (s["end_time"] or "")[:10],
                "start_time": s["start_time"],
                "end_time": s["end_time"],
            })

        return {
            "okres": okres,
            "sesji": summary["sesji"] if summary else 0,
            "sztuki": summary["sztuki"] if summary else 0,
            "godz": round((summary["min_total"] or 0) / 60, 2),
            "normy_ok": normy_ok,
            "normy_total": normy_total,
            "koszt_pracy": round(koszt_pracy, 2),
            "koszt_zbrojenia": round(koszt_zbrojenia, 2),
            "koszt_total": round(koszt_pracy + koszt_zbrojenia, 2),
            "sesje": sesje_list,
        }

@app.get("/api/zlecenia/{zid}/sesje", dependencies=[Depends(verify_key)])
def get_zlecenie_sesje(zid: int):
    with get_db() as conn:
        rows = conn.execute("""
            SELECT s.id, s.start_time, s.end_time, s.pauzy, s.ilosc_sztuk, s.uwagi, s.typ,
                   u.full_name, o.nazwa as op_nazwa, o.stanowisko, o.kolejnosc
            FROM sesje_pracy s
            JOIN users u ON s.user_id=u.id
            LEFT JOIN operacje o ON s.operacja_id=o.id
            WHERE o.zlecenie_id=? AND s.status='zakonczona'
            ORDER BY s.end_time DESC
        """, (zid,)).fetchall()
        return [dict(r) for r in rows]

@app.get("/api/zlecenia/{zid}/szczegoly", dependencies=[Depends(verify_key)])
def get_zlecenie_szczegoly(zid: int):
    with get_db() as conn:
        zlecenie = conn.execute("""
            SELECT z.*, (z.cena_brutto_szt * z.ilosc_sztuk) as wartosc_total
            FROM zlecenia z WHERE z.id=?
        """, (zid,)).fetchone()
        if not zlecenie:
            raise HTTPException(404, "Zlecenie nie znaleziono")

        # ➤ Używamy COALESCE(stawki_zlecen, stawki globalne)
        sesje = conn.execute("""
            SELECT s.id, s.start_time, s.end_time, s.pauzy, s.ilosc_sztuk, s.uwagi, s.typ,
                   u.full_name,
                   COALESCE(st_zl.stawka_godz, st.stawka_godz, 0) as stawka_godz,
                   COALESCE(o.nazwa, s.uwagi) as op_nazwa,
                   COALESCE(o.kolejnosc, 999) as kolejnosc,
                   COALESCE(o.stanowisko, '') as stanowisko,
                   COALESCE(o.czas_norma, 0) as czas_norma,
                   z2.numer as zl_numer
            FROM sesje_pracy s
            JOIN users u ON s.user_id=u.id
            LEFT JOIN operacje o ON s.operacja_id=o.id
            LEFT JOIN zlecenia z2 ON o.zlecenie_id=z2.id
            LEFT JOIN stawki_zlecen st_zl ON st_zl.zlecenie_id=z2.id AND st_zl.stanowisko=o.stanowisko
            LEFT JOIN stawki st ON o.stanowisko=st.stanowisko
            WHERE (o.zlecenie_id=? OR s.zlecenie_id_inne=?) AND s.status='zakonczona'
            ORDER BY COALESCE(o.kolejnosc,999), s.end_time
        """, (zid, zid)).fetchall()

        result_sesje = []
        koszt_total = 0.0
        for s in sesje:
            sd = dict(s)
            elapsed_sec = 0.0
            if sd['start_time'] and sd['end_time']:
                elapsed_sec = (_parse(sd['end_time']) - _parse(sd['start_time'])).total_seconds()
                for p in json.loads(sd['pauzy'] or '[]'):
                    if p.get('koniec'):
                        elapsed_sec -= (_parse(p['koniec']) - _parse(p['start'])).total_seconds()
                elapsed_sec = max(0, elapsed_sec)
            stawka = sd.get('stawka_godz') or 0
            koszt = (elapsed_sec / 3600.0) * float(stawka)
            sd['koszt_sesji'] = round(koszt, 2)
            sd['elapsed_sec'] = round(elapsed_sec, 1)
            koszt_total += koszt
            result_sesje.append(sd)

        produkty = conn.execute(
            "SELECT * FROM produkty_zlecenia WHERE zlecenie_id=? ORDER BY id", (zid,)
        ).fetchall()
        produkty_list = [dict(p) for p in produkty]
        koszt_produktow = sum(float(p['ilosc']) * float(p['cena']) for p in produkty_list)

        return {
            "sesje": result_sesje,
            "koszt_total": round(koszt_total, 2),
            "koszt_produktow": round(koszt_produktow, 2),
            "produkty": produkty_list,
            "wartosc": float(zlecenie['wartosc_total'] or 0),
        }

# ─── Podsumowanie kosztów zlecenia ────────────────────────────────────────────
@app.get("/api/zlecenia/{zid}/koszty", dependencies=[Depends(verify_key)])
def koszty_zlecenia(zid: int):
    with get_db() as conn:
        zl = conn.execute("SELECT * FROM zlecenia WHERE id=?", (zid,)).fetchone()
        if not zl:
            raise HTTPException(404, "Zlecenie nie znalezione")

        # ➤ Używamy COALESCE(stawki_zlecen, stawki globalne)
        sesje = conn.execute("""
            SELECT s.*, u.full_name, o.stanowisko, o.nazwa as op_nazwa, o.czas_norma,
                   COALESCE(st_zl.stawka_godz, st.stawka_godz) as stawka_godz
            FROM sesje_pracy s
            JOIN users u ON s.user_id = u.id
            JOIN operacje o ON s.operacja_id = o.id
            LEFT JOIN stawki_zlecen st_zl ON st_zl.zlecenie_id=o.zlecenie_id AND st_zl.stanowisko=o.stanowisko
            LEFT JOIN stawki st ON o.stanowisko = st.stanowisko
            WHERE o.zlecenie_id=? AND s.status='zakonczona'
        """, (zid,)).fetchall()

        total_koszt = 0
        total_godz = 0
        rows = []

        # ➤ Zbrojenie: priorytet stawki_zlecen
        sesje_zbrojenie = conn.execute("""
            SELECT s.start_time, s.end_time, s.pauzy,
                   COALESCE(o.nazwa,'—') as op_nazwa,
                   COALESCE(st_zl.zbrojenie_stawka_godz, st.zbrojenie_stawka_godz, 0) as zbr_stawka
            FROM sesje_pracy s
            LEFT JOIN operacje o ON s.operacja_id = o.id
            LEFT JOIN stawki_zlecen st_zl ON st_zl.zlecenie_id=o.zlecenie_id AND st_zl.stanowisko=o.stanowisko
            LEFT JOIN stawki st ON o.stanowisko = st.stanowisko
            WHERE o.zlecenie_id=? AND s.typ='zbrojenie' AND s.status='zakonczona'
        """, (zid,)).fetchall()

        koszt_zbrojenia_total = 0.0
        zbrojenia_info = []
        for sz in sesje_zbrojenie:
            if not sz["end_time"]: continue
            elapsed_zb = (_parse(sz["end_time"]) - _parse(sz["start_time"])).total_seconds() / 3600
            for p in json.loads(sz["pauzy"] or "[]"):
                if p.get("koniec"):
                    elapsed_zb -= (_parse(p["koniec"]) - _parse(p["start"])).total_seconds() / 3600
            elapsed_zb = max(0, elapsed_zb)
            koszt_zb = elapsed_zb * (sz["zbr_stawka"] or 0)
            koszt_zbrojenia_total += koszt_zb
            zbrojenia_info.append({
                "op_nazwa": sz["op_nazwa"],
                "czas_min": round(elapsed_zb * 60, 1),
                "stawka_godz": sz["zbr_stawka"],
                "koszt": round(koszt_zb, 2)
            })

        for s in sesje:
            elapsed = (_parse(s["end_time"]) - _parse(s["start_time"])).total_seconds() / 3600
            pauzy = json.loads(s["pauzy"] or "[]")
            for p in pauzy:
                if p.get("koniec"):
                    pause_sec = (_parse(p["koniec"]) - _parse(p["start"])).total_seconds()
                    elapsed -= pause_sec / 3600
            koszt = elapsed * (s["stawka_godz"] or 0)
            total_koszt += koszt
            total_godz += elapsed
            rows.append({
                "sesja_id": s["id"],
                "pracownik": s["full_name"],
                "operacja": s["op_nazwa"],
                "stanowisko": s["stanowisko"],
                "godz": round(elapsed, 2),
                "stawka_godz": s["stawka_godz"],
                "koszt": round(koszt, 2),
                "ilosc_sztuk": s["ilosc_sztuk"],
            })

        przychod = (zl["cena_brutto_szt"] or 0) * (zl["ilosc_sztuk"] or 0)
        produkty = conn.execute(
            "SELECT * FROM produkty_zlecenia WHERE zlecenie_id=? ORDER BY id", (zid,)
        ).fetchall()
        produkty_list = [dict(p) for p in produkty]
        koszt_produktow = sum(float(p['ilosc']) * float(p['cena']) for p in produkty_list)
        total_koszty = total_koszt + koszt_produktow
        total_koszty_z_zbrojeniem = total_koszt + koszt_produktow + koszt_zbrojenia_total
        return {
            "zlecenie": dict(zl),
            "sesje": rows,
            "produkty": produkty_list,
            "total_godz": round(total_godz, 2),
            "total_koszt": round(total_koszt, 2),
            "koszt_produktow": round(koszt_produktow, 2),
            "koszt_zbrojenia": round(koszt_zbrojenia_total, 2),
            "zbrojenia": zbrojenia_info,
            "total_koszty": round(total_koszty_z_zbrojeniem, 2),
            "przychod": round(przychod, 2),
            "marza": round(przychod - total_koszty_z_zbrojeniem, 2),
        }

# ─── Powiadomienia ────────────────────────────────────────────────────────────
@app.get("/api/powiadomienia/{rola}", dependencies=[Depends(verify_key)])
def get_powiadomienia(rola: str):
    with get_db() as conn:
        rows = conn.execute("""
            SELECT p.*, o.nazwa as op_nazwa, z.numer as zl_numer
            FROM powiadomienia p
            LEFT JOIN operacje o ON p.operacja_id = o.id
            LEFT JOIN zlecenia z ON p.zlecenie_id = z.id
            WHERE (p.dla_roli=? OR p.dla_roli='all') AND p.odczytane=0
            ORDER BY p.created_at DESC LIMIT 50
        """, (rola,)).fetchall()
        return [dict(r) for r in rows]

@app.post("/api/powiadomienia/{pid}/przeczytaj", dependencies=[Depends(verify_key)])
def mark_read(pid: int):
    with get_db() as conn:
        conn.execute("UPDATE powiadomienia SET odczytane=1 WHERE id=?", (pid,))
        return {"ok": True}

# ─── Produkty zlecenia ────────────────────────────────────────────────────────
class ProduktZleceniaRequest(BaseModel):
    zlecenie_id: int
    nazwa: str
    ilosc: float = 1
    cena: float = 0

@app.get("/api/zlecenia/{zid}/produkty", dependencies=[Depends(verify_key)])
def get_produkty_zlecenia(zid: int):
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM produkty_zlecenia WHERE zlecenie_id=? ORDER BY id", (zid,)
        ).fetchall()
        return [dict(r) for r in rows]

@app.post("/api/zlecenia/{zid}/produkty", dependencies=[Depends(verify_key)])
def add_produkt_zlecenia(zid: int, req: ProduktZleceniaRequest):
    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO produkty_zlecenia (zlecenie_id, nazwa, ilosc, cena) VALUES (?,?,?,?)",
            (zid, req.nazwa, req.ilosc, req.cena)
        )
        return {"id": cur.lastrowid}

@app.put("/api/zlecenia/{zid}/produkty/{pid}", dependencies=[Depends(verify_key)])
def update_produkt_zlecenia(zid: int, pid: int, req: ProduktZleceniaRequest):
    with get_db() as conn:
        conn.execute(
            "UPDATE produkty_zlecenia SET nazwa=?, ilosc=?, cena=? WHERE id=? AND zlecenie_id=?",
            (req.nazwa, req.ilosc, req.cena, pid, zid)
        )
        return {"ok": True}

@app.delete("/api/zlecenia/{zid}/produkty/{pid}", dependencies=[Depends(verify_key)])
def delete_produkt_zlecenia(zid: int, pid: int):
    with get_db() as conn:
        conn.execute("DELETE FROM produkty_zlecenia WHERE id=? AND zlecenie_id=?", (pid, zid))
        return {"ok": True}

# ─── Wydajność majstra – raport z zakresu dat ────────────────────────────────
@app.get("/api/stats/wydajnosc_raport", dependencies=[Depends(verify_key)])
def stats_wydajnosc_raport(data_od: str = "", data_do: str = ""):
    if not data_od:
        data_od = (_dt.datetime.utcnow() - _dt.timedelta(days=30)).strftime("%Y-%m-%d")
    if not data_do:
        data_do = _dt.datetime.utcnow().strftime("%Y-%m-%d")
    filter_sql = f"date(s.end_time) BETWEEN '{data_od}' AND '{data_do}'"

    with get_db() as conn:
        users_rows = conn.execute(f"""
            SELECT u.id, u.full_name,
                   COUNT(CASE WHEN s.typ IN ('operacja','inne_zlecenie') THEN 1 END) as sesji,
                   COALESCE(SUM(CASE WHEN s.typ IN ('operacja','inne_zlecenie') THEN s.ilosc_sztuk ELSE 0 END),0) as sztuki,
                   COALESCE(SUM(CASE WHEN s.typ IN ('operacja','inne_zlecenie') THEN
                     (strftime('%s', s.end_time) - strftime('%s', s.start_time)) / 60.0 ELSE 0 END), 0) as min_roboczy,
                   COALESCE(SUM(CASE WHEN s.typ='zbrojenie' THEN
                     (strftime('%s', s.end_time) - strftime('%s', s.start_time)) / 60.0 ELSE 0 END), 0) as min_zbrojenie,
                   COALESCE(SUM(CASE WHEN s.typ='nieprodukcyjna' THEN
                     (strftime('%s', s.end_time) - strftime('%s', s.start_time)) / 60.0 ELSE 0 END), 0) as min_nieproduktywny,
                   COUNT(DISTINCT date(s.start_time)) as dni_pracy
            FROM sesje_pracy s
            JOIN users u ON s.user_id = u.id
            WHERE s.status='zakonczona' AND {filter_sql}
            GROUP BY u.id ORDER BY min_roboczy DESC
        """).fetchall()

        wyniki = []
        for r in users_rows:
            # ➤ Używamy COALESCE(stawki_zlecen, stawki globalne)
            sesje = conn.execute(f"""
                SELECT s.ilosc_sztuk, s.start_time, s.end_time, s.pauzy, s.typ,
                       s.uwagi, s.sesja_glowna,
                       o.nazwa as op_nazwa, o.czas_norma, o.stanowisko, o.zlecenie_id,
                       z.numer as zl_numer,
                       zi.numer as zl_inne_numer,
                       COALESCE(st_zl_rz.stawka_godz, st_rz.stawka_godz, st_zl.stawka_godz, st.stawka_godz, 0) as stawka_godz,
                       COALESCE(st_zl_rz.zbrojenie_stawka_godz, st_rz.zbrojenie_stawka_godz, st_zl.zbrojenie_stawka_godz, st.zbrojenie_stawka_godz, 0) as zbrojenie_stawka_godz,
                       COALESCE(s.rzeczywiste_stanowisko, o.stanowisko) as efektywne_stanowisko
                FROM sesje_pracy s
                LEFT JOIN operacje o ON s.operacja_id = o.id
                LEFT JOIN zlecenia z ON o.zlecenie_id = z.id
                LEFT JOIN zlecenia zi ON s.zlecenie_id_inne = zi.id
                LEFT JOIN stawki_zlecen st_zl ON st_zl.zlecenie_id=o.zlecenie_id AND st_zl.stanowisko=o.stanowisko
                LEFT JOIN stawki st ON o.stanowisko = st.stanowisko
                LEFT JOIN stawki_zlecen st_zl_rz ON s.rzeczywiste_stanowisko IS NOT NULL AND st_zl_rz.zlecenie_id=o.zlecenie_id AND st_zl_rz.stanowisko=s.rzeczywiste_stanowisko
                LEFT JOIN stawki st_rz ON s.rzeczywiste_stanowisko IS NOT NULL AND st_rz.stanowisko=s.rzeczywiste_stanowisko
                WHERE s.user_id=? AND s.status='zakonczona' AND {filter_sql}
                ORDER BY s.end_time DESC
            """, (r["id"],)).fetchall()

            sesje_list = []
            normy_ok = 0; normy_total = 0
            koszt_pracy = 0.0; koszt_zbrojenia = 0.0
            suma_norma_min = 0.0
            suma_fakty_min = 0.0
            for s in sesje:
                elapsed = (_parse(s["end_time"]) - _parse(s["start_time"])).total_seconds() / 60
                pauzy = json.loads(s["pauzy"] or "[]")
                for p in pauzy:
                    if p.get("koniec"):
                        elapsed -= (_parse(p["koniec"]) - _parse(p["start"])).total_seconds() / 60
                elapsed = max(0.1, elapsed)
                czas_norma = s["czas_norma"]
                ilosc = s["ilosc_sztuk"] or 1
                sesja_glowna = s["sesja_glowna"] if s["sesja_glowna"] is not None else 1
                wyd_pct = round(czas_norma * ilosc / elapsed * 100) if czas_norma else None
                if s["typ"] in ("operacja", "inne_zlecenie"):
                    if sesja_glowna == 1 and czas_norma and czas_norma > 0:
                        normy_total += 1
                        suma_norma_min += czas_norma * ilosc
                        suma_fakty_min += elapsed
                        if wyd_pct is not None and wyd_pct >= 90:
                            normy_ok += 1
                if s["typ"] in ("operacja", "inne_zlecenie"):
                    if sesja_glowna == 1:
                        koszt_pracy += (elapsed / 60.0) * float(s["stawka_godz"] or 0)
                    else:
                        if czas_norma and czas_norma > 0:
                            koszt_pracy += (czas_norma * ilosc / 60.0) * float(s["stawka_godz"] or 0)
                        else:
                            koszt_pracy += (elapsed / 60.0) * float(s["stawka_godz"] or 0)
                elif s["typ"] == "zbrojenie":
                    koszt_zbrojenia += (elapsed / 60.0) * float(s["zbrojenie_stawka_godz"] or 0)

                typ = s["typ"]
                if typ == "nieprodukcyjna":
                    display_op = s["uwagi"] or s["op_nazwa"] or "—"
                    display_zl = ""
                elif typ == "inne_zlecenie":
                    display_op = s["uwagi"] or s["op_nazwa"] or "—"
                    display_zl = s["zl_inne_numer"] or s["zl_numer"] or "—"
                elif typ == "zbrojenie":
                    display_op = (s["op_nazwa"] + " (zbr.)") if s["op_nazwa"] else "— (zbr.)"
                    display_zl = s["zl_numer"] or "—"
                else:
                    display_op = s["op_nazwa"] or "—"
                    display_zl = s["zl_numer"] or "—"

                sesje_list.append({
                    "op_nazwa": display_op,
                    "stanowisko": s["stanowisko"] or "—",
                    "zl_numer": display_zl,
                    "ilosc_sztuk": s["ilosc_sztuk"],
                    "czas_min": round(elapsed, 1),
                    "norma_min": czas_norma,
                    "wyd_pct": wyd_pct,
                    "typ": s["typ"],
                    "data": (s["end_time"] or "")[:10],
                    "sesja_glowna": sesja_glowna,
                })

            norma_wydajnosc_pct = round(suma_norma_min / suma_fakty_min * 100) if suma_fakty_min > 0 else None

            wyniki.append({
                "user_id": r["id"],
                "full_name": r["full_name"],
                "sesji": r["sesji"],
                "sztuki": r["sztuki"],
                "min_roboczy": round(r["min_roboczy"], 1),
                "min_zbrojenie": round(r["min_zbrojenie"], 1),
                "min_nieproduktywny": round(r["min_nieproduktywny"], 1),
                "dni_pracy": r["dni_pracy"],
                "normy_ok": normy_ok,
                "normy_total": normy_total,
                "norma_wydajnosc_pct": norma_wydajnosc_pct,
                "suma_norma_min": round(suma_norma_min, 1),
                "suma_fakty_min": round(suma_fakty_min, 1),
                "koszt_pracy": round(koszt_pracy, 2),
                "koszt_zbrojenia": round(koszt_zbrojenia, 2),
                "koszt_total": round(koszt_pracy + koszt_zbrojenia, 2),
                "sesje": sesje_list,
            })

        return {"data_od": data_od, "data_do": data_do, "pracownicy": wyniki}

# ─── Raport zleceń PDF-data ──────────────────────────────────────────────────
@app.get("/api/raporty/zlecenia", dependencies=[Depends(verify_key)])
def raport_zlecenia(data_od: str = "", data_do: str = "", zlecenie_id: Optional[int] = None):
    if not data_od:
        data_od = (_dt.datetime.utcnow() - _dt.timedelta(days=30)).strftime("%Y-%m-%d")
    if not data_do:
        data_do = _dt.datetime.utcnow().strftime("%Y-%m-%d")

    with get_db() as conn:
        if zlecenie_id:
            zlecenia = conn.execute("""
                SELECT z.*,
                       COUNT(DISTINCT o.id) as op_total,
                       COALESCE(SUM(o.ilosc_wykonana),0) as sztuki_done
                FROM zlecenia z
                LEFT JOIN operacje o ON o.zlecenie_id=z.id
                WHERE z.id=?
                GROUP BY z.id
            """, (zlecenie_id,)).fetchall()
        else:
            zlecenia = conn.execute("""
                SELECT z.*,
                       COUNT(DISTINCT o.id) as op_total,
                       COALESCE(SUM(o.ilosc_wykonana),0) as sztuki_done
                FROM zlecenia z
                LEFT JOIN operacje o ON o.zlecenie_id=z.id
                WHERE date(z.created_at) BETWEEN ? AND ?
                GROUP BY z.id ORDER BY z.id DESC
            """, (data_od, data_do)).fetchall()

        result = []
        for z in zlecenia:
            zid = z["id"]
            # ➤ Używamy COALESCE(stawki_zlecen, stawki globalne)
            sesje = conn.execute("""
                SELECT s.start_time, s.end_time, s.pauzy, s.ilosc_sztuk, s.uwagi, s.typ,
                       u.full_name,
                       COALESCE(o.nazwa, s.uwagi) as op_nazwa,
                       COALESCE(o.kolejnosc, 999) as kolejnosc,
                       COALESCE(o.stanowisko, '') as stanowisko,
                       COALESCE(o.czas_norma, 0) as czas_norma,
                       COALESCE(st_zl.stawka_godz, st.stawka_godz, 0) as stawka_godz,
                       COALESCE(st_zl.zbrojenie_stawka_godz, st.zbrojenie_stawka_godz, 0) as zbrojenie_stawka_godz
                FROM sesje_pracy s
                JOIN users u ON s.user_id=u.id
                LEFT JOIN operacje o ON s.operacja_id=o.id
                LEFT JOIN stawki_zlecen st_zl ON st_zl.zlecenie_id=o.zlecenie_id AND st_zl.stanowisko=o.stanowisko
                LEFT JOIN stawki st ON o.stanowisko=st.stanowisko
                WHERE (o.zlecenie_id=? OR s.zlecenie_id_inne=?) AND s.status='zakonczona'
                ORDER BY COALESCE(o.kolejnosc,999), s.end_time
            """, (zid, zid)).fetchall()

            sesje_list = []
            koszt_pracy = 0
            koszt_zbrojenia = 0.0
            for s in sesje:
                elapsed = (_parse(s["end_time"]) - _parse(s["start_time"])).total_seconds() / 3600
                pauzy = json.loads(s["pauzy"] or "[]")
                for p in pauzy:
                    if p.get("koniec"):
                        elapsed -= (_parse(p["koniec"]) - _parse(p["start"])).total_seconds() / 3600
                elapsed = max(0, elapsed)
                if s["typ"] == "zbrojenie":
                    koszt = round(elapsed * float(s["zbrojenie_stawka_godz"] or 0), 2)
                    koszt_zbrojenia += koszt
                else:
                    koszt = round(elapsed * float(s["stawka_godz"] or 0), 2)
                    koszt_pracy += koszt
                sesje_list.append({
                    "pracownik": s["full_name"],
                    "operacja": (s["op_nazwa"] + " (zbr.)" if s["typ"] == "zbrojenie" else s["op_nazwa"]) or "—",
                    "kolejnosc": s["kolejnosc"],
                    "typ": s["typ"],
                    "data": (s["end_time"] or "")[:16].replace("T", "  "),
                    "czas_min": round(elapsed * 60, 1),
                    "ilosc_sztuk": s["ilosc_sztuk"],
                    "czas_norma": s["czas_norma"] or 0,
                    "koszt": koszt,
                    "uwagi": s["uwagi"] or "",
                })

            produkty = conn.execute(
                "SELECT * FROM produkty_zlecenia WHERE zlecenie_id=?", (zid,)
            ).fetchall()
            produkty_list = [dict(p) for p in produkty]
            koszt_produktow = sum(float(p["ilosc"])*float(p["cena"]) for p in produkty_list)
            wartosc = (z["cena_brutto_szt"] or 0) * (z["ilosc_sztuk"] or 0)
            koszt_total = koszt_pracy + koszt_zbrojenia + koszt_produktow
            zysk = wartosc - koszt_total

            result.append({
                "id": zid,
                "numer": z["numer"],
                "nazwa": z["nazwa"],
                "status": z["status"],
                "ilosc_sztuk": z["ilosc_sztuk"],
                "cena_szt": z["cena_brutto_szt"],
                "wartosc": round(wartosc, 2),
                "created_at": (z["created_at"] or "")[:10],
                "sesje": sesje_list,
                "produkty": produkty_list,
                "koszt_pracy": round(koszt_pracy, 2),
                "koszt_zbrojenia": round(koszt_zbrojenia, 2),
                "koszt_produktow": round(koszt_produktow, 2),
                "koszt_total": round(koszt_total, 2),
                "zysk": round(zysk, 2),
            })
        return {"data_od": data_od, "data_do": data_do, "zlecenia": result}

# ─── Inicjalizacja bazy danych ────────────────────────────────────────────────
def init_db_on_start():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL, password TEXT NOT NULL,
        full_name TEXT NOT NULL, role TEXT NOT NULL)""")

    c.execute("""CREATE TABLE IF NOT EXISTS zlecenia (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        numer TEXT UNIQUE NOT NULL, nazwa TEXT NOT NULL, opis TEXT,
        status TEXT DEFAULT 'nowe', created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        created_by INTEGER, termin TEXT, ilosc_sztuk INTEGER DEFAULT 1,
        cena_brutto_szt REAL DEFAULT 0, material_od_klienta INTEGER DEFAULT 0,
        qr_code TEXT)""")

    try: c.execute("ALTER TABLE zlecenia ADD COLUMN model_3d_url TEXT DEFAULT NULL")
    except: pass

    c.execute("""CREATE TABLE IF NOT EXISTS operacje (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        zlecenie_id INTEGER NOT NULL, nazwa TEXT NOT NULL,
        kolejnosc INTEGER DEFAULT 0, czas_norma REAL DEFAULT 0,
        stanowisko TEXT, status TEXT DEFAULT 'oczekuje', qr_code TEXT,
        ilosc_wykonana INTEGER DEFAULT 0, opis_czynnosci TEXT DEFAULT '',
        czas_zbrojenia_min REAL DEFAULT 0.0,
        FOREIGN KEY (zlecenie_id) REFERENCES zlecenia(id))""")
    try: c.execute("ALTER TABLE operacje ADD COLUMN czas_zbrojenia_min REAL DEFAULT 0.0")
    except: pass
    try: c.execute("ALTER TABLE operacje ADD COLUMN typ_operacji TEXT DEFAULT 'produkcja'")
    except: pass
    try: c.execute("ALTER TABLE operacje ADD COLUMN parametry_kj TEXT DEFAULT NULL")
    except: pass
    try: c.execute("ALTER TABLE operacje ADD COLUMN kj_wynik TEXT DEFAULT NULL")
    except: pass
    try: c.execute("ALTER TABLE operacje ADD COLUMN czas_tpz_min REAL DEFAULT 0.0")
    except: pass

    c.execute("""CREATE TABLE IF NOT EXISTS sesje_pracy (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        operacja_id INTEGER, user_id INTEGER NOT NULL,
        typ TEXT NOT NULL, start_time TIMESTAMP, end_time TIMESTAMP,
        pauzy TEXT DEFAULT '[]', ilosc_sztuk INTEGER DEFAULT 0,
        uwagi TEXT, status TEXT DEFAULT 'aktywna',
        zlecenie_id_inne INTEGER DEFAULT NULL,
        FOREIGN KEY (user_id) REFERENCES users(id))""")

    try:
        c.execute("ALTER TABLE sesje_pracy ADD COLUMN zlecenie_id_inne INTEGER DEFAULT NULL")
    except Exception:
        pass

    c.execute("""CREATE TABLE IF NOT EXISTS stawki (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        stanowisko TEXT UNIQUE NOT NULL, stawka_godz REAL NOT NULL,
        czas_norma_min REAL DEFAULT 0, opis TEXT DEFAULT '',
        zbrojenie_aktywne INTEGER DEFAULT 0,
        zbrojenie_stawka_godz REAL DEFAULT 0.0)""")
    for col, typ in [("zbrojenie_aktywne", "INTEGER DEFAULT 0"), ("zbrojenie_stawka_godz", "REAL DEFAULT 0.0"), ("typ_maszyny", "TEXT DEFAULT ''")]:
        try: c.execute(f"ALTER TABLE stawki ADD COLUMN {col} {typ}")
        except: pass

    c.execute("""CREATE TABLE IF NOT EXISTS katalog_produktow (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nazwa TEXT UNIQUE NOT NULL, opis TEXT,
        ilosc_domyslna INTEGER DEFAULT 1, cena_szt REAL DEFAULT 0.0,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")

    c.execute("""CREATE TABLE IF NOT EXISTS produkty_zlecenia (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        zlecenie_id INTEGER NOT NULL,
        nazwa TEXT NOT NULL,
        ilosc REAL DEFAULT 1,
        cena REAL DEFAULT 0,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (zlecenie_id) REFERENCES zlecenia(id))""")

    c.execute("""CREATE TABLE IF NOT EXISTS opcje_zlecen (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        zlecenie_id INTEGER NOT NULL, typ TEXT NOT NULL,
        opis TEXT, kwota REAL DEFAULT 0, ilosc REAL DEFAULT 0)""")

    c.execute("""CREATE TABLE IF NOT EXISTS powiadomienia (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        zlecenie_id INTEGER, operacja_id INTEGER, typ TEXT NOT NULL,
        tytul TEXT, tresc TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        odczytane INTEGER DEFAULT 0, dla_roli TEXT DEFAULT 'all')""")

    # ➤ NOWA TABELA: stawki per zlecenie
    c.execute("""CREATE TABLE IF NOT EXISTS stawki_zlecen (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        zlecenie_id INTEGER NOT NULL,
        stanowisko TEXT NOT NULL,
        stawka_godz REAL NOT NULL DEFAULT 0,
        zbrojenie_stawka_godz REAL DEFAULT 0.0,
        FOREIGN KEY (zlecenie_id) REFERENCES zlecenia(id) ON DELETE CASCADE,
        UNIQUE(zlecenie_id, stanowisko))""")

    # Dane startowe
    c.execute("SELECT id FROM users WHERE username='admin'")
    if not c.fetchone():
        def h(p): return hashlib.sha256(p.encode()).hexdigest()
        users = [
            ('admin',     h('admin123'),   'Administrator',     'admin'),
            ('jan.tech',  h('tech123'),    'Jan Kowalski',      'technolog'),
            ('piotr.p',   h('pracownik1'), 'Piotr Nowak',       'pracownik'),
            ('anna.p',    h('pracownik2'), 'Anna Wiśniewska',   'pracownik'),
            ('mag.jan',   h('magazyn123'), 'Jan Magazynowski',  'magazynier'),
            ('majster.k', h('majster123'), 'Krzysztof Majster', 'majster'),
        ]
        c.executemany("INSERT INTO users (username,password,full_name,role) VALUES (?,?,?,?)", users)

    c.execute("SELECT id FROM stawki LIMIT 1")
    if not c.fetchone():
        stawki = [
            ('Toczenie konwencjonalne',     65.0, 15.0),
            ('Toczenie CNC',                90.0, 10.0),
            ('Frezowanie konwencjonalne',   70.0, 20.0),
            ('Frezowanie CNC - mała brama', 95.0, 12.0),
            ('Frezowanie CNC - duża brama',110.0, 18.0),
            ('Cięcie na pile',              45.0,  5.0),
            ('Wypalanie laserowe',         120.0,  8.0),
            ('Wypalanie plazmowe',          80.0, 10.0),
            ('Prace ślusarskie',            60.0, 25.0),
            ('Szlifowanie',                 75.0, 12.0),
        ]
        c.executemany("INSERT INTO stawki (stanowisko,stawka_godz,czas_norma_min) VALUES (?,?,?)", stawki)

    try: c.execute("ALTER TABLE sesje_pracy ADD COLUMN sesja_glowna INTEGER DEFAULT 1")
    except: pass
    try: c.execute("ALTER TABLE sesje_pracy ADD COLUMN rzeczywiste_stanowisko TEXT DEFAULT NULL")
    except: pass

    c.execute("""CREATE TABLE IF NOT EXISTS user_permissions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL UNIQUE,
        tabs TEXT NOT NULL DEFAULT '[]',
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE)""")

    c.execute("""CREATE TABLE IF NOT EXISTS app_feedback (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        user_name TEXT,
        ocena INTEGER NOT NULL,
        wiadomosc TEXT DEFAULT '',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE SET NULL)""")

    # ─── Fakturowanie ────────────────────────────────────────────────────────
    c.execute("""CREATE TABLE IF NOT EXISTS kontrahenci (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nazwa TEXT NOT NULL,
        nip TEXT DEFAULT '',
        adres TEXT DEFAULT '',
        kod_pocztowy TEXT DEFAULT '',
        miasto TEXT DEFAULT '',
        kraj TEXT DEFAULT 'PL',
        email TEXT DEFAULT '',
        telefon TEXT DEFAULT '',
        uwagi TEXT DEFAULT '',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")

    c.execute("""CREATE TABLE IF NOT EXISTS faktury (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        numer TEXT UNIQUE NOT NULL,
        kontrahent_id INTEGER,
        zlecenie_id INTEGER,
        data_wystawienia TEXT NOT NULL,
        data_sprzedazy TEXT DEFAULT '',
        termin_platnosci TEXT DEFAULT '',
        forma_platnosci TEXT DEFAULT 'przelew',
        status TEXT DEFAULT 'szkic',
        uwagi TEXT DEFAULT '',
        waluta TEXT DEFAULT 'PLN',
        total_netto REAL DEFAULT 0,
        total_vat REAL DEFAULT 0,
        total_brutto REAL DEFAULT 0,
        created_by INTEGER,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (kontrahent_id) REFERENCES kontrahenci(id),
        FOREIGN KEY (zlecenie_id) REFERENCES zlecenia(id))""")

    c.execute("""CREATE TABLE IF NOT EXISTS pozycje_faktury (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        faktura_id INTEGER NOT NULL,
        lp INTEGER DEFAULT 1,
        nazwa TEXT NOT NULL,
        jm TEXT DEFAULT 'szt',
        ilosc REAL DEFAULT 1,
        cena_netto REAL DEFAULT 0,
        vat_procent REAL DEFAULT 23,
        wartosc_netto REAL DEFAULT 0,
        wartosc_brutto REAL DEFAULT 0,
        FOREIGN KEY (faktura_id) REFERENCES pozycje_faktury(id))""")

    try: c.execute("ALTER TABLE faktury ADD COLUMN data_sprzedazy TEXT DEFAULT ''")
    except: pass
    try: c.execute("ALTER TABLE faktury ADD COLUMN forma_platnosci TEXT DEFAULT 'przelew'")
    except: pass

    # ➤ Baza materiałów (import z xlsx)
    c.execute("""CREATE TABLE IF NOT EXISTS materialy (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        kod TEXT,
        indeks TEXT UNIQUE NOT NULL,
        opis TEXT NOT NULL,
        jm TEXT DEFAULT 'kg',
        do_dyspozycji REAL DEFAULT 0,
        stan_rzeczywisty REAL DEFAULT 0,
        rezerwacja REAL DEFAULT 0,
        kod_paskowy TEXT,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")

    # ➤ BOM – pozycje materiałowe zlecenia
    c.execute("""CREATE TABLE IF NOT EXISTS bom_pozycje (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        zlecenie_id INTEGER NOT NULL,
        material_id INTEGER NOT NULL,
        ilosc REAL NOT NULL DEFAULT 1,
        ilosc_wykonana REAL DEFAULT 0,
        uwagi TEXT DEFAULT '',
        masa_kg REAL DEFAULT 0,
        gatunek_stali TEXT DEFAULT '',
        wymiary_str TEXT DEFAULT '',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (zlecenie_id) REFERENCES zlecenia(id) ON DELETE CASCADE,
        FOREIGN KEY (material_id) REFERENCES materialy(id))""")
    # Migracja: dodaj kolumny jeśli nie istnieją (dla istniejących baz)
    for _col, _def in [("masa_kg","REAL DEFAULT 0"),("gatunek_stali","TEXT DEFAULT ''"),("wymiary_str","TEXT DEFAULT ''")]:
        try:
            c.execute(f"ALTER TABLE bom_pozycje ADD COLUMN {_col} {_def}")
        except Exception:
            pass

    # ➤ Narzędziownia – baza narzędzi
    c.execute("""CREATE TABLE IF NOT EXISTS narzedzia (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        indeks TEXT UNIQUE NOT NULL,
        nazwa TEXT NOT NULL,
        typ TEXT DEFAULT '',
        jm TEXT DEFAULT 'szt',
        stan INTEGER DEFAULT 0,
        stan_min INTEGER DEFAULT 1,
        lokalizacja TEXT DEFAULT '',
        uwagi TEXT DEFAULT '',
        kod_paskowy TEXT DEFAULT '',
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
    for _col, _def in [
        ("lokalizacja","TEXT DEFAULT ''"),
        ("uwagi","TEXT DEFAULT ''"),
        ("kod_paskowy","TEXT DEFAULT ''"),
        ("stan_min","INTEGER DEFAULT 1"),
        ("typ","TEXT DEFAULT ''"),
    ]:
        try: c.execute(f"ALTER TABLE narzedzia ADD COLUMN {_col} {_def}")
        except: pass

    # ➤ MRP – rezerwacje materiałów pod zlecenia G
    c.execute("""CREATE TABLE IF NOT EXISTS mrp_rezerwacje (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        zlecenie_id INTEGER NOT NULL,
        zlecenie_nr TEXT NOT NULL,
        material_indeks TEXT NOT NULL,
        ilosc REAL NOT NULL DEFAULT 0,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(zlecenie_id, material_indeks),
        FOREIGN KEY (zlecenie_id) REFERENCES zlecenia(id) ON DELETE CASCADE)""")

    # ➤ Magazyn – rezerwacje materiałów (poprzednio w localStorage)
    c.execute("""CREATE TABLE IF NOT EXISTS mag_rezerwacje (
        id TEXT PRIMARY KEY,
        material_id INTEGER NOT NULL,
        material_indeks TEXT NOT NULL,
        material_opis TEXT NOT NULL,
        material_jm TEXT DEFAULT 'kg',
        ilosc REAL NOT NULL DEFAULT 0,
        zlecenie_nr TEXT DEFAULT '',
        uwagi TEXT DEFAULT '',
        status TEXT DEFAULT 'aktywna',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")

    # ➤ Narzędziownia – rezerwacje narzędzi pod zlecenia
    c.execute("""CREATE TABLE IF NOT EXISTS narzedzia_rezerwacje (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        narzedzie_id INTEGER NOT NULL,
        zlecenie_nr TEXT NOT NULL,
        cel TEXT DEFAULT 'produkcja',
        ilosc INTEGER DEFAULT 1,
        data_od TEXT,
        data_do TEXT,
        uwagi TEXT DEFAULT '',
        status TEXT DEFAULT 'aktywna',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (narzedzie_id) REFERENCES narzedzia(id) ON DELETE CASCADE)""")


    # ➤ Narzędziownia – historia pobrań (materiały eksploatacyjne)
    c.execute("""CREATE TABLE IF NOT EXISTS narzedzia_pobrania (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        narzedzie_id INTEGER NOT NULL,
        zlecenie_nr TEXT DEFAULT '—',
        ilosc REAL NOT NULL DEFAULT 1,
        uwagi TEXT DEFAULT '',
        user_name TEXT DEFAULT '',
        status TEXT DEFAULT 'wydane',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (narzedzie_id) REFERENCES narzedzia(id) ON DELETE CASCADE)""")

    # ➤ Migracja: uzupełnij stawki_zlecen dla istniejących zleceń
    existing_zl = c.execute("SELECT id FROM zlecenia").fetchall()
    for zl_row in existing_zl:
        try:
            zid = zl_row[0]
            ops = c.execute(
                "SELECT DISTINCT stanowisko FROM operacje WHERE zlecenie_id=? AND stanowisko IS NOT NULL AND stanowisko!=''",
                (zid,)
            ).fetchall()
            global_stawki = {
                r[0]: (r[1] or 0, r[2] or 0)
                for r in c.execute("SELECT stanowisko, stawka_godz, zbrojenie_stawka_godz FROM stawki").fetchall()
            }
            for op in ops:
                st = op[0]
                if not st:
                    continue
                existing = c.execute(
                    "SELECT id FROM stawki_zlecen WHERE zlecenie_id=? AND stanowisko=?",
                    (zid, st)
                ).fetchone()
                if existing:
                    continue
                g = global_stawki.get(st, (0, 0))
                c.execute(
                    """INSERT INTO stawki_zlecen (zlecenie_id, stanowisko, stawka_godz, zbrojenie_stawka_godz)
                       VALUES (?, ?, ?, ?)""",
                    (zid, st, g[0], g[1])
                )
        except Exception as e:
            print(f"  Migracja stawek zlecenia {zl_row[0]}: {e}")


    # ─── Drzewo G/P (struktura wyrobów z ERP) ────────────────────────────────
    # Katalog wyrobów / półproduktów (każde G i P to jeden rekord)
    c.execute("""CREATE TABLE IF NOT EXISTS wyroby (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        symbol TEXT UNIQUE NOT NULL,          -- np. G14402, P18653
        typ TEXT NOT NULL DEFAULT 'P',         -- 'G' = wyrób główny, 'P' = półprodukt
        nazwa TEXT NOT NULL,
        numer_rysunku TEXT DEFAULT '',
        jednostka TEXT DEFAULT 'szt',
        opis TEXT DEFAULT '',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
    for _col, _def in [
        ("numer_rysunku","TEXT DEFAULT ''"),
        ("jednostka","TEXT DEFAULT 'szt'"),
        ("opis","TEXT DEFAULT ''"),
    ]:
        try: c.execute(f"ALTER TABLE wyroby ADD COLUMN {_col} {_def}")
        except: pass

    # Struktura BOM wyrobów: wyrob_id zawiera skladnik_id w ilosci X
    # (n-poziomowa hierarchia, wyrob P moze byc skladnikiem wielu G lub P)
    c.execute("""CREATE TABLE IF NOT EXISTS wyroby_bom (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        wyrob_id INTEGER NOT NULL,             -- rodzic (G lub P)
        skladnik_id INTEGER NOT NULL,          -- dziecko (P lub M)
        typ_skladnika TEXT NOT NULL DEFAULT 'P', -- 'P'=polprodukt, 'M'=material
        material_indeks TEXT DEFAULT '',       -- dla typ='M': indeks z tabeli materialy
        ilosc REAL NOT NULL DEFAULT 1,
        jednostka TEXT DEFAULT 'szt',
        pozycja INTEGER DEFAULT 0,            -- kolejnosc na liscie
        uwagi TEXT DEFAULT '',
        FOREIGN KEY (wyrob_id) REFERENCES wyroby(id) ON DELETE CASCADE,
        UNIQUE(wyrob_id, skladnik_id, typ_skladnika, material_indeks))""")
    for _col, _def in [
        ("pozycja","INTEGER DEFAULT 0"),
        ("uwagi","TEXT DEFAULT ''"),
        ("material_indeks","TEXT DEFAULT ''"),
    ]:
        try: c.execute(f"ALTER TABLE wyroby_bom ADD COLUMN {_col} {_def}")
        except: pass
    # Migracja UNIQUE: usuń stary indeks i utwórz nowy z material_indeks
    try:
        c.execute("DROP INDEX IF EXISTS sqlite_autoindex_wyroby_bom_1")
    except: pass
    try:
        c.execute("""CREATE UNIQUE INDEX IF NOT EXISTS idx_bom_unique
                     ON wyroby_bom(wyrob_id, skladnik_id, typ_skladnika, material_indeks)""")
    except: pass
    # Indeksy wydajnościowe
    try:
        c.execute("CREATE INDEX IF NOT EXISTS idx_bom_wyrob ON wyroby_bom(wyrob_id)")
    except: pass
    try:
        c.execute("CREATE INDEX IF NOT EXISTS idx_bom_skladnik ON wyroby_bom(skladnik_id) WHERE typ_skladnika='P'")
    except: pass
    try:
        c.execute("CREATE INDEX IF NOT EXISTS idx_wyroby_symbol ON wyroby(symbol)")
    except: pass
    try:
        c.execute("CREATE INDEX IF NOT EXISTS idx_zlecenia_numer ON zlecenia(numer)")
    except: pass

    # Zapotrzebowania produkcyjne: G zleca określoną ilość P do wykonania
    # Powiązanie konkretnego zlecenia G z półproduktem P i jego zleceniem
    c.execute("""CREATE TABLE IF NOT EXISTS zapotrzebowania (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        zlecenie_g_id INTEGER NOT NULL,        -- zlecenie G (rodzic)
        wyrob_p_symbol TEXT NOT NULL,          -- symbol P (np. P18653)
        zlecenie_p_id INTEGER,                 -- zlecenie P (może być NULL jeśli nie założone)
        ilosc_wymagana REAL NOT NULL DEFAULT 1,
        ilosc_wykonana REAL DEFAULT 0,
        status TEXT DEFAULT 'oczekuje',        -- oczekuje / w_toku / zakonczone / anulowane
        priorytet INTEGER DEFAULT 0,
        uwagi TEXT DEFAULT '',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (zlecenie_g_id) REFERENCES zlecenia(id) ON DELETE CASCADE,
        FOREIGN KEY (zlecenie_p_id) REFERENCES zlecenia(id) ON DELETE SET NULL)""")
    for _col, _def in [
        ("priorytet","INTEGER DEFAULT 0"),
        ("uwagi","TEXT DEFAULT ''"),
    ]:
        try: c.execute(f"ALTER TABLE zapotrzebowania ADD COLUMN {_col} {_def}")
        except: pass

    # Indeks importów drzewa G/P z PDF/ERP
    c.execute("""CREATE TABLE IF NOT EXISTS import_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        typ TEXT NOT NULL,                     -- 'drzewo_gp', 'karta_technologiczna'
        symbol_glowny TEXT NOT NULL,
        ilosc_wyrobow INTEGER DEFAULT 0,
        ilosc_pozycji_bom INTEGER DEFAULT 0,
        bledy TEXT DEFAULT '[]',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        created_by INTEGER)""")


    conn.commit()
    conn.close()
    print(f"✓ Baza danych gotowa: {DB_PATH}")

# ─── System backupu i przywracania danych ─────────────────────────────────────
import threading as _threading
_TABLES_TO_BACKUP = [
    "zlecenia", "operacje", "sesje_pracy", "users", "stawki",
    "katalog_produktow", "produkty_zlecenia", "opcje_zlecen", "user_permissions",
    "stawki_zlecen",
    "materialy",            # baza materialow (import xlsx)
    "bom_pozycje",          # zapotrzebowanie materialowe zlecen
    "narzedzia",            # baza narzedzi (narzedzialnia)
    "narzedzia_rezerwacje", # rezerwacje narzedzi pod zlecenia
    "narzedzia_pobrania",   # historia pobran materialow z narzedzilalni
    "wyroby",               # katalog wyrobow i polproduktow G/P
    "wyroby_bom",           # struktura BOM wyrobow
    "zapotrzebowania",      # zapotrzebowania P dla zlecen G
    "mrp_rezerwacje",       # rezerwacje materialow MRP pod zlecenia G
    "mag_rezerwacje",       # rezerwacje materialow magazynowych (dawniej localStorage)
]

# ─── GitHub Gist helpers ────────────────────────────────────────────────────────
def _gist_get_id() -> str:
    if GIST_ID:
        return GIST_ID
    if os.path.exists(GIST_ID_FILE):
        try:
            return open(GIST_ID_FILE).read().strip()
        except Exception:
            pass
    if not GIST_TOKEN:
        return ""
    try:
        req = urllib.request.Request(
            "https://api.github.com/gists",
            headers={"Authorization": f"token {GIST_TOKEN}", "Accept": "application/vnd.github+json"}
        )
        with urllib.request.urlopen(req, timeout=10) as r:
            gists = json.loads(r.read())
            for g in gists:
                if "produkcja_backup.json" in g.get("files", {}):
                    gid = g["id"]
                    open(GIST_ID_FILE, "w").write(gid)
                    print(f"✓ Znaleziono istniejący Gist: {gid}")
                    return gid
    except Exception as e:
        print(f"✗ Gist search error: {e}")
        return ""

def _gist_save(data: dict) -> bool:
    if not GIST_TOKEN:
        return False
    content = json.dumps(data, ensure_ascii=False, default=str)
    gist_id = _gist_get_id()
    try:
        payload = json.dumps({
            "description": "Produkcja DB backup – auto",
            "public": False,
            "files": {"produkcja_backup.json": {"content": content}}
        }).encode()
        if gist_id:
            req = urllib.request.Request(
                f"https://api.github.com/gists/{gist_id}",
                data=payload,
                method="PATCH",
                headers={
                    "Authorization": f"token {GIST_TOKEN}",
                    "Accept": "application/vnd.github+json",
                    "Content-Type": "application/json"
                }
            )
        else:
            req = urllib.request.Request(
                "https://api.github.com/gists",
                data=payload,
                method="POST",
                headers={
                    "Authorization": f"token {GIST_TOKEN}",
                    "Accept": "application/vnd.github+json",
                    "Content-Type": "application/json"
                }
            )
        with urllib.request.urlopen(req, timeout=15) as r:
            resp = json.loads(r.read())
            new_id = resp.get("id", gist_id)
            if new_id and new_id != gist_id:
                open(GIST_ID_FILE, "w").write(new_id)
                print(f"✓ Gist utworzony: {new_id}")
            print(f"✓ Gist backup zapisany ({len(content)} B, {data.get('_ts','?')})")
            return True
    except Exception as e:
        print(f"✗ Błąd zapisu Gist: {e}")
        return False

def _gist_load() -> dict | None:
    if not GIST_TOKEN:
        return None
    gist_id = _gist_get_id()
    if not gist_id:
        print("✗ Gist: brak ID – nie można pobrać backupu")
        return None
    try:
        req = urllib.request.Request(
            f"https://api.github.com/gists/{gist_id}",
            headers={"Authorization": f"token {GIST_TOKEN}", "Accept": "application/vnd.github+json"}
        )
        with urllib.request.urlopen(req, timeout=15) as r:
            resp = json.loads(r.read())
            file_info = resp.get("files", {}).get("produkcja_backup.json", {})
            raw_url = file_info.get("raw_url")
            if not raw_url:
                print("✗ Gist: brak pliku produkcja_backup.json")
                return None
            with urllib.request.urlopen(raw_url, timeout=15) as r:
                data = json.loads(r.read())
                print(f"✓ Gist backup pobrany (ts: {data.get('_ts','?')})")
                return data
    except Exception as e:
        print(f"✗ Błąd odczytu Gist: {e}")
        return None

# ─── Główne funkcje backup/restore ─────────────────────────────────────────────
def _db_backup_to_json(path: str = None) -> dict:
    path = path or BACKUP_PATH
    data = {"_ts": _now(), "_ver": "v19", "tables": {}}
    try:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        conn.row_factory = sqlite3.Row
        for tbl in _TABLES_TO_BACKUP:
            try:
                rows = conn.execute(f"SELECT * FROM {tbl}").fetchall()
                data["tables"][tbl] = [dict(r) for r in rows]
            except Exception:
                data["tables"][tbl] = []
        conn.close()
        try:
            os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, default=str)
            print(f"✓ Backup lokalny: {path}")
        except Exception as e:
            print(f"✗ Backup lokalny nieudany: {e}")
        _gist_save(data)
    except Exception as e:
        print(f"✗ Błąd backupu: {e}")
    return data

def _db_restore_from_json(path: str = None) -> bool:
    path = path or BACKUP_PATH
    if not os.path.exists(path):
        print(f"✗ Brak pliku backupu: {path}")
        return False
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return _db_restore_from_dict(data)
    except Exception as e:
        print(f"✗ Błąd przywracania z pliku: {e}")
        return False

def _db_restore_from_dict(data: dict) -> bool:
    tables = data.get("tables", {})
    try:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        conn.execute("PRAGMA foreign_keys = OFF")
        for tbl, rows in tables.items():
            if not rows:
                continue
            try:
                cols = list(rows[0].keys())
                placeholders = ", ".join(["?" for _ in cols])
                col_list = ", ".join(cols)
                conn.executemany(
                    f"INSERT OR REPLACE INTO {tbl} ({col_list}) VALUES ({placeholders})",
                    [[r.get(c) for c in cols] for r in rows]
                )
            except Exception as e:
                print(f"✗ Błąd przywracania tabeli {tbl}: {e}")
        conn.execute("PRAGMA foreign_keys = ON")
        conn.commit()
        conn.close()
        total = sum(len(v) for v in tables.values())
        aktywne = len([r for r in tables.get("sesje_pracy", []) if r.get("status") == "aktywna"])
        print(f"✓ Przywrócono {total} rekordów (ts: {data.get('_ts','?')})")
        if aktywne:
            print(f"  ↳ W tym {aktywne} aktywnych sesji pracy")
        return True
    except Exception as e:
        print(f"✗ Błąd przywracania: {e}")
        return False

def _auto_backup_loop():
    import time as _time
    _time.sleep(60)
    while True:
        try:
            db_mtime  = os.path.getmtime(DB_PATH) if os.path.exists(DB_PATH) else 0
            bak_mtime = os.path.getmtime(BACKUP_PATH) if os.path.exists(BACKUP_PATH) else 0
            if db_mtime > bak_mtime:
                _db_backup_to_json()
            elif db_mtime > 0 and (_time.time() - bak_mtime > 3600):
                _db_backup_to_json()
        except Exception as e:
            print(f"auto-backup loop error: {e}")
        _time.sleep(BACKUP_INTERVAL)

# ─── Endpointy backupu ────────────────────────────────────────────────────────
@app.get("/api/admin/backup", dependencies=[Depends(verify_key)])
def admin_backup():
    data = _db_backup_to_json()
    summary = {tbl: len(rows) for tbl, rows in data.get("tables", {}).items()}
    return {"ok": True, "ts": data["_ts"], "path": BACKUP_PATH, "rows": summary}

@app.get("/api/admin/backup/download")
def admin_backup_download(
    x_api_key: str = "",
    x_api_key_h: str = Header(None, alias="x-api-key")
):
    key = x_api_key or x_api_key_h or ""
    if key != API_KEY:
        raise HTTPException(403, "Nieprawidłowy klucz API")
    data = _db_backup_to_json()
    content = json.dumps(data, ensure_ascii=False, default=str, indent=2)
    return StreamingResponse(
        iter([content.encode("utf-8")]),
        media_type="application/json",
        headers={"Content-Disposition": 'attachment; filename="produkcja_backup.json"'}
    )

@app.post("/api/admin/backup/restore", dependencies=[Depends(verify_key)])
async def admin_backup_restore(request: Request):
    body = await request.body()
    if not body:
        ok = _db_restore_from_json()
        return {"ok": ok, "source": "local_file"}
    tmp_path = BACKUP_PATH + ".tmp"
    try:
        with open(tmp_path, "wb") as f:
            f.write(body)
        ok = _db_restore_from_json(tmp_path)
        os.remove(tmp_path)
        return {"ok": ok, "source": "uploaded"}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.post("/api/admin/backup/restore-upload", dependencies=[Depends(verify_key)])
async def admin_backup_restore_upload(request: Request):
    try:
        body = await request.body()
        data = json.loads(body)
        ok = _db_restore_from_dict(data)
        if ok:
            rows = {k: len(v) if isinstance(v, list) else 1
                    for k, v in data.items() if not k.startswith("_")}
            return {"ok": True, "source": "disk_upload", "rows": rows}
        return {"ok": False, "error": "Przywracanie nie powiodło się"}
    except json.JSONDecodeError:
        return {"ok": False, "error": "Nieprawidłowy format JSON"}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/api/admin/backup/status", dependencies=[Depends(verify_key)])
def admin_backup_status():
    exists = os.path.exists(BACKUP_PATH)
    size = os.path.getsize(BACKUP_PATH) if exists else 0
    ts = None
    if exists:
        try:
            with open(BACKUP_PATH, "r") as f:
                d = json.load(f)
                ts = d.get("_ts")
        except Exception:
            pass
    gist_id = _gist_get_id() if GIST_TOKEN else None
    return {
        "backup_path": BACKUP_PATH,
        "db_path": DB_PATH,
        "backup_exists": exists,
        "backup_size_kb": round(size / 1024, 1),
        "backup_ts": ts,
        "backup_interval_sec": BACKUP_INTERVAL,
        "gist_enabled": bool(GIST_TOKEN),
        "gist_id": gist_id or None,
    }

@app.get("/api/oblozenie", dependencies=[Depends(verify_key)])
def get_oblozenie():
    with get_db() as conn:
        stawki_rows = conn.execute(
            "SELECT stanowisko, stawka_godz, opis FROM stawki ORDER BY stanowisko"
        ).fetchall()
        stawki_set = {r["stanowisko"]: dict(r) for r in stawki_rows}
        op_stanowiska = conn.execute("""
            SELECT DISTINCT o.stanowisko
            FROM operacje o JOIN zlecenia z ON o.zlecenie_id = z.id
            WHERE z.status IN ('nowe','w_realizacji')
              AND o.stanowisko IS NOT NULL AND o.stanowisko != ''
              AND o.status != 'zakonczona'
        """).fetchall()
        extra = [r["stanowisko"] for r in op_stanowiska if r["stanowisko"] not in stawki_set]

        ops = conn.execute("""
            SELECT o.id, o.nazwa, o.stanowisko, o.status, o.kolejnosc,
                   o.ilosc_wykonana, o.czas_norma, o.czas_zbrojenia_min, o.opis_czynnosci,
                   z.id as zlecenie_id, z.numer, z.nazwa as zlecenie_nazwa,
                   z.termin, z.ilosc_sztuk, z.status as zlecenie_status
            FROM operacje o JOIN zlecenia z ON o.zlecenie_id = z.id
            WHERE z.status IN ('nowe','w_realizacji')
              AND o.status != 'zakonczona'
              AND o.stanowisko IS NOT NULL AND o.stanowisko != ''
            ORDER BY z.termin ASC, z.id ASC, o.kolejnosc ASC
        """).fetchall()

        # Pobierz aktywne sesje dla operacji w_toku
        active_sess = {}
        sess_rows = conn.execute("""
            SELECT sp.operacja_id, sp.start_time, sp.pauzy
            FROM sesje_pracy sp
            WHERE sp.status='aktywna'
            ORDER BY sp.id ASC
        """).fetchall()
        for sr in sess_rows:
            oid = sr["operacja_id"]
            if oid and oid not in active_sess:
                active_sess[oid] = {"start_time": sr["start_time"], "pauzy": sr["pauzy"] or "[]"}

        stanowiska_ops = {}
        for o in ops:
            st = o["stanowisko"]
            if st not in stanowiska_ops:
                stanowiska_ops[st] = []
            sess = active_sess.get(o["id"], {})
            stanowiska_ops[st].append({
                "op_id": o["id"], "op_nazwa": o["nazwa"], "op_status": o["status"],
                "op_kolejnosc": o["kolejnosc"], "ilosc_wykonana": o["ilosc_wykonana"],
                "czas_norma": o["czas_norma"], "czas_zbrojenia_min": o["czas_zbrojenia_min"] or 0,
                "opis_czynnosci": o["opis_czynnosci"] or "", "zlecenie_id": o["zlecenie_id"],
                "zlecenie_numer": o["numer"], "zlecenie_nazwa": o["zlecenie_nazwa"],
                "zlecenie_status": o["zlecenie_status"],
                "termin": o["termin"], "ilosc_sztuk": o["ilosc_sztuk"],
                "sesja_start": sess.get("start_time"), "sesja_pauzy": sess.get("pauzy", "[]"),
            })

        result = []
        all_names = sorted(set(list(stawki_set.keys()) + extra))
        for name in all_names:
            info = stawki_set.get(name, {})
            result.append({
                "stanowisko": name, "stawka_godz": info.get("stawka_godz"),
                "opis": info.get("opis", ""), "in_stawki": name in stawki_set,
                "operacje": stanowiska_ops.get(name, []),
            })
        return result

# ─── Materiały – import i CRUD ────────────────────────────────────────────────

@app.post("/api/materialy/import", dependencies=[Depends(verify_key)])
async def import_materialy(file: UploadFile = File(...)):
    """Import bazy materiałów z pliku xlsx. Nadpisuje istniejące rekordy (upsert po indeksie)."""
    content = await file.read()
    try:
        wb = _openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        raise HTTPException(400, f"Błąd odczytu pliku xlsx: {e}")

    if not rows:
        raise HTTPException(400, "Plik jest pusty")

    # Znajdź nagłówki (pierwsza niepusta linia)
    header_row = [str(c).strip() if c else '' for c in rows[0]]
    # Mapowanie kolumn (elastyczne)
    col_map = {}
    for i, h in enumerate(header_row):
        hl = h.lower()
        if 'kod' in hl and 'paskowy' not in hl: col_map.setdefault('kod', i)
        elif 'indeks' in hl:  col_map['indeks'] = i
        elif 'opis' in hl:    col_map['opis'] = i
        elif 'jm' in hl:      col_map['jm'] = i
        elif 'dyspozycji' in hl: col_map['do_dyspozycji'] = i
        elif 'rzeczywisty' in hl: col_map['stan_rzeczywisty'] = i
        elif 'rezerwacja' in hl: col_map['rezerwacja'] = i
        elif 'paskowy' in hl: col_map['kod_paskowy'] = i

    required = {'indeks', 'opis'}
    missing = required - set(col_map.keys())
    if missing:
        raise HTTPException(400, f"Brak wymaganych kolumn: {missing}. Znalezione: {header_row}")

    def cell(row, key):
        i = col_map.get(key)
        if i is None: return None
        v = row[i] if i < len(row) else None
        if v is None: return None
        return str(v).strip() if isinstance(v, str) else v

    now = _now()
    imported = 0
    skipped = 0
    with get_db() as conn:
        for row in rows[1:]:
            indeks = cell(row, 'indeks')
            opis = cell(row, 'opis')
            if not indeks or not opis:
                skipped += 1
                continue
            kod = cell(row, 'kod')
            jm = cell(row, 'jm') or 'kg'
            do_dysp = float(cell(row, 'do_dyspozycji') or 0)
            stan = float(cell(row, 'stan_rzeczywisty') or 0)
            rez = float(cell(row, 'rezerwacja') or 0)
            kp = cell(row, 'kod_paskowy')
            if kp: kp = str(int(float(kp))) if isinstance(kp, float) else str(kp)
            conn.execute("""
                INSERT INTO materialy (kod, indeks, opis, jm, do_dyspozycji, stan_rzeczywisty, rezerwacja, kod_paskowy, updated_at)
                VALUES (?,?,?,?,?,?,?,?,?)
                ON CONFLICT(indeks) DO UPDATE SET
                  kod=excluded.kod, opis=excluded.opis, jm=excluded.jm,
                  do_dyspozycji=excluded.do_dyspozycji, stan_rzeczywisty=excluded.stan_rzeczywisty,
                  rezerwacja=excluded.rezerwacja, kod_paskowy=excluded.kod_paskowy, updated_at=excluded.updated_at
            """, (kod, indeks, opis, jm, do_dysp, stan, rez, kp, now))
            imported += 1
    return {"ok": True, "imported": imported, "skipped": skipped}

@app.get("/api/materialy", dependencies=[Depends(verify_key)])
def get_materialy(q: str = "", limit: int = 50):
    with get_db() as conn:
        if q:
            pattern = f"%{q}%"
            rows = conn.execute("""
                SELECT id, kod, indeks, opis, jm, do_dyspozycji, stan_rzeczywisty, rezerwacja, kod_paskowy
                FROM materialy
                WHERE opis LIKE ? OR indeks LIKE ? OR kod LIKE ?
                ORDER BY opis LIMIT ?
            """, (pattern, pattern, pattern, limit)).fetchall()
        else:
            rows = conn.execute("""
                SELECT id, kod, indeks, opis, jm, do_dyspozycji, stan_rzeczywisty, rezerwacja, kod_paskowy
                FROM materialy ORDER BY opis LIMIT ?
            """, (limit,)).fetchall()
        return [dict(r) for r in rows]

@app.get("/api/materialy/count", dependencies=[Depends(verify_key)])
def count_materialy():
    with get_db() as conn:
        n = conn.execute("SELECT COUNT(*) FROM materialy").fetchone()[0]
        return {"count": n}

# ─── Materiały – ręczny CRUD ──────────────────────────────────────────────────
class MaterialIn(BaseModel):
    indeks: str
    opis: str
    kod: Optional[str] = ""
    jm: Optional[str] = "kg"
    do_dyspozycji: Optional[float] = 0
    stan_rzeczywisty: Optional[float] = 0
    rezerwacja: Optional[float] = 0
    kod_paskowy: Optional[str] = ""

@app.post("/api/materialy", dependencies=[Depends(verify_key)])
def create_material(req: MaterialIn):
    """Ręczne dodanie nowego materiału do bazy."""
    now = _now()
    with get_db() as conn:
        existing = conn.execute("SELECT id FROM materialy WHERE indeks=?", (req.indeks,)).fetchone()
        if existing:
            raise HTTPException(400, f"Materiał o indeksie '{req.indeks}' już istnieje")
        cur = conn.execute(
            """INSERT INTO materialy (kod, indeks, opis, jm, do_dyspozycji, stan_rzeczywisty, rezerwacja, kod_paskowy, updated_at)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (req.kod or "", req.indeks, req.opis, req.jm or "kg",
             req.do_dyspozycji or 0, req.stan_rzeczywisty or 0,
             req.rezerwacja or 0, req.kod_paskowy or "", now)
        )
        new_id = cur.lastrowid
    _threading.Thread(target=_db_backup_to_json, daemon=True).start()
    return {"ok": True, "id": new_id}

@app.put("/api/materialy/{mid}", dependencies=[Depends(verify_key)])
def update_material(mid: int, req: MaterialIn):
    """Aktualizacja istniejącego materiału."""
    now = _now()
    with get_db() as conn:
        existing = conn.execute("SELECT id FROM materialy WHERE id=?", (mid,)).fetchone()
        if not existing:
            raise HTTPException(404, "Materiał nie istnieje")
        conflict = conn.execute("SELECT id FROM materialy WHERE indeks=? AND id!=?", (req.indeks, mid)).fetchone()
        if conflict:
            raise HTTPException(400, f"Inny materiał już posiada indeks '{req.indeks}'")
        conn.execute(
            """UPDATE materialy SET kod=?,indeks=?,opis=?,jm=?,do_dyspozycji=?,
               stan_rzeczywisty=?,rezerwacja=?,kod_paskowy=?,updated_at=? WHERE id=?""",
            (req.kod or "", req.indeks, req.opis, req.jm or "kg",
             req.do_dyspozycji or 0, req.stan_rzeczywisty or 0,
             req.rezerwacja or 0, req.kod_paskowy or "", now, mid)
        )
    _threading.Thread(target=_db_backup_to_json, daemon=True).start()
    return {"ok": True}

@app.delete("/api/materialy/{mid}", dependencies=[Depends(verify_key)])
def delete_material(mid: int):
    """Usunięcie materiału z bazy."""
    with get_db() as conn:
        existing = conn.execute("SELECT id FROM materialy WHERE id=?", (mid,)).fetchone()
        if not existing:
            raise HTTPException(404, "Materiał nie istnieje")
        conn.execute("DELETE FROM mag_rezerwacje WHERE material_id=?", (mid,))
        conn.execute("DELETE FROM bom_pozycje WHERE material_id=?", (mid,))
        conn.execute("DELETE FROM materialy WHERE id=?", (mid,))
    _threading.Thread(target=_db_backup_to_json, daemon=True).start()
    return {"ok": True}

# ─── Rezerwacje magazynowe (przeniesione z localStorage na serwer) ─────────────
@app.get("/api/mag-rezerwacje", dependencies=[Depends(verify_key)])
def get_mag_rezerwacje(status: Optional[str] = None):
    with get_db() as conn:
        if status:
            rows = conn.execute(
                "SELECT * FROM mag_rezerwacje WHERE status=? ORDER BY created_at DESC", (status,)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM mag_rezerwacje ORDER BY created_at DESC"
            ).fetchall()
        return [dict(r) for r in rows]

@app.post("/api/mag-rezerwacje", dependencies=[Depends(verify_key)])
def create_mag_rezerwacja(body: dict):
    """Tworzy rezerwację materiału w magazynie."""
    now = _now()
    rez_id = body.get("id") or f"rez_{int(time.time()*1000)}"
    material_id = body.get("material_id")
    ilosc = float(body.get("ilosc", 0))
    if not material_id or ilosc <= 0:
        raise HTTPException(400, "Wymagane: material_id i ilosc > 0")
    with get_db() as conn:
        mat = conn.execute(
            "SELECT id, indeks, opis, jm, do_dyspozycji FROM materialy WHERE id=?", (material_id,)
        ).fetchone()
        if not mat:
            raise HTTPException(404, "Materiał nie istnieje")
        conn.execute(
            """INSERT INTO mag_rezerwacje (id, material_id, material_indeks, material_opis, material_jm,
               ilosc, zlecenie_nr, uwagi, status, created_at, updated_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (rez_id, mat["id"], mat["indeks"], mat["opis"], mat["jm"],
             ilosc, body.get("zlecenie_nr",""), body.get("uwagi",""),
             "aktywna", now, now)
        )
        conn.execute(
            "UPDATE materialy SET do_dyspozycji = do_dyspozycji - ? WHERE id=?",
            (ilosc, material_id)
        )
    _threading.Thread(target=_db_backup_to_json, daemon=True).start()
    return {"ok": True, "id": rez_id}

@app.patch("/api/mag-rezerwacje/{rid}/zwolnij", dependencies=[Depends(verify_key)])
def zwolnij_mag_rezerwacje(rid: str):
    """Zwalnia rezerwację i przywraca stan do dyspozycji."""
    now = _now()
    with get_db() as conn:
        rez = conn.execute("SELECT * FROM mag_rezerwacje WHERE id=?", (rid,)).fetchone()
        if not rez:
            raise HTTPException(404, "Rezerwacja nie istnieje")
        if rez["status"] != "aktywna":
            raise HTTPException(400, "Rezerwacja nie jest aktywna")
        conn.execute(
            "UPDATE mag_rezerwacje SET status='zwolniona', updated_at=? WHERE id=?", (now, rid)
        )
        conn.execute(
            "UPDATE materialy SET do_dyspozycji = do_dyspozycji + ? WHERE id=?",
            (rez["ilosc"], rez["material_id"])
        )
    _threading.Thread(target=_db_backup_to_json, daemon=True).start()
    return {"ok": True}

@app.delete("/api/mag-rezerwacje/{rid}", dependencies=[Depends(verify_key)])
def delete_mag_rezerwacja(rid: str):
    """Usuwa rezerwację. Jeśli aktywna – przywraca stan do dyspozycji."""
    now = _now()
    with get_db() as conn:
        rez = conn.execute("SELECT * FROM mag_rezerwacje WHERE id=?", (rid,)).fetchone()
        if not rez:
            raise HTTPException(404, "Rezerwacja nie istnieje")
        if rez["status"] == "aktywna":
            conn.execute(
                "UPDATE materialy SET do_dyspozycji = do_dyspozycji + ? WHERE id=?",
                (rez["ilosc"], rez["material_id"])
            )
        conn.execute("DELETE FROM mag_rezerwacje WHERE id=?", (rid,))
    _threading.Thread(target=_db_backup_to_json, daemon=True).start()
    return {"ok": True}

# ─── Import rezerwacji z localStorage (migracja jednorazowa) ──────────────────
@app.post("/api/mag-rezerwacje/import-local", dependencies=[Depends(verify_key)])
def import_mag_rezerwacje_from_local(body: dict):
    """Importuje rezerwacje zapisane w localStorage przeglądarki do bazy serwera."""
    lista = body.get("rezerwacje", [])
    if not lista:
        return {"ok": True, "imported": 0}
    now = _now()
    imported = 0
    with get_db() as conn:
        for rez in lista:
            rez_id = rez.get("id")
            if not rez_id:
                continue
            exists = conn.execute("SELECT id FROM mag_rezerwacje WHERE id=?", (rez_id,)).fetchone()
            if exists:
                continue
            material_id = rez.get("material_id")
            mat = conn.execute("SELECT id, indeks, opis, jm FROM materialy WHERE id=?", (material_id,)).fetchone() if material_id else None
            if not mat:
                continue
            try:
                conn.execute(
                    """INSERT OR IGNORE INTO mag_rezerwacje
                       (id, material_id, material_indeks, material_opis, material_jm,
                        ilosc, zlecenie_nr, uwagi, status, created_at, updated_at)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                    (rez_id, mat["id"], mat["indeks"], mat["opis"], mat["jm"],
                     float(rez.get("ilosc", 0)), rez.get("zlecenie_nr",""),
                     rez.get("uwagi",""), rez.get("status","aktywna"),
                     rez.get("created_at", now), now)
                )
                imported += 1
            except Exception:
                pass
    _threading.Thread(target=_db_backup_to_json, daemon=True).start()
    return {"ok": True, "imported": imported}


class BomPozycjaIn(BaseModel):
    material_id: int
    ilosc: float
    uwagi: Optional[str] = ""

@app.get("/api/zlecenia/{zid}/bom", dependencies=[Depends(verify_key)])
def get_bom(zid: int):
    with get_db() as conn:
        rows = conn.execute("""
            SELECT bp.id, bp.zlecenie_id, bp.material_id, bp.ilosc, bp.ilosc_wykonana, bp.uwagi, bp.created_at,
                   bp.masa_kg, bp.gatunek_stali, bp.wymiary_str,
                   m.indeks, m.opis, m.jm, m.do_dyspozycji, m.stan_rzeczywisty, m.rezerwacja, m.kod
            FROM bom_pozycje bp
            JOIN materialy m ON m.id = bp.material_id
            WHERE bp.zlecenie_id = ?
            ORDER BY bp.id
        """, (zid,)).fetchall()
        result = []
        for r in rows:
            row = dict(r)
            # Dla starych pozycji bez masa_kg – oblicz na bieżąco z opisu
            if not row.get("masa_kg"):
                masa_kg, wymiary_str, _ = _calc_mass_kg(row["opis"], row["ilosc"])
                gatunek = _detect_steel_grade(row["opis"])
                row["masa_kg"] = masa_kg
                row["gatunek_stali"] = gatunek
                row["wymiary_str"] = wymiary_str
            result.append(row)
        return result

@app.post("/api/zlecenia/{zid}/bom", dependencies=[Depends(verify_key)])
def add_bom(zid: int, b: BomPozycjaIn):
    with get_db() as conn:
        # Sprawdź czy zlecenie istnieje
        if not conn.execute("SELECT id FROM zlecenia WHERE id=?", (zid,)).fetchone():
            raise HTTPException(404, "Zlecenie nie istnieje")
        if not conn.execute("SELECT id FROM materialy WHERE id=?", (b.material_id,)).fetchone():
            raise HTTPException(404, "Materiał nie istnieje")
        # Sprawdź duplikaty
        existing = conn.execute(
            "SELECT id FROM bom_pozycje WHERE zlecenie_id=? AND material_id=?", (zid, b.material_id)
        ).fetchone()
        if existing:
            raise HTTPException(409, "Ten materiał jest już w BOM tego zlecenia. Usuń i dodaj ponownie.")
        conn.execute(
            "INSERT INTO bom_pozycje (zlecenie_id, material_id, ilosc, uwagi, created_at) VALUES (?,?,?,?,?)",
            (zid, b.material_id, b.ilosc, b.uwagi or "", _now())
        )
        return {"ok": True}

@app.put("/api/bom/{bid}", dependencies=[Depends(verify_key)])
def update_bom(bid: int, b: BomPozycjaIn):
    with get_db() as conn:
        conn.execute(
            "UPDATE bom_pozycje SET ilosc=?, uwagi=? WHERE id=?",
            (b.ilosc, b.uwagi or "", bid)
        )
        return {"ok": True}

@app.delete("/api/bom/{bid}", dependencies=[Depends(verify_key)])
def delete_bom(bid: int):
    with get_db() as conn:
        conn.execute("DELETE FROM bom_pozycje WHERE id=?", (bid,))
        return {"ok": True}

# ─── Narzędziownia ────────────────────────────────────────────────────────────

@app.post("/api/narzedzia/import", dependencies=[Depends(verify_key)])
async def import_narzedzia_xlsx(file: UploadFile = File(...)):
    """Import bazy narzędzi z pliku xlsx.
    Wymagane kolumny: Indeks, Nazwa.
    Opcjonalne: Typ, Jm, Stan, Stan_min, Lokalizacja, Uwagi, Kod_paskowy.
    """
    content = await file.read()
    try:
        wb = _openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        raise HTTPException(400, f"Błąd odczytu xlsx: {e}")
    if not rows:
        raise HTTPException(400, "Plik xlsx jest pusty")

    # Mapowanie nagłówków (case-insensitive, bez polskich znaków)
    def _norm(s):
        import unicodedata
        s = str(s or "").strip().lower()
        return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

    header = [_norm(c) for c in rows[0]]
    def col(names):
        for n in names:
            if n in header: return header.index(n)
        return None

    i_indeks   = col(["indeks","index","kod","symbol"])
    i_nazwa    = col(["nazwa","name","opis","description","artykul"])
    i_typ      = col(["typ","type","kategoria","category"])
    i_jm       = col(["jm","jednostka","unit"])
    i_stan     = col(["stan","ilosc","qty","quantity","stan rzeczywisty","stan_rzeczywisty"])
    i_stan_min = col(["stan_min","min","minimum","prog","prog_min","ilosc_min"])
    i_lok      = col(["lokalizacja","location","miejsce","polka","shelf"])
    i_uwagi    = col(["uwagi","uwaga","notes","note","komentarz"])
    i_kod      = col(["kod paskowy","kod_paskowy","barcode","ean"])

    if i_indeks is None or i_nazwa is None:
        raise HTTPException(400, "Brak wymaganych kolumn: Indeks i Nazwa")

    imported = skipped = 0
    with get_db() as conn:
        for row in rows[1:]:
            def g(i): return str(row[i]).strip() if i is not None and i < len(row) and row[i] is not None else ""
            def gf(i, default=0):
                try: return float(str(row[i]).replace(",", ".")) if i is not None and i < len(row) and row[i] is not None else default
                except: return default
            def gi(i, default=0):
                try: return int(float(str(row[i]).replace(",", "."))) if i is not None and i < len(row) and row[i] is not None else default
                except: return default

            indeks = g(i_indeks)
            nazwa  = g(i_nazwa)
            if not indeks or not nazwa:
                skipped += 1
                continue
            try:
                conn.execute("""
                    INSERT INTO narzedzia (indeks, nazwa, typ, jm, stan, stan_min, lokalizacja, uwagi, kod_paskowy, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(indeks) DO UPDATE SET
                        nazwa=excluded.nazwa, typ=excluded.typ, jm=excluded.jm,
                        stan=excluded.stan, stan_min=excluded.stan_min,
                        lokalizacja=excluded.lokalizacja, uwagi=excluded.uwagi,
                        kod_paskowy=excluded.kod_paskowy, updated_at=excluded.updated_at
                """, (indeks, nazwa,
                      g(i_typ) or "ogólne",
                      g(i_jm) or "szt",
                      gi(i_stan),
                      gi(i_stan_min, 1),
                      g(i_lok),
                      g(i_uwagi),
                      g(i_kod),
                      _now()))
                imported += 1
            except Exception as e:
                print(f"Narzędzie skip {indeks}: {e}")
                skipped += 1
    return {"ok": True, "imported": imported, "skipped": skipped}


@app.get("/api/narzedzia", dependencies=[Depends(verify_key)])
def get_narzedzia(q: str = "", limit: int = 50):
    with get_db() as conn:
        if q:
            like = f"%{q}%"
            rows = conn.execute("""
                SELECT n.*, COALESCE(SUM(CASE WHEN r.status='aktywna' THEN r.ilosc ELSE 0 END),0) as zarezerwowane
                FROM narzedzia n
                LEFT JOIN narzedzia_rezerwacje r ON r.narzedzie_id=n.id
                WHERE n.indeks LIKE ? OR n.nazwa LIKE ? OR n.typ LIKE ?
                GROUP BY n.id ORDER BY n.nazwa LIMIT ?
            """, (like, like, like, limit)).fetchall()
        else:
            rows = conn.execute("""
                SELECT n.*, COALESCE(SUM(CASE WHEN r.status='aktywna' THEN r.ilosc ELSE 0 END),0) as zarezerwowane
                FROM narzedzia n
                LEFT JOIN narzedzia_rezerwacje r ON r.narzedzie_id=n.id
                GROUP BY n.id ORDER BY n.nazwa LIMIT ?
            """, (limit,)).fetchall()
        return [dict(r) for r in rows]


@app.get("/api/narzedzia/count", dependencies=[Depends(verify_key)])
def count_narzedzia():
    with get_db() as conn:
        n = conn.execute("SELECT COUNT(*) FROM narzedzia").fetchone()[0]
        niskie = conn.execute("""
            SELECT COUNT(*) FROM narzedzia n
            WHERE n.stan - COALESCE((
                SELECT SUM(r.ilosc) FROM narzedzia_rezerwacje r
                WHERE r.narzedzie_id=n.id AND r.status='aktywna'
            ),0) < n.stan_min
        """).fetchone()[0]
        return {"count": n, "niskie_stany": niskie}


@app.get("/api/narzedzia/niskie-stany", dependencies=[Depends(verify_key)])
def narzedzia_niskie_stany():
    """Narzędzia, których wolny stan < stan_min (do ostrzeżeń)."""
    with get_db() as conn:
        rows = conn.execute("""
            SELECT n.*,
                   COALESCE(SUM(CASE WHEN r.status='aktywna' THEN r.ilosc ELSE 0 END),0) as zarezerwowane
            FROM narzedzia n
            LEFT JOIN narzedzia_rezerwacje r ON r.narzedzie_id=n.id
            GROUP BY n.id
            HAVING (n.stan - zarezerwowane) < n.stan_min
            ORDER BY (n.stan - zarezerwowane) ASC
        """).fetchall()
        return [dict(r) for r in rows]


@app.put("/api/narzedzia/{nid}", dependencies=[Depends(verify_key)])
async def update_narzedzie(nid: int, request: Request):
    body = await request.json()
    fields = ["nazwa","typ","jm","stan","stan_min","lokalizacja","uwagi","kod_paskowy"]
    sets = ", ".join(f"{f}=?" for f in fields if f in body)
    vals = [body[f] for f in fields if f in body]
    if not sets:
        raise HTTPException(400, "Brak pól do aktualizacji")
    vals.append(_now())
    vals.append(nid)
    with get_db() as conn:
        conn.execute(f"UPDATE narzedzia SET {sets}, updated_at=? WHERE id=?", vals)
        return {"ok": True}


@app.delete("/api/narzedzia/{nid}", dependencies=[Depends(verify_key)])
def delete_narzedzie(nid: int):
    with get_db() as conn:
        conn.execute("DELETE FROM narzedzia WHERE id=?", (nid,))
        return {"ok": True}


# ─── Rezerwacje narzędzi ──────────────────────────────────────────────────────

class NarzedzieRezerwacjaIn(BaseModel):
    narzedzie_id: int
    zlecenie_nr: str
    cel: Optional[str] = "produkcja"
    ilosc: int = 1
    data_od: Optional[str] = None
    data_do: Optional[str] = None
    uwagi: Optional[str] = ""


@app.get("/api/narzedzia-rezerwacje", dependencies=[Depends(verify_key)])
def get_narzedzia_rezerwacje(status: str = "aktywna", narzedzie_id: Optional[int] = None):
    with get_db() as conn:
        if narzedzie_id:
            rows = conn.execute("""
                SELECT r.*, n.nazwa as narzedzie_nazwa, n.indeks as narzedzie_indeks, n.jm
                FROM narzedzia_rezerwacje r JOIN narzedzia n ON r.narzedzie_id=n.id
                WHERE r.narzedzie_id=? AND r.status=?
                ORDER BY r.created_at DESC
            """, (narzedzie_id, status)).fetchall()
        else:
            rows = conn.execute("""
                SELECT r.*, n.nazwa as narzedzie_nazwa, n.indeks as narzedzie_indeks, n.jm
                FROM narzedzia_rezerwacje r JOIN narzedzia n ON r.narzedzie_id=n.id
                WHERE r.status=?
                ORDER BY r.created_at DESC
            """, (status,)).fetchall()
        return [dict(r) for r in rows]


@app.post("/api/narzedzia-rezerwacje", dependencies=[Depends(verify_key)])
def add_narzedzie_rezerwacja(req: NarzedzieRezerwacjaIn):
    # Sprawdź dostępność
    with get_db() as conn:
        tool = conn.execute("SELECT * FROM narzedzia WHERE id=?", (req.narzedzie_id,)).fetchone()
        if not tool:
            raise HTTPException(404, "Narzędzie nie znalezione")
        zarezerwowane = conn.execute("""
            SELECT COALESCE(SUM(ilosc),0) FROM narzedzia_rezerwacje
            WHERE narzedzie_id=? AND status='aktywna'
        """, (req.narzedzie_id,)).fetchone()[0]
        wolne = tool["stan"] - zarezerwowane
        if req.ilosc > wolne:
            raise HTTPException(409, f"Niewystarczający stan. Wolne: {wolne} {tool['jm']}")
        conn.execute("""
            INSERT INTO narzedzia_rezerwacje
                (narzedzie_id, zlecenie_nr, cel, ilosc, data_od, data_do, uwagi, status, created_at)
            VALUES (?,?,?,?,?,?,?,'aktywna',?)
        """, (req.narzedzie_id, req.zlecenie_nr, req.cel, req.ilosc,
              req.data_od, req.data_do, req.uwagi or "", _now()))
        return {"ok": True}


@app.patch("/api/narzedzia-rezerwacje/{rid}/zwolnij", dependencies=[Depends(verify_key)])
def zwolnij_narzedzie_rezerwacje(rid: int):
    with get_db() as conn:
        conn.execute("UPDATE narzedzia_rezerwacje SET status='zwolniona' WHERE id=?", (rid,))
        return {"ok": True}


@app.patch("/api/narzedzia-rezerwacje/{rid}/wydaj", dependencies=[Depends(verify_key)])
def wydaj_narzedzie(rid: int):
    """Oznacza narzędzie jako wydane i zmniejsza stan w bazie."""
    with get_db() as conn:
        rez = conn.execute("SELECT * FROM narzedzia_rezerwacje WHERE id=?", (rid,)).fetchone()
        if not rez:
            raise HTTPException(404, "Rezerwacja nie znaleziona")
        conn.execute("UPDATE narzedzia_rezerwacje SET status='wydane' WHERE id=?", (rid,))
        conn.execute("UPDATE narzedzia SET stan = MAX(0, stan - ?), updated_at=? WHERE id=?",
                     (rez["ilosc"], _now(), rez["narzedzie_id"]))
        return {"ok": True}


@app.patch("/api/narzedzia-rezerwacje/{rid}/zwrot", dependencies=[Depends(verify_key)])
def zwrot_narzedzia(rid: int):
    """Przyjęcie zwrotu – zwiększa stan i zamyka rezerwację."""
    with get_db() as conn:
        rez = conn.execute("SELECT * FROM narzedzia_rezerwacje WHERE id=?", (rid,)).fetchone()
        if not rez:
            raise HTTPException(404, "Rezerwacja nie znaleziona")
        conn.execute("UPDATE narzedzia_rezerwacje SET status='zwrocone' WHERE id=?", (rid,))
        conn.execute("UPDATE narzedzia SET stan = stan + ?, updated_at=? WHERE id=?",
                     (rez["ilosc"], _now(), rez["narzedzie_id"]))
        return {"ok": True}


@app.delete("/api/narzedzia-rezerwacje/{rid}", dependencies=[Depends(verify_key)])
def delete_narzedzie_rezerwacja(rid: int):
    with get_db() as conn:
        conn.execute("DELETE FROM narzedzia_rezerwacje WHERE id=?", (rid,))
        return {"ok": True}


# ─── Narzędziownia: dodawanie pojedynczego narzędzia ──────────────────────────

class NarzedzieIn(BaseModel):
    indeks: str
    nazwa: str
    typ: Optional[str] = "Inne"
    jm: Optional[str] = "szt"
    stan: Optional[float] = 0
    stan_min: Optional[float] = 1
    lokalizacja: Optional[str] = ""
    uwagi: Optional[str] = ""

@app.post("/api/narzedzia", dependencies=[Depends(verify_key)])
def add_narzedzie(req: NarzedzieIn):
    with get_db() as conn:
        exists = conn.execute("SELECT id FROM narzedzia WHERE indeks=?", (req.indeks,)).fetchone()
        if exists:
            raise HTTPException(409, f"Indeks '{req.indeks}' już istnieje")
        conn.execute("""
            INSERT INTO narzedzia (indeks, nazwa, typ, jm, stan, stan_min, lokalizacja, uwagi, updated_at)
            VALUES (?,?,?,?,?,?,?,?,?)
        """, (req.indeks, req.nazwa, req.typ or "Inne", req.jm or "szt",
              req.stan or 0, req.stan_min or 1, req.lokalizacja or "", req.uwagi or "", _now()))
    return {"ok": True}


# ─── Narzędziownia: historia pobrań (zastępuje rezerwacje dla materiałów eksploatacyjnych) ──

@app.get("/api/narzedzia/niskie-stany", dependencies=[Depends(verify_key)])
def narzedzia_niskie_stany_v2():
    """Narzędzia, których stan < stan_min."""
    with get_db() as conn:
        rows = conn.execute("""
            SELECT * FROM narzedzia
            WHERE stan < stan_min
            ORDER BY (stan * 1.0 / NULLIF(stan_min,0)) ASC
        """).fetchall()
        return [dict(r) for r in rows]

class NarzedzePobranieIn(BaseModel):
    narzedzie_id: int
    zlecenie_nr: Optional[str] = "—"
    ilosc: float = 1
    uwagi: Optional[str] = ""

@app.post("/api/narzedzia-pobrania", dependencies=[Depends(verify_key)])
def pobierz_narzedzie(req: NarzedzePobranieIn):
    with get_db() as conn:
        tool = conn.execute("SELECT * FROM narzedzia WHERE id=?", (req.narzedzie_id,)).fetchone()
        if not tool:
            raise HTTPException(404, "Narzędzie nie znalezione")
        if req.ilosc > tool["stan"]:
            raise HTTPException(409, f"Niewystarczający stan. Dostępne: {tool['stan']} {tool['jm']}")
        # Zmniejsz stan
        conn.execute("UPDATE narzedzia SET stan = MAX(0, stan - ?), updated_at=? WHERE id=?",
                     (req.ilosc, _now(), req.narzedzie_id))
        # Zapisz historię
        conn.execute("""
            INSERT INTO narzedzia_pobrania (narzedzie_id, zlecenie_nr, ilosc, uwagi, status, created_at)
            VALUES (?,?,?,?,'wydane',?)
        """, (req.narzedzie_id, req.zlecenie_nr or "—", req.ilosc, req.uwagi or "", _now()))
    return {"ok": True}

@app.get("/api/narzedzia-pobrania", dependencies=[Depends(verify_key)])
def get_narzedzia_pobrania(limit: int = 100):
    with get_db() as conn:
        rows = conn.execute("""
            SELECT p.*, n.nazwa as narzedzie_nazwa, n.indeks as narzedzie_indeks, n.jm
            FROM narzedzia_pobrania p
            JOIN narzedzia n ON p.narzedzie_id = n.id
            ORDER BY p.created_at DESC LIMIT ?
        """, (limit,)).fetchall()
        return [dict(r) for r in rows]

@app.patch("/api/narzedzia-pobrania/{pid}/zwrot", dependencies=[Depends(verify_key)])
def zwrot_narzedzia_v2(pid: int):
    """Przyjęcie zwrotu – zwiększa stan z powrotem."""
    with get_db() as conn:
        p = conn.execute("SELECT * FROM narzedzia_pobrania WHERE id=?", (pid,)).fetchone()
        if not p:
            raise HTTPException(404, "Pobranie nie znalezione")
        if p["status"] == "zwrocone":
            raise HTTPException(409, "Już zwrócone")
        conn.execute("UPDATE narzedzia_pobrania SET status='zwrocone' WHERE id=?", (pid,))
        conn.execute("UPDATE narzedzia SET stan = stan + ?, updated_at=? WHERE id=?",
                     (p["ilosc"], _now(), p["narzedzie_id"]))
    return {"ok": True}


class FeedbackIn(BaseModel):
    user_id: Optional[int] = None
    user_name: Optional[str] = None
    ocena: int
    wiadomosc: Optional[str] = ""

@app.post("/api/feedback", dependencies=[Depends(verify_key)])
def post_feedback(fb: FeedbackIn):
    if not (1 <= fb.ocena <= 5):
        raise HTTPException(400, "Ocena musi być od 1 do 5")
    with get_db() as conn:
        conn.execute(
            "INSERT INTO app_feedback (user_id, user_name, ocena, wiadomosc, created_at) VALUES (?,?,?,?,?)",
            (fb.user_id, fb.user_name, fb.ocena, fb.wiadomosc or "", _now())
        )
    return {"ok": True}

@app.get("/api/feedback", dependencies=[Depends(verify_key)])
def get_feedback():
    with get_db() as conn:
        rows = conn.execute(
            "SELECT id, user_id, user_name, ocena, wiadomosc, created_at FROM app_feedback ORDER BY created_at DESC"
        ).fetchall()
        return [dict(r) for r in rows]

# ─── Upload pliku STEP do Cloudinary ──────────────────────────────────────────
@app.post("/api/step-upload", dependencies=[Depends(verify_key)])
async def step_upload(request: Request):
    import hashlib as _hl, hmac as _hmac, base64 as _b64, time as _time
    if not (CLOUDINARY_CLOUD and CLOUDINARY_KEY and CLOUDINARY_SECRET):
        raise HTTPException(503, "Cloudinary nie skonfigurowany")
    body = await request.body()
    if not body:
        raise HTTPException(400, "Brak danych pliku")
    if len(body) > 100 * 1024 * 1024:
        raise HTTPException(413, "Plik za duży (maks. 100 MB)")

    ts = str(int(_time.time()))
    public_id = "produkcja_step/step_" + _hl.md5(body[:1024]).hexdigest()[:12]

    sign_params = f"public_id={public_id}&timestamp={ts}"
    sig = _hl.sha1(f"{sign_params}{CLOUDINARY_SECRET}".encode()).hexdigest()

    boundary = "----CLD" + _hl.md5(ts.encode()).hexdigest()[:16]
    def _field(name, value):
        return f"--{boundary}\r\nContent-Disposition: form-data; name=\"{name}\"\r\n\r\n{value}\r\n".encode()

    fname = request.headers.get("x-filename", "model.step")
    try:
        import urllib.parse as _up
        fname = _up.unquote(fname)
    except Exception:
        pass

    multipart = (
        _field("api_key", CLOUDINARY_KEY) +
        _field("timestamp", ts) +
        _field("signature", sig) +
        _field("public_id", public_id) +
        f"--{boundary}\r\nContent-Disposition: form-data; name=\"file\"; filename=\"{fname}\"\r\nContent-Type: application/octet-stream\r\n\r\n".encode() +
        body + f"\r\n--{boundary}--\r\n".encode()
    )

    try:
        req = urllib.request.Request(
            f"https://api.cloudinary.com/v1_1/{CLOUDINARY_CLOUD}/raw/upload",
            data=multipart,
            method="POST",
            headers={"Content-Type": f"multipart/form-data; boundary={boundary}"}
        )
        with urllib.request.urlopen(req, timeout=60) as r:
            resp = json.loads(r.read())
        url = resp.get("secure_url", "")
        if not url:
            raise HTTPException(500, "Cloudinary nie zwrócił URL")
        return {"ok": True, "url": url, "public_id": resp.get("public_id"), "bytes": resp.get("bytes")}
    except urllib.error.HTTPError as e:
        err = e.read().decode()
        raise HTTPException(502, f"Cloudinary error: {err}")
    except Exception as e:
        raise HTTPException(502, f"Upload error: {e}")

# ─── Proxy pobierania pliku STEP ──────────────────────────────────────────────
@app.get("/api/step-proxy")
async def step_proxy(url: str):
    import re as _re
    gdrive = _re.search(r'drive.google.com/file/d/([^/?]+)', url)
    if gdrive:
        file_id = gdrive.group(1)
        url = f"https://drive.google.com/uc?export=download&id={file_id}"
    elif 'onedrive.live.com' in url or '1drv.ms' in url:
        sep = '&' if '?' in url else '?'
        url = url + sep + 'download=1'
    try:
        req = urllib.request.Request(url, headers={
            "User-Agent": "Mozilla/5.0",
            "Accept": "*/*",
            "Cache-Control": "no-cache"
        })
        response = urllib.request.urlopen(req, timeout=120)
        ct = response.headers.get("Content-Type", "application/octet-stream")
        if "text/html" in ct:
            response.close()
            raise HTTPException(502, "Plik niedostępny – serwer zwrócił stronę HTML")

        def _stream():
            try:
                while True:
                    chunk = response.read(65536)
                    if not chunk: break
                    yield chunk
            finally:
                response.close()

        return StreamingResponse(
            _stream(),
            media_type="application/octet-stream",
            headers={
                "Access-Control-Allow-Origin": "*",
                "Content-Disposition": "inline; filename=model.step",
                "Cache-Control": "public, max-age=3600",
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(502, f"Nie można pobrać pliku STEP: {e}")

# ─── Serwowanie aplikacji ──────────────────────────────────────────────────────
STATIC_DIR = os.path.join(os.path.dirname(__file__), "static")
os.makedirs(STATIC_DIR, exist_ok=True)

BUILD_VERSION = os.environ.get("BUILD_VERSION", _dt.datetime.utcnow().strftime("%Y%m%d%H%M%S"))

@app.get("/app", response_class=HTMLResponse)
@app.get("/app/", response_class=HTMLResponse)
def serve_app():
    html_path = os.path.join(STATIC_DIR, "index.html")
    if not os.path.exists(html_path):
        raise HTTPException(404, "Brak pliku index.html")
    with open(html_path, "r", encoding="utf-8") as f:
        content = f.read()
    return HTMLResponse(
        content=content,
        headers={
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0",
            "ETag": f'"{BUILD_VERSION}"',
        }
    )

@app.get("/", response_class=HTMLResponse)
def root():
    return """
    <html><body style="font-family:monospace;background:#1a1f2e;color:#e8eaf0;padding:40px">
    <h2>⚙ Produkcja API v4.1</h2>
    <p>Status: <span style="color:#27ae60">● online</span></p>
    <p><a href="/docs" style="color:#e8a020">/docs</a> – dokumentacja API</p>
    <p><a href="/app" style="color:#e8a020">/app</a> – aplikacja mobilna</p>
    </body></html>
    """

@app.get("/health")
def health():
    return {"status": "ok", "db": os.path.exists(DB_PATH)}

# ─── Start ─────────────────────────────────────────────────────────────────────
init_db_on_start()

try:
    _conn_chk = sqlite3.connect(DB_PATH, timeout=5)
    _row_count = _conn_chk.execute("SELECT COUNT(*) FROM zlecenia").fetchone()[0]
    _conn_chk.close()
    if _row_count == 0:
        print("⚠ Baza pusta – próbuję przywrócić dane...")
        _restored = False
        if GIST_TOKEN:
            _gist_data = _gist_load()
            if _gist_data:
                init_db_on_start()
                _restored = _db_restore_from_dict(_gist_data)
                if _restored:
                    try:
                        with open(BACKUP_PATH, "w", encoding="utf-8") as _f:
                            json.dump(_gist_data, _f, ensure_ascii=False, default=str)
                    except Exception:
                        pass
                    print("✓ Dane przywrócone z GitHub Gist")
        if not _restored and os.path.exists(BACKUP_PATH):
            _restored = _db_restore_from_json()
            if _restored:
                print("✓ Dane przywrócone z lokalnego backupu")
        if not _restored:
            print("✗ Brak backupu – serwer startuje z pustą bazą")
    else:
        print(f"✓ Baza zawiera {_row_count} zleceń – backup nie jest potrzebny")
except Exception as _e:
    print(f"⚠ Nie sprawdzono stanu bazy: {_e}")

_bk_thread = _threading.Thread(target=_auto_backup_loop, daemon=True, name="auto-backup")
_bk_thread.start()
print(f"✓ Auto-backup uruchomiony co {BACKUP_INTERVAL}s → {BACKUP_PATH}")


# ═══════════════════════════════════════════════════════════════════════════════
# FAKTUROWANIE – Kontrahenci + Faktury
# ═══════════════════════════════════════════════════════════════════════════════

class KontrahentIn(BaseModel):
    nazwa: str
    nip: Optional[str] = ''
    adres: Optional[str] = ''
    kod_pocztowy: Optional[str] = ''
    miasto: Optional[str] = ''
    kraj: Optional[str] = 'PL'
    email: Optional[str] = ''
    telefon: Optional[str] = ''
    uwagi: Optional[str] = ''

class PozycjaFakturyIn(BaseModel):
    lp: Optional[int] = 1
    nazwa: str
    jm: Optional[str] = 'szt'
    ilosc: float = 1
    cena_netto: float = 0
    vat_procent: float = 23

class FakturaIn(BaseModel):
    kontrahent_id: Optional[int] = None
    zlecenie_id: Optional[int] = None
    data_wystawienia: str
    data_sprzedazy: Optional[str] = ''
    termin_platnosci: Optional[str] = ''
    forma_platnosci: Optional[str] = 'przelew'
    uwagi: Optional[str] = ''
    waluta: Optional[str] = 'PLN'
    created_by: Optional[int] = None
    pozycje: List[PozycjaFakturyIn] = []

def _next_numer_faktury(conn) -> str:
    from datetime import datetime
    year = datetime.now().year
    last = conn.execute(
        "SELECT numer FROM faktury WHERE numer LIKE ? ORDER BY id DESC LIMIT 1",
        (f"F/{year}/%",)
    ).fetchone()
    if last:
        try: seq = int(last["numer"].split("/")[-1]) + 1
        except: seq = 1
    else:
        seq = 1
    return f"F/{year}/{seq:04d}"

def _calc_faktura_totals(pozycje: List[PozycjaFakturyIn]):
    total_netto = sum(p.ilosc * p.cena_netto for p in pozycje)
    total_vat   = sum(p.ilosc * p.cena_netto * p.vat_procent / 100 for p in pozycje)
    return round(total_netto, 2), round(total_vat, 2), round(total_netto + total_vat, 2)

# ─── Kontrahenci ──────────────────────────────────────────────────────────────
@app.get("/api/kontrahenci", dependencies=[Depends(verify_key)])
def get_kontrahenci():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM kontrahenci ORDER BY nazwa COLLATE NOCASE").fetchall()
        return [dict(r) for r in rows]

@app.post("/api/kontrahenci", dependencies=[Depends(verify_key)])
def add_kontrahent(k: KontrahentIn):
    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO kontrahenci (nazwa,nip,adres,kod_pocztowy,miasto,kraj,email,telefon,uwagi) VALUES (?,?,?,?,?,?,?,?,?)",
            (k.nazwa, k.nip, k.adres, k.kod_pocztowy, k.miasto, k.kraj, k.email, k.telefon, k.uwagi)
        )
        return {"id": cur.lastrowid, "ok": True}

@app.patch("/api/kontrahenci/{kid}", dependencies=[Depends(verify_key)])
def update_kontrahent(kid: int, k: KontrahentIn):
    with get_db() as conn:
        conn.execute(
            "UPDATE kontrahenci SET nazwa=?,nip=?,adres=?,kod_pocztowy=?,miasto=?,kraj=?,email=?,telefon=?,uwagi=? WHERE id=?",
            (k.nazwa, k.nip, k.adres, k.kod_pocztowy, k.miasto, k.kraj, k.email, k.telefon, k.uwagi, kid)
        )
        return {"ok": True}

@app.delete("/api/kontrahenci/{kid}", dependencies=[Depends(verify_key)])
def delete_kontrahent(kid: int):
    with get_db() as conn:
        used = conn.execute("SELECT COUNT(*) as c FROM faktury WHERE kontrahent_id=?", (kid,)).fetchone()["c"]
        if used > 0:
            raise HTTPException(400, f"Kontrahent jest powiązany z {used} fakturami")
        conn.execute("DELETE FROM kontrahenci WHERE id=?", (kid,))
        return {"ok": True}

# ─── Faktury ──────────────────────────────────────────────────────────────────
@app.get("/api/faktury", dependencies=[Depends(verify_key)])
def get_faktury(status: Optional[str] = None, rok: Optional[int] = None):
    with get_db() as conn:
        q = """SELECT f.*, k.nazwa as kontrahent_nazwa, k.nip as kontrahent_nip,
                      z.numer as zlecenie_numer, z.nazwa as zlecenie_nazwa
               FROM faktury f
               LEFT JOIN kontrahenci k ON f.kontrahent_id = k.id
               LEFT JOIN zlecenia z ON f.zlecenie_id = z.id
               WHERE 1=1"""
        params = []
        if status and status != 'wszystkie':
            q += " AND f.status=?"; params.append(status)
        if rok:
            q += " AND f.numer LIKE ?"; params.append(f"F/{rok}/%")
        q += " ORDER BY f.id DESC"
        return [dict(r) for r in conn.execute(q, params).fetchall()]

@app.get("/api/faktury/{fid}", dependencies=[Depends(verify_key)])
def get_faktura(fid: int):
    with get_db() as conn:
        f = conn.execute("""
            SELECT f.*, k.nazwa as kontrahent_nazwa, k.nip as kontrahent_nip,
                   k.adres as kontrahent_adres, k.kod_pocztowy as kontrahent_kp,
                   k.miasto as kontrahent_miasto, k.kraj as kontrahent_kraj,
                   k.email as kontrahent_email, k.telefon as kontrahent_telefon,
                   z.numer as zlecenie_numer, z.nazwa as zlecenie_nazwa,
                   u.full_name as wystawil
            FROM faktury f
            LEFT JOIN kontrahenci k ON f.kontrahent_id = k.id
            LEFT JOIN zlecenia z ON f.zlecenie_id = z.id
            LEFT JOIN users u ON f.created_by = u.id
            WHERE f.id=?""", (fid,)).fetchone()
        if not f:
            raise HTTPException(404, "Faktura nie znaleziona")
        poz = conn.execute(
            "SELECT * FROM pozycje_faktury WHERE faktura_id=? ORDER BY lp, id", (fid,)
        ).fetchall()
        return {"faktura": dict(f), "pozycje": [dict(p) for p in poz]}

@app.post("/api/faktury", dependencies=[Depends(verify_key)])
def create_faktura(f: FakturaIn):
    with get_db() as conn:
        numer = _next_numer_faktury(conn)
        tn, tv, tb = _calc_faktura_totals(f.pozycje)
        cur = conn.execute(
            """INSERT INTO faktury (numer,kontrahent_id,zlecenie_id,data_wystawienia,
               data_sprzedazy,termin_platnosci,forma_platnosci,uwagi,waluta,
               total_netto,total_vat,total_brutto,created_by)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (numer, f.kontrahent_id, f.zlecenie_id, f.data_wystawienia,
             f.data_sprzedazy, f.termin_platnosci, f.forma_platnosci,
             f.uwagi, f.waluta, tn, tv, tb, f.created_by)
        )
        fid = cur.lastrowid
        for i, p in enumerate(f.pozycje, 1):
            wn = round(p.ilosc * p.cena_netto, 2)
            wb = round(wn * (1 + p.vat_procent / 100), 2)
            conn.execute(
                """INSERT INTO pozycje_faktury (faktura_id,lp,nazwa,jm,ilosc,cena_netto,
                   vat_procent,wartosc_netto,wartosc_brutto) VALUES (?,?,?,?,?,?,?,?,?)""",
                (fid, i, p.nazwa, p.jm, p.ilosc, p.cena_netto, p.vat_procent, wn, wb)
            )
        return {"id": fid, "numer": numer}

@app.patch("/api/faktury/{fid}/status", dependencies=[Depends(verify_key)])
def update_faktura_status(fid: int, body: dict = Body(...)):
    status = body.get("status")
    if status not in ("szkic", "wystawiona", "oplacona", "anulowana"):
        raise HTTPException(400, "Nieprawidłowy status")
    with get_db() as conn:
        conn.execute("UPDATE faktury SET status=? WHERE id=?", (status, fid))
        return {"ok": True}

@app.delete("/api/faktury/{fid}", dependencies=[Depends(verify_key)])
def delete_faktura(fid: int):
    with get_db() as conn:
        f = conn.execute("SELECT status FROM faktury WHERE id=?", (fid,)).fetchone()
        if not f: raise HTTPException(404)
        if f["status"] not in ("szkic", "anulowana"):
            raise HTTPException(400, "Można usunąć tylko szkice i anulowane faktury")
        conn.execute("DELETE FROM pozycje_faktury WHERE faktura_id=?", (fid,))
        conn.execute("DELETE FROM faktury WHERE id=?", (fid,))
        return {"ok": True}

@app.get("/api/zlecenia/{zid}/faktura-template", dependencies=[Depends(verify_key)])
def faktura_template(zid: int):
    """Generuje pozycje faktury na podstawie danych zlecenia"""
    with get_db() as conn:
        zl = conn.execute("SELECT * FROM zlecenia WHERE id=?", (zid,)).fetchone()
        if not zl: raise HTTPException(404)
        ilosc = zl["ilosc_sztuk"] or 1
        cena_brutto = float(zl["cena_brutto_szt"] or 0)
        cena_netto = round(cena_brutto / 1.23, 4) if cena_brutto else 0
        pozycje = [{"nazwa": f"{zl['nazwa']} [{zl['numer']}]",
                    "jm": "szt", "ilosc": ilosc,
                    "cena_netto": round(cena_netto, 2), "vat_procent": 23}]
        # Dodaj koszty dodatkowe jeśli brak ceny
        prods = conn.execute(
            "SELECT * FROM produkty_zlecenia WHERE zlecenie_id=?", (zid,)
        ).fetchall()
        return {"zlecenie": dict(zl), "pozycje": pozycje, "produkty": [dict(p) for p in prods]}

@app.get("/api/faktury/export/all", dependencies=[Depends(verify_key)])
def export_all_faktury():
    """Export pełnych danych faktur do importu w systemie księgowym"""
    with get_db() as conn:
        faktury = conn.execute("""
            SELECT f.*, k.nazwa as k_nazwa, k.nip as k_nip, k.adres as k_adres,
                   k.kod_pocztowy as k_kp, k.miasto as k_miasto, k.kraj as k_kraj
            FROM faktury f LEFT JOIN kontrahenci k ON f.kontrahent_id=k.id
            ORDER BY f.data_wystawienia, f.id
        """).fetchall()
        result = []
        for fak in faktury:
            poz = conn.execute(
                "SELECT * FROM pozycje_faktury WHERE faktura_id=? ORDER BY lp, id", (fak["id"],)
            ).fetchall()
            d = dict(fak); d["pozycje"] = [dict(p) for p in poz]
            result.append(d)
        return result

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=PORT, reload=False)
